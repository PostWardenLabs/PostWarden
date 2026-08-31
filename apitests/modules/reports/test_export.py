"""Tests for `modules.reports.export` — one CSV and one XLSX check per
report, against the same `book` fixture `test_service.py`/`test_router.py`
already use. Each test computes the real `service.py` result first (the
same call the corresponding router route would make) and hands it to
the export function under test, so a drift between `service.py`'s
output shape and what `export.py` expects would fail here, not silently
in production.

CSV assertions parse the response body with `csv.reader` rather than
string-matching, so a formatting change in `csv.writer`'s own quoting
doesn't produce a spurious failure. XLSX assertions load the response
body back with `openpyxl.load_workbook` and read specific cells/
formulas by address — `export/test_xlsx.py` already covers the shared
style helpers in isolation, so these only check that each report wires
the right *data* into the right cells, not the styling again.
"""
import csv
import io
from decimal import Decimal

from openpyxl import load_workbook

from postwarden.modules.reports import export, service

from ...conftest import mk_account, mk_account_level, mk_entry, mk_line, mk_scenario


def csv_rows(resp) -> list[list[str]]:
    text = resp.body.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


def load_wb(resp):
    return load_workbook(io.BytesIO(resp.body))


# ---------------------------------------------------------------------------
# Trial balance
# ---------------------------------------------------------------------------


def test_trial_balance_csv_lists_every_account_with_its_balance(book, conn):
    result = service.trial_balance(conn, "ACTUAL", "2026-02-28", zeros=0, raw=1)
    rows = csv_rows(export.trial_balance_csv(result, "ACTUAL", "2026-02-28", raw=1))
    assert rows[0] == ["Code", "Account", "Path", "Debit", "Credit"]
    by_code = {r[0]: r for r in rows[1:]}
    assert by_code["1100"][3] == "2200.00"   # Checking's raw debit balance
    assert by_code["5100"][3] == "800.00"    # Rent
    assert by_code["4100"][4] == "2000.00"   # Salary, credit side
    assert by_code["3100"][4] == "1000.00"   # Opening Balance Equity


def test_trial_balance_csv_filename_includes_scenario_as_of_and_raw(book, conn):
    result = service.trial_balance(conn, "ACTUAL", "2026-02-28", zeros=0, raw=1)
    resp = export.trial_balance_csv(result, "ACTUAL", "2026-02-28", raw=1)
    assert resp.headers["content-disposition"] == \
        'attachment; filename="postwarden-trial-balance-ACTUAL_2026-02-28_raw.csv"'


def test_trial_balance_xlsx_grand_total_row_is_in_balance(book, conn):
    result = service.trial_balance(conn, "ACTUAL", "2026-02-28", zeros=0, raw=1)
    ws = load_wb(export.trial_balance_xlsx(result, "ACTUAL", "2026-02-28", raw=1)).active
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    last = values[-1]
    assert last[1] == "In balance"
    assert last[2] == 3000.0
    assert last[3] == 3000.0


# ---------------------------------------------------------------------------
# Balance sheet
# ---------------------------------------------------------------------------


def test_balance_sheet_csv_sections_and_totals(book, conn):
    result = service.balance_sheet(conn, "ACTUAL", "2026-02-28", raw=0, zeros=0)
    rows = csv_rows(export.balance_sheet_csv(result, "ACTUAL", "2026-02-28", raw=0))
    assert rows[0] == ["Section", "Code", "Account", "Path", "Amount"]
    total_assets_row = next(r for r in rows if r and r[0] == "Total assets")
    assert total_assets_row[4] == "2200.00"
    assert result["in_balance"] is True


def test_balance_sheet_xlsx_liabilities_and_equity_are_sign_flipped(book, conn):
    """CSV/XLSX both negate the stored (credit-normal) subtotal for
    Liabilities/Equity rows so the sheet reads in the conventional
    positive direction — ported verbatim from `_balance_sheet_rows`'
    own sign convention."""
    result = service.balance_sheet(conn, "ACTUAL", "2026-02-28", raw=0, zeros=0)
    ws = load_wb(export.balance_sheet_xlsx(result, "ACTUAL", "2026-02-28", raw=0)).active
    equity_row = next(row for row in ws.iter_rows(values_only=True) if row[0] == "3100")
    assert equity_row[2] == 1000.0  # stored subtotal is -1000 (credit-normal); flipped positive


def test_balance_sheet_csv_includes_retained_earnings_as_a_plain_equity_row(book, conn):
    # No more separate "earnings_lines" pass — the "Retained Earnings"
    # node rides through as part of `result["equity"]` and gets the same
    # sign flip every real Equity row gets, same reasoning the sign-flip
    # test above already covers for account 3100.
    result = service.balance_sheet(conn, "ACTUAL", "2026-02-28", raw=0, zeros=0)
    rows = csv_rows(export.balance_sheet_csv(result, "ACTUAL", "2026-02-28", raw=0))
    retained_row = next(r for r in rows if r and r[2] == "Retained Earnings")
    assert retained_row[0] == "Equity"
    assert retained_row[4] == "1200.00"  # 2000 salary - 800 rent, unclosed


def test_balance_sheet_csv_raw_omits_retained_earnings_entirely(book, conn):
    result = service.balance_sheet(conn, "ACTUAL", "2026-02-28", raw=1, zeros=0)
    rows = csv_rows(export.balance_sheet_csv(result, "ACTUAL", "2026-02-28", raw=1))
    assert not any(r and r[2] == "Retained Earnings" for r in rows)
    total_assets_row = next(r for r in rows if r and r[0] == "Total assets")
    total_le_row = next(r for r in rows if r and r[0] == "Total liabilities + equity")
    assert total_assets_row[4] != total_le_row[4]
    assert result["in_balance"] is False


# ---------------------------------------------------------------------------
# Income statement
# ---------------------------------------------------------------------------


def test_income_statement_csv_single_range_has_no_periods_column(book, conn):
    result = service.income_statement_rows(conn, "ACTUAL", "2026-02-01", "2026-02-28", "", 0, False)
    rows = csv_rows(export.income_statement_csv(result, "ACTUAL", "", "2026-02-01", "2026-02-28", ""))
    assert rows[0] == ["Section", "Code", "Account", "Path", "ACTUAL"]
    net_income_row = next(r for r in rows if r and r[2] == "Net income")
    assert net_income_row[4] == "1200.00"  # 2000 salary - 800 rent


def test_income_statement_csv_split_prefixes_every_column_with_its_period(book, conn):
    from postwarden.domain.periods import split_periods
    periods = split_periods("2026-01-01", "2026-02-28", "monthly")
    result = service.income_statement_matrix(conn, "ACTUAL", periods, "2026-01-01", "2026-02-28", "", 0, False)
    rows = csv_rows(export.income_statement_csv(result, "ACTUAL", "", "2026-01-01", "2026-02-28", "monthly"))
    assert rows[0][4:8] == ["2026-01 ACTUAL", "2026-02 ACTUAL", "Total ACTUAL", "Average ACTUAL"]


def test_income_statement_xlsx_net_income_is_a_live_formula(book, conn):
    result = service.income_statement_rows(conn, "ACTUAL", "2026-02-01", "2026-02-28", "", 0, False)
    ws = load_wb(export.income_statement_xlsx(result, "ACTUAL", "", "2026-02-01", "2026-02-28", "",
                                                pct_of_base=False)).active
    net_income_row = next(row for row in ws.iter_rows(values_only=True) if row[1] == "Net income")
    # One income root (Salary, row 5) minus one expense root (Rent) —
    # exercises xlsx_sum_formula's "=C{income}-C{expense}" shape.
    assert net_income_row[2].startswith("=C")


def test_income_statement_xlsx_with_compare_writes_variance_formula(book, conn):
    result = service.income_statement_rows(conn, "ACTUAL", "2026-02-01", "2026-02-28", "ACTUAL", 0, False)
    ws = load_wb(export.income_statement_xlsx(result, "ACTUAL", "ACTUAL", "2026-02-01", "2026-02-28", "",
                                                pct_of_base=False)).active
    salary_r = next(i for i, row in enumerate(ws.iter_rows(values_only=True), start=1) if row[1] == "Salary")
    # Variance (column D) = base - compare, referencing this row's own C/F cells only.
    assert ws.cell(row=salary_r, column=4).value == f"=C{salary_r}-F{salary_r}"


# ---------------------------------------------------------------------------
# Cash flow
# ---------------------------------------------------------------------------


def test_cash_flow_csv_ties_out(book, conn):
    result = service.cash_flow_rows(conn, "ACTUAL", "2026-01-01", "2026-02-28")
    rows = csv_rows(export.cash_flow_csv(result, "ACTUAL", "2026-01-01", "2026-02-28"))
    tie_out_row = next(r for r in rows if r and r[2] == "Tie-out check")
    assert tie_out_row[3] == "PASS"
    net_change_row = next(r for r in rows if r and r[2] == "Net change in cash")
    assert net_change_row[3] == "2200.00"


def test_cash_flow_xlsx_includes_the_equity_contra_ledger_adjustment(book, conn):
    """The book's own Opening Balance entry (Dr Checking / Cr Opening
    Balance Equity — an equity-typed contra leg) is exactly the "ledger
    adjustment" case `cash_flow_rows`' rule 1 always peels into its own
    section rather than blending it into inflows/outflows — see that
    function's own docstring."""
    result = service.cash_flow_rows(conn, "ACTUAL", "2026-01-01", "2026-02-28")
    assert len(result["ledger_adjustments"]) == 1
    ws = load_wb(export.cash_flow_xlsx(result, "ACTUAL", "2026-01-01", "2026-02-28")).active
    labels = [row[1] for row in ws.iter_rows(values_only=True)]
    assert "Ledger adjustments" in labels
    assert "Opening Balance Equity" in labels


def test_cash_flow_xlsx_omits_ledger_adjustments_section_when_empty(conn):
    # A book with no equity-contra activity at all — Rent paid straight
    # out of Checking, no opening-balance seeding.
    scenario = mk_scenario(conn, "ACTUAL")
    checking = mk_account(conn, "1100", "Checking", "asset", is_cashflow=True)
    rent = mk_account(conn, "5100", "Rent", "expense")
    e = mk_entry(conn, scenario["id"], "2026-02-05", "Rent payment")
    mk_line(conn, e, rent["id"], 800, 1)
    mk_line(conn, e, checking["id"], -800, 2)

    result = service.cash_flow_rows(conn, "ACTUAL", "2026-01-01", "2026-02-28")
    assert result["ledger_adjustments"] == []
    ws = load_wb(export.cash_flow_xlsx(result, "ACTUAL", "2026-01-01", "2026-02-28")).active
    labels = [row[1] for row in ws.iter_rows(values_only=True)]
    assert "Ledger adjustments" not in labels


# ---------------------------------------------------------------------------
# Variance
# ---------------------------------------------------------------------------


def test_variance_csv_native_depth(book, conn):
    result = service.compute_variance(conn, "ACTUAL", "", "", "2026-02-28")
    rows = csv_rows(export.variance_csv(result, "ACTUAL", "2026-02-28"))
    assert rows[0] == ["Code", "Account", "Path", "ACTUAL", "Variance", ""]


def test_variance_xlsx_rolled_up_gives_every_row_a_subtotal(book, conn):
    level = mk_account_level(conn, "Top level", depth=1)
    budget = mk_scenario(conn, "BUDGET", scenario_type="budget", base_level_id=level["id"])
    e = mk_entry(conn, budget["id"], "2026-02-01", "Budgeted opening")
    mk_line(conn, e, book["assets"]["id"], 600, 1)
    mk_line(conn, e, book["equity"]["id"], -600, 2)

    result = service.compute_variance(conn, "ACTUAL", "BUDGET", "", "2026-02-28")
    resp = export.variance_xlsx(result, "ACTUAL", "2026-02-28", pct_of_base=False)
    assert resp.headers["content-disposition"] == \
        'attachment; filename="postwarden-variance-ACTUAL-vs-BUDGET_2026-02-28.xlsx"'
    ws = load_wb(resp).active
    subtotal_rows = [row for row in ws.iter_rows(values_only=True) if row[1] and "subtotal" in row[1]]
    assert len(subtotal_rows) == len(result["grouped"])  # every populated section gets one, rolled up

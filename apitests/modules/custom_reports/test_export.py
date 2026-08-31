"""`modules/custom_reports/export.py` against export.py directly — row
content, total row, filenames, and the count-vs-money number format.
No database: exports take an already-resolved service result, so these
build one by hand (the same division `modules/reports/`'s own
`test_export.py` draws)."""
import io
from decimal import Decimal

from openpyxl import load_workbook

from postwarden.modules.custom_reports import export
from postwarden.modules.custom_reports.enums import Dimension, Metric

RESULT = {
    "rows": [
        {"key": "2026-01", "label": "2026-01", "value": Decimal("350.00")},
        {"key": "2026-02", "label": "2026-02", "value": Decimal("60.00")},
    ],
    "total": Decimal("410.00"),
    "row_count": 2,
}


def test_csv_rows_and_total():
    resp = export.custom_report_csv(RESULT, Metric.net_amount, Dimension.month,
                                    "ACTUAL", "", "")
    lines = resp.body.decode("utf-8-sig").splitlines()
    assert lines[0] == "Month,Net amount"
    assert lines[1] == "2026-01,350.00"
    assert lines[3] == "Total,410.00"


def test_filename_carries_config():
    resp = export.custom_report_csv(RESULT, Metric.debit_total, Dimension.account,
                                    "ACTUAL", "2026-01-01", "2026-02-28")
    assert 'filename="postwarden-custom-debit_total-by-account-ACTUAL_2026-01-01_2026-02-28.csv"' \
        in resp.headers["content-disposition"]
    # dimension=scenario drops the scenario from the name — the filter
    # was dropped from the query too (service.py's rule).
    resp = export.custom_report_csv(RESULT, Metric.net_amount, Dimension.scenario,
                                    "ACTUAL", "", "")
    assert 'filename="postwarden-custom-net_amount-by-scenario.csv"' \
        in resp.headers["content-disposition"]


def test_xlsx_rows_total_and_formats():
    resp = export.custom_report_xlsx(RESULT, Metric.net_amount, Dimension.month,
                                     "ACTUAL", "", "")
    ws = load_workbook(io.BytesIO(resp.body)).active
    assert ws["A4"].value == "Month"
    assert ws["B5"].value == Decimal("350.00")
    assert ws["A7"].value == "Total"
    assert ws["B7"].value == Decimal("410.00")
    assert ws["B5"].number_format == export.xlsx.XLSX_MONEY_FMT


def test_xlsx_entry_count_uses_count_format():
    result = {"rows": [{"key": "2026-01", "label": "2026-01", "value": 2}],
              "total": 4, "row_count": 1}
    resp = export.custom_report_xlsx(result, Metric.entry_count, Dimension.month,
                                     "ACTUAL", "", "")
    ws = load_workbook(io.BytesIO(resp.body)).active
    assert ws["B5"].number_format == export._COUNT_FMT


def test_labels_cover_every_enum_member():
    assert set(export.METRIC_LABELS) == set(Metric)
    assert set(export.DIMENSION_LABELS) == set(Dimension)

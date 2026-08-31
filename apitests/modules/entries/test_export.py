"""Unit tests for `modules.entries.export` — built directly against
hand-built row dicts (the same shape `repository.export_rows` returns),
no database needed: `test_repository.py`/`test_service.py` already
cover that the right rows come back; this file only checks that
`journal_csv`/`journal_xlsx` turn a given row list into the right
spreadsheet."""
import csv
import io
from decimal import Decimal

from openpyxl import load_workbook

from postwarden.modules.entries import export


def _leg(entry_id="ABC123", entry_date="2026-03-01", scenario_code="ACTUAL",
         description="Paycheck", reference=None, payee_name=None,
         account_code="1100", account_name="Checking", debit=None, credit=None, memo=None):
    return {"entry_id": entry_id, "entry_date": entry_date, "scenario_code": scenario_code,
            "description": description, "reference": reference, "payee_name": payee_name,
            "account_code": account_code, "account_name": account_name,
            "debit": debit, "credit": credit, "memo": memo}


def _two_leg_entry():
    return [
        _leg(debit=Decimal("500.00")),
        _leg(account_code="4100", account_name="Salary", credit=Decimal("500.00")),
    ]


def test_journal_csv_writes_one_row_per_leg():
    resp = export.journal_csv(_two_leg_entry())
    rows = list(csv.reader(io.StringIO(resp.body.decode("utf-8-sig"))))
    assert rows[0] == ["Entry #", "Date", "Scenario", "Description", "Reference",
                        "Payee", "Account code", "Account name", "Debit", "Credit", "Memo"]
    assert rows[1] == ["ABC123", "2026-03-01", "ACTUAL", "Paycheck", "", "",
                        "1100", "Checking", "500.00", "", ""]
    assert rows[2][8:10] == ["", "500.00"]


def test_journal_csv_filename_is_fixed():
    resp = export.journal_csv([])
    assert resp.headers["content-disposition"] == 'attachment; filename="postwarden-journal.csv"'


def test_journal_xlsx_merges_entry_level_columns_across_every_leg():
    ws = load_workbook(io.BytesIO(export.journal_xlsx(_two_leg_entry(), "", "", "").body)).active
    merged = {str(r) for r in ws.merged_cells.ranges}
    # Two data rows (5, 6) for the one entry's two legs, plus the title/
    # subtitle rows (1, 2) — Entry # (col A) merged down both leg rows.
    assert "A5:A6" in merged
    assert "D5:D6" in merged  # Description


def test_journal_xlsx_does_not_merge_a_single_leg_entry():
    ws = load_workbook(io.BytesIO(export.journal_xlsx([_leg(debit=Decimal("100.00"))], "", "", "").body)).active
    # Rows 1-2 (title/subtitle) always merge; no data row (5+) should.
    assert all(r.min_row < 5 for r in ws.merged_cells.ranges)


def test_journal_xlsx_grand_total_sums_debits_and_credits_when_balanced():
    ws = load_workbook(io.BytesIO(export.journal_xlsx(_two_leg_entry(), "", "", "").body)).active
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    last = values[-1]
    assert last[3] == "Total"
    assert last[8] == "=SUM(I5:I6)"
    assert last[9] == "=SUM(J5:J6)"


def test_journal_xlsx_flags_an_out_of_balance_filtered_result():
    # A single debit leg with no matching credit leg in the export — the
    # filtered result itself is out of balance, which the export flags
    # with a "this filter includes a scenario that allows single-sided
    # entries" label (single-sided entries are a real, if rare, legal
    # state — SPEC.md's own scenario.enforce_balance flag).
    ws = load_workbook(io.BytesIO(export.journal_xlsx([_leg(debit=Decimal("100.00"))], "", "", "").body)).active
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    assert "out of balance" in values[-1][3]


def test_journal_xlsx_with_no_rows_writes_a_zero_total():
    ws = load_workbook(io.BytesIO(export.journal_xlsx([], "", "", "").body)).active
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    last = values[-1]
    assert last[8] == 0
    assert last[9] == 0


def test_journal_xlsx_subtitle_reflects_scenario_and_date_range():
    ws = load_workbook(io.BytesIO(export.journal_xlsx([], "ACTUAL", "2026-01-01", "2026-02-28").body)).active
    assert ws.cell(row=2, column=1).value == "ACTUAL · 2026-01-01 to 2026-02-28"


def test_journal_xlsx_credit_leg_is_indented():
    ws = load_workbook(io.BytesIO(export.journal_xlsx(_two_leg_entry(), "", "", "").body)).active
    debit_row = ws.cell(row=5, column=10)  # Credit column, debit leg — blank, no indent to check
    credit_row_account = ws.cell(row=6, column=7)  # Account code column, credit leg
    assert credit_row_account.alignment.indent == 1

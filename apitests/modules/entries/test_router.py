"""End-to-end tests of modules.entries.router — real HTTP requests
through a throwaway FastAPI() + include_router(), the same pattern
modules/reports/test_router.py established, proving the whole chain
(request body/query params -> service -> repository -> real Postgres ->
JSON response) works together, including that Decimal values round-trip
as strings (Phase 1.3's json.py) and that a domain `ValueError` or a
deferred-trigger `SQLAlchemyError` both come back as a 400 with a plain
message, not a bare 500."""
from fastapi import FastAPI
from fastapi.testclient import TestClient

from postwarden.db import get_connection
from postwarden.modules.auth.deps import get_current_session, require_csrf_header
from postwarden.modules.entries.router import router

from ...conftest import mk_user


def client_for(conn) -> TestClient:
    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[get_connection] = lambda: conn
    # As of Phase 1.14, every route here requires a session
    # (`APIRouter(dependencies=[Depends(get_current_session)])`), and every
    # write route additionally requires `require_csrf_header` — override both
    # to a fixed fake session rather than simulate a real login/CSRF-token
    # round-trip in every test below. A *real* `users` row (`mk_user`), not a
    # made-up id: `create_entry`/`reverse_entry`/`reverse_entries_bulk` all
    # thread this session's `user_id` into `created_by_user_id`, which has a
    # real FK against `users(id)`.
    session = {"user_id": mk_user(conn)["id"], "username": "test"}
    app.dependency_overrides[get_current_session] = lambda: session
    app.dependency_overrides[require_csrf_header] = lambda: session
    return TestClient(app)


def _entry_body(book, **kw):
    body = {
        "scenario_id": book["scenario"]["id"], "description": "Paycheck",
        "lines": [{"account": "1100", "debit": "500", "credit": ""},
                  {"account": "4100", "debit": "", "credit": "500"}],
    }
    body.update(kw)
    return body


def test_create_entry_returns_201_and_the_new_entry_id(book, conn):
    resp = client_for(conn).post("/entries", json=_entry_body(book))
    assert resp.status_code == 201
    entry_id = resp.json()["entry_id"]
    assert len(entry_id) == 6


def test_create_entry_unbalanced_lines_returns_400_with_the_trigger_message(book, conn):
    body = _entry_body(book, lines=[{"account": "1100", "debit": "100", "credit": ""},
                                     {"account": "4100", "debit": "", "credit": "50"}])
    resp = client_for(conn).post("/entries", json=body)
    assert resp.status_code == 400
    assert "is not balanced" in resp.json()["detail"]


def test_create_entry_unknown_account_returns_400(book, conn):
    body = _entry_body(book, lines=[{"account": "9999", "debit": "10", "credit": ""},
                                     {"account": "4100", "debit": "", "credit": "10"}])
    resp = client_for(conn).post("/entries", json=body)
    assert resp.status_code == 400
    assert "Unknown account code" in resp.json()["detail"]


def test_list_entries_returns_decimal_totals_as_strings(book, conn):
    client = client_for(conn)
    client.post("/entries", json=_entry_body(book))
    resp = client.get("/entries")
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["entries"]) == 1
    assert body["entries"][0]["total_debits"] == "500.00"
    # lines come back ordered by line_no; line 1 is the Checking (debit) leg.
    assert body["entries"][0]["lines"][0]["debit"] == "500.00"


def test_export_csv_endpoint(book, conn):
    client = client_for(conn)
    client.post("/entries", json=_entry_body(book))
    resp = client.get("/entries/export.csv")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"
    assert "postwarden-journal.csv" in resp.headers["content-disposition"]
    text = resp.text
    assert "Entry #,Date,Scenario" in text
    assert "500" in text


def test_export_xlsx_endpoint_respects_the_same_filters_as_the_list(book, conn):
    client = client_for(conn)
    client.post("/entries", json=_entry_body(book))
    resp = client.get("/entries/export.xlsx", params={"account": "9999-no-such-account"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == \
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # No entry has a line on that account, so the workbook comes back
    # with a zero grand total rather than the one entry we just posted —
    # same build_filter() the paged list itself uses.
    from openpyxl import load_workbook
    import io
    ws = load_workbook(io.BytesIO(resp.content)).active
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    assert values[-1][8] == 0  # Debit total


def test_reverse_entry_endpoint(book, conn):
    client = client_for(conn)
    entry_id = client.post("/entries", json=_entry_body(book)).json()["entry_id"]
    resp = client.post(f"/entries/{entry_id}/reverse")
    assert resp.status_code == 200
    assert resp.json()["entry_id"] == entry_id
    assert len(resp.json()["reversed_by"]) == 6


def test_reverse_entry_not_found_returns_400(book, conn):
    resp = client_for(conn).post("/entries/ZZZZZZ/reverse")
    assert resp.status_code == 400
    assert "not found" in resp.json()["detail"]


def test_reverse_entries_bulk_endpoint_reports_partial_success(book, conn):
    client = client_for(conn)
    good = client.post("/entries", json=_entry_body(book)).json()["entry_id"]
    resp = client.post("/entries/reverse", json={"entry_ids": ["ZZZZZZ", good]})
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["reversed"]) == 1
    assert body["errors"] == ["Entry #ZZZZZZ not found"]


def test_reverse_entries_bulk_endpoint_rejects_an_empty_selection(book, conn):
    resp = client_for(conn).post("/entries/reverse", json={"entry_ids": []})
    assert resp.status_code == 400


def test_edit_entries_tags_endpoint(book, conn):
    client = client_for(conn)
    entry_id = client.post("/entries", json=_entry_body(book)).json()["entry_id"]
    resp = client.post("/entries/tags", json={"entry_ids": [entry_id], "action": "add", "tag": "urgent"})
    assert resp.status_code == 200
    assert resp.json() == {"tag": "urgent", "action": "add"}


def test_edit_description_endpoint(book, conn):
    client = client_for(conn)
    entry_id = client.post("/entries", json=_entry_body(book)).json()["entry_id"]
    resp = client.post(f"/entries/{entry_id}/edit-description", json={"description": "Updated"})
    assert resp.status_code == 200
    assert resp.json()["description"] == "Updated"


def test_edit_line_memo_endpoint(book, conn):
    client = client_for(conn)
    entry_id = client.post("/entries", json=_entry_body(book)).json()["entry_id"]
    line_id = client.get("/entries", params={"entry_id": entry_id}).json()["entries"][0]["lines"][0]["id"]
    resp = client.post(f"/entries/lines/{line_id}/edit-memo", json={"memo": "a note"})
    assert resp.status_code == 200
    assert resp.json()["memo"] == "a note"

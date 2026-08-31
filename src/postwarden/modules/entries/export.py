"""CSV/XLSX assembly for the Journal's export routes — ported from
`app/main.py`'s `entries_export_csv`/`entries_export_xlsx`, unchanged in
shape. Both take the flat leg-per-row list `service.export_rows` already
produces (CSV in each line's own posting order, XLSX with debits grouped
ahead of credits within an entry — see that function's own docstring)
and hand it to `export.csv`/`export.xlsx` (Phase 1.12) to write, the
same shared plumbing `modules/reports/export.py` draws from.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border
from fastapi import Response

from ...export import xlsx
from ...export.csv import csv_response

# Journal's own leg indent, used on both the account columns and the
# Credit amount for a credit leg — the classic printed-journal
# convention (debit side flush left, credit side stepped in underneath
# it) that a flat Debit/Credit column pair alone doesn't convey.
# Right-aligned explicitly (not left to Excel's default General
# alignment the way every other report's money cells are) because
# indent's edge is relative to whichever horizontal alignment is set —
# General resolves per-type at render time in a way that doesn't
# reliably indent a right-aligned number, so both the debit and credit
# amount columns below pin "right" outright to keep the indent's effect
# predictable. Local to this module, unlike the shared palette in
# `export.xlsx` — no other report needs a per-leg indent.
_JOURNAL_INDENT = Alignment(horizontal="right", indent=1)
_JOURNAL_ACCOUNT_INDENT = Alignment(indent=1)


def journal_csv(rows: list[dict]) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Entry #", "Date", "Scenario", "Description", "Reference",
                "Payee", "Account code", "Account name", "Debit", "Credit", "Memo"])
    for r in rows:
        w.writerow([r["entry_id"], r["entry_date"], r["scenario_code"],
                    r["description"], r["reference"] or "", r["payee_name"] or "",
                    r["account_code"], r["account_name"],
                    r["debit"] or "", r["credit"] or "", r["memo"] or ""])
    return csv_response(buf, "postwarden-journal.csv")


def journal_xlsx(rows: list[dict], scenario: str, date_from: str, date_to: str) -> Response:
    """XLSX counterpart to `journal_csv` above — same rows, same
    DESC-by-date order, but formatted the way a printed general journal
    traditionally reads rather than as a flat line-per-row dump: each
    entry's own legs grouped together with debits listed before credits
    (`service.export_rows(group_legs=True)`'s own ordering, not
    `line_no`'s original posting order — the standard journal-entry
    presentation), credit legs indented under the debit lines above
    them, and a rule only under an entry's *last* leg — never between
    two legs of the same entry — so a glance down the sheet reads "these
    lines are one transaction" the way the grid alone wouldn't. Every
    entry-level column (Entry #, Date, Scenario, Description, Reference,
    Payee — everything that describes the transaction rather than one
    leg of it) is merged and vertically centered down every leg, written
    once on the entry's first leg row rather than repeated —
    `merge_cells()` discards whatever's in a merged range's non-anchor
    cells on save regardless, so writing it again on every leg would
    just be work openpyxl throws away.

    Doesn't reuse `export.xlsx.xlsx_data_row`/`xlsx_header_row`'s
    tree-shaped row model (label_cols then value_cols, one flat font per
    row) — this report's per-leg indent only ever applies to the account
    and Credit-amount columns, and the entry-level columns are merged
    rather than repeated, neither of which fits that helper's
    "everything but the first label column" depth convention. Still
    built from the same shared palette (fonts, money format, grand-total
    border/coloring, title/subtitle style, no view gridlines) so it
    reads as one of this app's own exports rather than a one-off."""
    # Group the flat leg rows back into entries, preserving the SQL's
    # own order (a plain dict, not itertools.groupby, matching every
    # other "bucket rows under their parent id" spot in this codebase).
    legs_by_entry: dict[str, list] = {}
    for line in rows:
        legs_by_entry.setdefault(line["entry_id"], []).append(line)

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"

    subtitle = scenario or "All scenarios"
    if date_from and date_to:
        subtitle += f" · {date_from} to {date_to}"
    elif date_from:
        subtitle += f" · from {date_from}"
    elif date_to:
        subtitle += f" · through {date_to}"
    ws.cell(row=1, column=1, value="Journal").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    headers = ["Entry #", "Date", "Scenario", "Description", "Reference",
               "Payee", "Account code", "Account name", "Debit", "Credit", "Memo"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    xlsx.xlsx_header_row(ws, header_row, headers)

    # Entry-level columns (everything but the leg-level account/amount/
    # memo ones) — merged and centered down every leg of an entry, since
    # a value like "Salary — second half of August" or a payee name
    # describes the whole transaction, not any one leg of it. Written
    # once, on an entry's first leg row only — merge_cells() discards
    # whatever's in a merged range's non-anchor cells on save regardless,
    # so writing them again on every leg would just be work openpyxl
    # throws away.
    MERGED_ENTRY_COLS = (1, 2, 3, 4, 5, 6)  # Entry #, Date, Scenario, Description, Reference, Payee

    r = data_start
    total_debits = total_credits = 0
    for eid, legs in legs_by_entry.items():
        entry_first_row = r
        for i, line in enumerate(legs):
            is_credit = bool(line["credit"])
            if i == 0:
                ws.cell(row=r, column=1, value=eid).font = xlsx.XLSX_LINE_FONT
                # entry_date is a real date object (not a string) so the
                # column stays sortable/filterable in Excel — number_format
                # pins it to plain ISO text instead of Excel's own locale-
                # dependent default date display, matching the CSV export's
                # plain "YYYY-MM-DD" rendering of the same value.
                date_cell = ws.cell(row=r, column=2, value=line["entry_date"])
                date_cell.font = xlsx.XLSX_LINE_FONT
                date_cell.number_format = "yyyy-mm-dd"
                for col, value in ((3, line["scenario_code"]),
                                   (4, line["description"]), (5, line["reference"] or ""),
                                   (6, line["payee_name"] or "")):
                    ws.cell(row=r, column=col, value=value).font = xlsx.XLSX_LINE_FONT
            for col, value in ((7, line["account_code"]), (8, line["account_name"])):
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = xlsx.XLSX_LINE_FONT
                if is_credit:
                    cell.alignment = _JOURNAL_ACCOUNT_INDENT
            debit_cell = ws.cell(row=r, column=9, value=line["debit"] or None)
            debit_cell.font = xlsx.XLSX_LINE_FONT
            debit_cell.number_format = xlsx.XLSX_MONEY_FMT
            debit_cell.alignment = Alignment(horizontal="right")
            credit_cell = ws.cell(row=r, column=10, value=line["credit"] or None)
            credit_cell.font = xlsx.XLSX_LINE_FONT
            credit_cell.number_format = xlsx.XLSX_MONEY_FMT
            credit_cell.alignment = _JOURNAL_INDENT if is_credit else Alignment(horizontal="right")
            ws.cell(row=r, column=11, value=line["memo"] or "").font = xlsx.XLSX_LINE_FONT
            total_debits += line["debit"] or 0
            total_credits += line["credit"] or 0
            r += 1
        entry_last_row = r - 1
        for col in MERGED_ENTRY_COLS:
            if entry_last_row > entry_first_row:
                ws.merge_cells(start_row=entry_first_row, start_column=col,
                                end_row=entry_last_row, end_column=col)
            ws.cell(row=entry_first_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        # The one rule this report draws mid-sheet: under an entry's last
        # leg only, so it reads as "the next row starts a new
        # transaction" rather than a grid line between two legs of the
        # same one.
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=entry_last_row, column=col)
            cell.border = Border(left=cell.border.left, right=cell.border.right,
                                  top=cell.border.top, bottom=xlsx.XLSX_RULE)

    in_balance = total_debits == total_credits
    style = "grand" if in_balance else "grand_bad"
    label = "Total" if in_balance \
        else "Total (out of balance — this filter includes a scenario that allows single-sided entries)"
    ws.cell(row=r, column=4, value=label).font = xlsx.XLSX_ROW_FONTS[style]
    if rows:
        debit_cell = ws.cell(row=r, column=9, value=f"=SUM(I{data_start}:I{r - 1})")
        credit_cell = ws.cell(row=r, column=10, value=f"=SUM(J{data_start}:J{r - 1})")
    else:
        debit_cell = ws.cell(row=r, column=9, value=0)
        credit_cell = ws.cell(row=r, column=10, value=0)
    for cell in (debit_cell, credit_cell):
        cell.font = xlsx.XLSX_ROW_FONTS[style]
        cell.number_format = xlsx.XLSX_MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.border = xlsx.XLSX_GRAND_BORDER if in_balance else xlsx.XLSX_GRAND_BORDER_BAD

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 24
    ws.freeze_panes = f"B{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, "postwarden-journal.xlsx")

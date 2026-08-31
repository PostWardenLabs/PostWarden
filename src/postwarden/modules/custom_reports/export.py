"""CSV/XLSX writers for a custom report — the same shape every other
report module's `export.py` follows (take the already-resolved
`service.py` result plus the raw params that feed the filename/
subtitle, return a `Response`), just far simpler: a custom report is a
flat two-column table (label, value) plus a total row, whatever the
dimension — no sections, no hierarchy, no per-type subtotals.

The display-label maps below are the server's own copy for
filenames/headers/subtitles only — the frontend keeps its own map keyed
by the generated union types (see `enums.py`'s docstring), and the two
don't need to agree character-for-character, only to both cover every
member.
"""
import csv
import io

from fastapi import Response
from openpyxl import Workbook

from ...export import xlsx
from ...export.csv import csv_response
from .enums import Dimension, Metric

METRIC_LABELS: dict[Metric, str] = {
    Metric.net_amount: "Net amount",
    Metric.debit_total: "Debits",
    Metric.credit_total: "Credits",
    Metric.entry_count: "Entries",
}

DIMENSION_LABELS: dict[Dimension, str] = {
    Dimension.account: "Account",
    Dimension.account_level: "Account level",
    Dimension.tag: "Tag",
    Dimension.scenario: "Scenario",
    Dimension.month: "Month",
    Dimension.quarter: "Quarter",
    Dimension.year: "Year",
}

# Entries are a count, not money — no decimals, no paren-negative.
_COUNT_FMT = "#,##0"


def _filename(metric: Metric, dimension: Dimension, scenario: str,
              date_from: str, date_to: str, ext: str) -> str:
    name = f"postwarden-custom-{metric.value}-by-{dimension.value}"
    if scenario and dimension is not Dimension.scenario:
        name += f"-{scenario}"
    if date_from or date_to:
        name += f"_{date_from or 'start'}_{date_to or 'today'}"
    return f"{name}.{ext}"


def _subtitle(metric: Metric, dimension: Dimension, scenario: str,
              date_from: str, date_to: str) -> str:
    parts = [f"{METRIC_LABELS[metric]} by {DIMENSION_LABELS[dimension].lower()}"]
    if scenario and dimension is not Dimension.scenario:
        parts.append(scenario)
    if date_from or date_to:
        parts.append(f"{date_from or '…'} – {date_to or '…'}")
    return " · ".join(parts)


def custom_report_csv(result: dict, metric: Metric, dimension: Dimension,
                      scenario: str, date_from: str, date_to: str) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([DIMENSION_LABELS[dimension], METRIC_LABELS[metric]])
    for r in result["rows"]:
        w.writerow([r["label"], r["value"]])
    w.writerow(["Total", result["total"]])
    return csv_response(buf, _filename(metric, dimension, scenario, date_from, date_to, "csv"))


def custom_report_xlsx(result: dict, metric: Metric, dimension: Dimension,
                       scenario: str, date_from: str, date_to: str) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Custom Report"

    ws.cell(row=1, column=1, value="Custom Report").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1,
            value=_subtitle(metric, dimension, scenario, date_from, date_to)
            ).font = xlsx.XLSX_SUBTITLE_FONT

    xlsx.xlsx_header_row(ws, 4, [DIMENSION_LABELS[dimension], METRIC_LABELS[metric]])
    fmt = _COUNT_FMT if metric is Metric.entry_count else xlsx.XLSX_MONEY_FMT

    r = 5
    for row in result["rows"]:
        xlsx.xlsx_data_row(ws, r, [(1, row["label"])], [(2, row["value"], fmt)], style="line")
        r += 1
    xlsx.xlsx_data_row(ws, r, [(1, "Total")], [(2, result["total"], fmt)], style="grand")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    return xlsx.xlsx_response(wb, _filename(metric, dimension, scenario, date_from, date_to, "xlsx"))

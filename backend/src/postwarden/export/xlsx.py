"""XLSX export styling and writers — ported from `app/main.py`'s
module-level `_xlsx_*` helpers (the block sitting directly above the
Auth section), unchanged in behavior. Every name here drops the
original's leading underscore: those functions were private to one
5,908-line file; here the module boundary itself is the privacy
boundary, and `modules/reports/export.py`/`modules/entries/export.py`
are meant to import them directly.

Colors come from `style.css`'s default "Slate" theme (`--ink`/
`--paper-deep`) rather than whatever theme the browser has active — the
export is generated server-side with no idea which of the app's themes
Settings has picked, and a report someone opens in Excel a year later
shouldn't depend on that anyway. Kept as one small palette here rather
than a full port of `style.css`'s theme system: this only ever needs to
look like "a PostWarden document," not match the live page
pixel-for-pixel.
"""
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import Response
import io

XLSX_FONT = "Arial"
# Explicit "FF" alpha on every ARGB string below, not just the bare
# RGB — openpyxl silently zero-pads a 6-digit color to "00RRGGBB" (fully
# transparent alpha) rather than "FFRRGGBB" (opaque) if you don't. Excel
# itself ignores that byte for a solid fill and renders it opaque either
# way, but other readers (LibreOffice, Google Sheets) don't all make the
# same forgiving choice, so relying on Excel's leniency isn't worth it.
XLSX_HEADER_FILL = PatternFill("solid", fgColor="FF1B2430")   # --ink
XLSX_HEADER_FONT = Font(name=XLSX_FONT, size=10, bold=True, color="FFFFFFFF")
XLSX_GROUP_FONT = Font(name=XLSX_FONT, size=10, bold=True)
XLSX_LINE_FONT = Font(name=XLSX_FONT, size=10)
XLSX_RUNNING_FONT = Font(name=XLSX_FONT, size=10, italic=True)
XLSX_TITLE_FONT = Font(name=XLSX_FONT, size=14, bold=True, color="FF1B2430")
XLSX_SUBTITLE_FONT = Font(name=XLSX_FONT, size=9, italic=True, color="FF5B6B7C")  # --ink-soft
XLSX_LINE_FILL = PatternFill("solid", fgColor="FFEEF0F3")     # a shade off --paper-deep
XLSX_BOTTOM_BORDER = Border(bottom=Side(style="thin", color="FFAEBBC7"))  # --rule-strong
XLSX_RULE = Side(style="thin", color="FFAEBBC7")  # --rule-strong
XLSX_LINE_BORDER = Border(left=XLSX_RULE, right=XLSX_RULE, top=XLSX_RULE, bottom=XLSX_RULE)
# Split's period-group divider — heavier than the plain grid rule above,
# so the eye catches "new period starts here" scanning across a wide
# sheet the same way a ruled column break would in a printed ledger.
XLSX_PERIOD_DIVIDER = Side(style="medium", color="FF1B2430")  # --ink
# Same red/green the HTML report's own .neg (style.css) already uses for
# a negative figure — --red/--ok — so a variance reads the same way in
# the browser and in the spreadsheet. Real conditional-formatting rules
# (CellIsRule, in xlsx_variance_coloring below), not a color baked in at
# generation time, so the color still tracks correctly if a variance
# cell is edited by hand later — font-only (no fill/border), so it
# layers over whatever base style (line/group/running) that cell
# already has rather than replacing it.
XLSX_NEG_FONT = Font(color="FFB3392C")  # --red
XLSX_POS_FONT = Font(color="FF1F7A52")  # --ok
# The other three reports' own row styles, matching style.css's own
# tr.subtotal/tr.grand treatment exactly so a Trial Balance/Balance
# Sheet/Variance export reads the same way in a spreadsheet as it does
# on the page: "subtotal" (a rolled-up figure across more than one
# top-level account) is semi-bold and muted rather than fully bold, so
# it doesn't compete visually with a "group" row's own bold; "grand" is
# the accountant's double-rule under a total that balances — Excel's
# "double" border style — with a red variant for a report that doesn't
# (Trial Balance/Balance Sheet's own in_balance check, Cash Flow's
# tie-out).
XLSX_SUBTOTAL_FONT = Font(name=XLSX_FONT, size=10, bold=True, color="FF5B6B7C")  # --ink-soft
XLSX_GRAND_FONT = Font(name=XLSX_FONT, size=10, bold=True)
XLSX_GRAND_FONT_BAD = Font(name=XLSX_FONT, size=10, bold=True, color="FFB3392C")  # --red
XLSX_GRAND_BORDER = Border(bottom=Side(style="double", color="FF1B2430"))  # --ink
XLSX_GRAND_BORDER_BAD = Border(bottom=Side(style="double", color="FFB3392C"))  # --red
# Income Statement Split's own Total/Average column-group treatment —
# style.css's .period-agg/.period-agg-average, a fixed-value copy of the
# same color-mix() this palette targets everywhere else (the export has
# no idea which theme Settings has picked, so it targets the default
# Slate theme's own colors). Bold plus this tint on both, italic layered
# on top for Average only (xlsx_period_agg_font below), so Total and
# Average read as a different *kind* of column at a glance — not just
# one more period — while staying distinguishable from each other too.
XLSX_PERIOD_AGG_FILL = PatternFill("solid", fgColor="FFDCE2E8")  # 22% --rule-strong over --paper-deep
# No currency symbol — matches the app's own plain-text `money()` Jinja
# filter convention (display-only formatting is a rendering concern, and
# there's no renderer here). Parens for negatives, a bare dash for zero.
XLSX_MONEY_FMT = '#,##0.00;(#,##0.00);"-"'
# `pct_variance()` already returns the percentage figure itself (12.3
# meaning "12.3%"), not a 0-1 fraction, so this appends a literal "%"
# rather than using Excel's built-in 0.0% format, which would multiply
# the already-multiplied number by 100 again.
XLSX_PCT_FMT = '0.0"%";(0.0"%");"-"'


def xlsx_period_agg_font(base_font: Font, force_italic: bool = False) -> Font:
    """`base_font` with period-agg's bold layered on top — always
    forced, same as `.period-agg`'s own `font-weight:600` applies
    regardless of the row's own class — and italic forced on for
    Average specifically (`force_italic`), but never off: a
    running-total row (already italic via `XLSX_RUNNING_FONT`) stays
    italic in a Total column too, exactly as CSS would render it —
    `.period-agg` only ever *sets* `font-style`, it has no rule to
    override the row's own, so a running row's italic and a Total
    column's bold apply simultaneously rather than one replacing the
    other. Only name/size/color carry over unchanged; bold/italic are
    the two properties this treatment actually touches."""
    return Font(name=base_font.name, size=base_font.size, bold=True,
                italic=(base_font.italic or force_italic), color=base_font.color)


def xlsx_header_row(ws, row: int, headers: list[str], start_col: int = 1):
    for col, text in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = XLSX_HEADER_FONT
        cell.fill = XLSX_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def xlsx_merged_header(ws, r1: int, c1: int, r2: int, c2: int, text: str):
    """One header cell spanning r1:r2 x c1:c2, merged and centered —
    Income Statement Split's per-period date ("2026-01") sitting above
    that period's own ACTUAL/Variance/%/compare columns, or (c1==c2
    spanning both header rows) the Code/Account label sitting beside
    them. Only the anchor cell (top-left of the merge) needs the header
    styling — Excel and every reader that respects merges renders a
    merged range entirely from that cell, ignoring whatever the
    covered-but-hidden cells carry, so styling those individually would
    be dead work."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.font = XLSX_HEADER_FONT
    cell.fill = XLSX_HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def xlsx_thicken_right_border(ws, row: int, col: int):
    """Replace just the right edge of an already-styled cell with the
    heavier period-divider rule, keeping whatever left/top/bottom border
    that cell already carries. Works on a merged-but-not-anchor cell too
    (a period's header date spans several columns, but only its own
    left-most cell holds the merge) — openpyxl allows setting style on
    those, just not `.value`."""
    cell = ws.cell(row=row, column=col)
    b = cell.border
    cell.border = Border(left=b.left, right=XLSX_PERIOD_DIVIDER, top=b.top, bottom=b.bottom)


def xlsx_variance_coloring(ws, col: int, row_start: int, row_end: int):
    """Red text for a negative variance, green for a positive one, over
    one column's whole data range — applied to both the plain Variance
    column and the % Variance column, single-range or per-period alike.
    A zero (rendered as the money/pct format's own "-") gets neither."""
    cell_range = f"{get_column_letter(col)}{row_start}:{get_column_letter(col)}{row_end}"
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["0"], font=XLSX_NEG_FONT))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=["0"], font=XLSX_POS_FONT))


def xlsx_variance_formulas(base_cell: str, compare_cell: str, pct_of_base: bool) -> tuple[str, str]:
    """Live Excel formulas for a Variance/% Variance pair, replicating
    `domain.money.variance_amount`/`pct_variance`'s own two conventions
    exactly — default: base-minus-compare, % of compare (the standard
    percent-change reading, base as "new"); `pct_of_base` ("Flip
    variance direction" checked): compare-minus-base, % of base instead.
    Safe to derive live, unlike a group/subtotal row's own base/compare
    figures (which stay literals — see `modules/reports/export.py`'s own
    docstring on why): each formula only ever references the two cells
    already sitting in the same row, never a range that could
    double-count a rolled-up tree. `IF(...,"",...)` mirrors
    `pct_variance()` returning `None` (blank, not a literal 0%) when
    there's nothing to divide by."""
    if pct_of_base:
        return (f"={compare_cell}-{base_cell}",
                f'=IF({base_cell}=0,"",ROUND(({compare_cell}-{base_cell})/ABS({base_cell})*100,1))')
    return (f"={base_cell}-{compare_cell}",
            f'=IF({compare_cell}=0,"",ROUND(({base_cell}-{compare_cell})/ABS({compare_cell})*100,1))')


def xlsx_sum_formula(plus_cells: list[str], minus_cells: list[str] = ()) -> str:
    """A live formula adding/subtracting specific, individually-named
    cells — e.g. "=C6+C20-C34" — never a row range, so it stays safe
    regardless of how deep the tree under any one of those cells goes:
    each cell named here is a group's own root row, which already
    carries that subtree's full rolled-up total. Used for Income
    Statement's "Net income after X" running rows, each one just
    Income's root row(s) minus every expense group's root row seen so
    far. Falls back to a literal 0 rather than a bare "=" (not a valid
    formula) when both lists are empty — a report with no income and no
    expense rows at all, which shouldn't happen in practice but
    shouldn't crash either."""
    if not plus_cells and not minus_cells:
        return 0
    return "=" + "+".join(plus_cells) + "".join(f"-{c}" for c in minus_cells)


XLSX_ROW_FONTS = {
    "group": XLSX_GROUP_FONT, "line": XLSX_LINE_FONT, "running": XLSX_RUNNING_FONT,
    "subtotal": XLSX_SUBTOTAL_FONT, "grand": XLSX_GRAND_FONT, "grand_bad": XLSX_GRAND_FONT_BAD,
}


def xlsx_data_row(ws, row: int, label_cols: list, value_cols: list, style: str, depth: int = 0):
    """Write one report row. `label_cols` is `[(col, text), ...]` for the
    leading text columns (code, account name); `value_cols` is
    `[(col, value, number_format), ...]` for the money/percent columns.
    `style` picks one of six treatments (`XLSX_ROW_FONTS` above names the
    font for each):

    - "group": a section's own top-level account, or a bare
      section-title row when `value_cols` is empty (Trial Balance/
      Balance Sheet's own "Assets"/"Liabilities" headers) — bold, ruled
      underneath. For a real account row, that figure already *is* the
      section's total, since it's the root of the rolled-up tree.
    - "line": a plain account row — normal weight, its value cells
      shaded *and* fully gridded, matching the reference workbook's own
      "these are the numbers you'd read down a column" treatment.
    - "running": a running-total row like "Net income after Taxes" —
      italic, unshaded, unruled.
    - "subtotal": a rolled-up figure across more than one top-level
      account in the same section — semi-bold, muted, ruled, matching
      `style.css`'s `tr.subtotal` exactly.
    - "grand"/"grand_bad": the report's own bottom-line total — bold,
      with the accountant's double-rule under the value cells
      (`style.css`'s `tr.grand`), red instead of ink for "grand_bad"
      when that total doesn't actually balance/tie out.

    `depth` indents an account name under its parent, same meaning as
    the HTML report's own chevrons."""
    font = XLSX_ROW_FONTS[style]
    for col, text in label_cols:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = font
        if depth and col != label_cols[0][0]:
            cell.alignment = Alignment(indent=depth)
        if style in ("group", "subtotal"):
            cell.border = XLSX_BOTTOM_BORDER
    for col, value, number_format in value_cols:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.number_format = number_format
        if style in ("group", "subtotal"):
            cell.border = XLSX_BOTTOM_BORDER
        elif style == "line":
            cell.fill = XLSX_LINE_FILL
            cell.border = XLSX_LINE_BORDER
        elif style == "grand":
            cell.border = XLSX_GRAND_BORDER
        elif style == "grand_bad":
            cell.border = XLSX_GRAND_BORDER_BAD


def xlsx_response(wb: Workbook, filename: str) -> Response:
    """Wrap a finished Workbook as a download — the XLSX counterpart to
    `csv.csv_response()`. No BOM/codepage concerns here (XLSX is a real
    zip container, not a bare text stream), so this is just a stream and
    a content type."""
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(),
                     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     headers={"Content-Disposition": f'attachment; filename="{filename}"'})

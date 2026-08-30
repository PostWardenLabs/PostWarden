"""CSV/XLSX assembly for every report `modules/reports/router.py`
exports — ported from `app/main.py`'s six `/export/*` route pairs
(Trial Balance, Balance Sheet, Income Statement, Cash Flow, Variance —
Ledger's own export never existed in legacy either), unchanged in
shape. Each function here takes the *same* `service.py` result its
read-only sibling route already returns, plus whatever raw query params
feed the filename/subtitle, and returns a `fastapi.Response` — the
actual `csv.writer`/`openpyxl.Workbook` calls live here, `export.csv`/
`export.xlsx` (Phase 1.12) supply the shared plumbing (the BOM-prefixed
CSV wrapper, the XLSX style palette and row/formula helpers) every
report below draws from identically.

**Every account row's own base/compare figure in an XLSX export is a
literal, not a formula.** This is a rolled-up multi-root account tree
(`domain.accounts.build_account_tree`/`flatten_tree`), so a plain
`SUM()` over a *visible row range* would double-count wherever a group
is more than one level deep — what's written is exactly the same number
the JSON report and the CSV export already show for that row. Three
things layered on top of those literals *are* live formulas, each safe
for the identical reason — every cell referenced is named individually,
by row, never swept in as a range that could double-count: Income
Statement's Variance/% Variance pair and its "Net income after X"
running rows (`xlsx.xlsx_variance_formulas`/`xlsx_sum_formula`), and
Variance's own Variance/% Variance pair. Edit any base/compare figure by
hand later and every running total and variance downstream of it
recalculates instead of going stale.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi import Response

from ...export import xlsx
from ...export.csv import csv_response

# ---------------------------------------------------------------------------
# Trial balance
# ---------------------------------------------------------------------------


def _trial_balance_filename(scenario: str, as_of: str, raw: int, ext: str) -> str:
    name = f"postwarden-trial-balance-{scenario}"
    if as_of:
        name += f"_{as_of}"
    if raw:
        name += "_raw"
    return f"{name}.{ext}"


def trial_balance_csv(result: dict, scenario: str, as_of: str, raw: int) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Code", "Account", "Path", "Debit", "Credit"])
    for g in result["grouped"]:
        for r in g["rows"]:
            w.writerow([r["account_code"], r["account_name"], r["path"],
                        r["debit_balance"] or "", r["credit_balance"] or ""])
    return csv_response(buf, _trial_balance_filename(scenario, as_of, raw, "csv"))


def trial_balance_xlsx(result: dict, scenario: str, as_of: str, raw: int) -> Response:
    """XLSX counterpart to `trial_balance_csv` above, plus the
    section/subtotal/grand-total structure the CSV leaves out (it's a
    plain account list — no type-head, no per-type subtotal, no balance
    check) but `result["grouped"]` already carries: a bold section-title
    row per account type (`g["label"]`), that type's own subtotal row
    only when it actually sums more than one top-level account
    (`g["show_type_total"]` — a single-root type's own root row already
    *is* the total), and a bottom "In balance"/"Out of balance" row with
    the accountant's double-rule, red instead of ink when it doesn't."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    subtitle = f"{scenario} · {'As of ' + as_of if as_of else 'Through today'}"
    if not raw:
        subtitle += " · simulated monthly close"
    ws.cell(row=1, column=1, value="Trial Balance").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Debit", "Credit"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    xlsx.xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, debit, credit, style="line"):
        value_cols = [(3, debit or None, xlsx.XLSX_MONEY_FMT), (4, credit or None, xlsx.XLSX_MONEY_FMT)]
        xlsx.xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

    r = data_start
    for g in result["grouped"]:
        xlsx.xlsx_data_row(ws, r, [(1, ""), (2, g["label"])], [], style="group")
        r += 1
        for line in g["rows"]:
            row(r, line["account_code"], line["account_name"], line.get("depth", 2),
                line["debit_balance"], line["credit_balance"],
                style="group" if line.get("depth") == 1 else "line")
            r += 1
        if g["show_type_total"]:
            row(r, "", f"{g['label']} subtotal", 1, g["sub_debits"], g["sub_credits"], style="subtotal")
            r += 1
    grand_style = "grand" if result["in_balance"] else "grand_bad"
    label = "In balance" if result["in_balance"] else "Out of balance (this scenario allows single-sided entries)"
    row(r, "", label, 1, result["total_debits"], result["total_credits"], style=grand_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, _trial_balance_filename(scenario, as_of, raw, "xlsx"))


# ---------------------------------------------------------------------------
# Balance sheet
# ---------------------------------------------------------------------------


def _balance_sheet_filename(scenario: str, as_of: str, raw: int, ext: str) -> str:
    name = f"postwarden-balance-sheet-{scenario}"
    if as_of:
        name += f"_{as_of}"
    if raw:
        name += "_raw"
    return f"{name}.{ext}"


def balance_sheet_csv(result: dict, scenario: str, as_of: str, raw: int) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Section", "Code", "Account", "Path", "Amount"])
    for r in result["assets"]:
        w.writerow(["Assets", r["account_code"], r["account_name"], r["path"], r["subtotal"]])
    for r in result["liabilities"]:
        w.writerow(["Liabilities", r["account_code"], r["account_name"], r["path"], -r["subtotal"]])
    for r in result["equity"]:
        w.writerow(["Equity", r["account_code"], r["account_name"], r["path"], -r["subtotal"]])
    w.writerow([])
    w.writerow(["Total assets", "", "", "", result["total_assets"]])
    w.writerow(["Total liabilities + equity", "", "", "", result["total_liab_and_equity"]])
    return csv_response(buf, _balance_sheet_filename(scenario, as_of, raw, "csv"))


def balance_sheet_xlsx(result: dict, scenario: str, as_of: str, raw: int) -> Response:
    """XLSX counterpart to `balance_sheet_csv` above. Each section
    (Assets/Liabilities/Equity) gets a bold section-title row, same as
    Trial Balance; a section's own top-level account row (depth 1) gets
    the same "group" bold+ruled treatment Income Statement's groups use,
    since that row's figure already is that root's own rolled-up total —
    no separate "Total X" row needed unless a section actually has more
    than one root. "Total assets"/"Total liabilities + equity" are a
    real cross-section identity, not a duplicate of anything above them,
    so they keep their own grand-total row with the accountant's
    double-rule — red instead of ink when the sheet doesn't balance."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    subtitle = f"{scenario} · {'As of ' + as_of if as_of else 'Through today'}"
    if not raw:
        subtitle += " · simulated monthly close"
    ws.cell(row=1, column=1, value="Balance Sheet").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Amount"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    xlsx.xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, amount, style="line"):
        xlsx.xlsx_data_row(ws, r, [(1, code), (2, name)], [(3, amount, xlsx.XLSX_MONEY_FMT)], style,
                            max(depth - 1, 0))

    r = data_start
    sections = [("Assets", result["assets"], 1), ("Liabilities", result["liabilities"], -1),
                ("Equity", result["equity"], -1)]
    for label, rows, sign in sections:
        xlsx.xlsx_data_row(ws, r, [(1, ""), (2, label)], [], style="group")
        r += 1
        for line in rows:
            row(r, line["account_code"], line["account_name"], line.get("depth", 2),
                sign * line["subtotal"], style="group" if line.get("depth") == 1 else "line")
            r += 1
    r += 1  # blank separator, same breathing room the CSV gives with w.writerow([])
    grand_style = "grand" if result["in_balance"] else "grand_bad"
    row(r, "", "Total assets", 1, result["total_assets"], style=grand_style)
    r += 1
    row(r, "", "Total liabilities + equity", 1, result["total_liab_and_equity"], style=grand_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, _balance_sheet_filename(scenario, as_of, raw, "xlsx"))


# ---------------------------------------------------------------------------
# Income statement — branches on whether `result` came from
# `service.income_statement_rows` (a single range) or
# `service.income_statement_matrix` (Split view, one column-group per
# period) — the two share the same `income_groups`/`expense_groups`
# shape, but a matrix result's top-level `"periods"` key is what tells
# these functions which one they were handed, same as
# `router.income_statement` itself already branches on `split_periods`.
# ---------------------------------------------------------------------------


def _income_statement_filename(scenario: str, compare: str, date_from: str, date_to: str,
                                split: str, ext: str) -> str:
    name = f"postwarden-income-statement-{scenario}"
    if compare:
        name += f"-vs-{compare}"
    if date_from and date_to:
        name += f"_{date_from}_to_{date_to}"
    elif date_from:
        name += f"_from_{date_from}"
    elif date_to:
        name += f"_through_{date_to}"
    if split:
        name += f"_{split}"
    return f"{name}.{ext}"


def income_statement_csv(result: dict, scenario: str, compare: str, date_from: str, date_to: str,
                          split: str) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)

    if "periods" not in result:
        header = ["Section", "Code", "Account", "Path", scenario or "Amount"]
        if compare:
            header += ["Variance", "% variance", compare]
        w.writerow(header)

        def row(section, code, name, path, base, comp=None, variance=None, pct=None):
            line = [section, code, name, path, base]
            if compare:
                line += [variance if variance is not None else "",
                         pct if pct is not None else "",
                         comp if comp is not None else ""]
            w.writerow(line)

        for g in result["income_groups"]:
            for r in g["rows"]:
                row(g["name"], r["account_code"], r["account_name"], r["path"],
                    r["base_net"], r["compare_net"], r["variance"], r["pct_variance"])
        row("Income", "", "Total income", "", result["total_base_income"],
            result["total_compare_income"], result["income_variance_amount"], result["income_variance"])
        for i, g in enumerate(result["expense_groups"]):
            w.writerow([])
            for r in g["rows"]:
                row(g["name"], r["account_code"], r["account_name"], r["path"],
                    r["base_net"], r["compare_net"], r["variance"], r["pct_variance"])
            row(g["name"], "", f"Total {g['name']}", "", g["base_subtotal"],
                g["compare_subtotal"], g["variance"], g["pct_variance"])
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            row(g["name"], "", label, "", g["base_running_after"],
                g["compare_running_after"], g["running_variance"], g["running_pct_variance"])
        if not result["expense_groups"]:
            row("Income", "", "Net income", "", result["net_income"],
                result["compare_net_income"], result["net_income_variance_amount"],
                result["net_income_variance"])
        return csv_response(buf, _income_statement_filename(scenario, compare, date_from, date_to, split, "csv"))

    # Split view: one wide row per account, one group of columns per
    # period instead of one. Each period's own column group is prefixed
    # with that period's label so the header stays legible in a plain
    # spreadsheet with no merged/two-row header the way the JSON->grid
    # rendering can afford — "2026-08 ACTUAL" reads fine as a single
    # Excel column header, "ACTUAL" repeated 3x with a separate period
    # row above it wouldn't survive a CSV round trip at all.
    header = ["Section", "Code", "Account", "Path"]
    for p in result["periods"]:  # real periods + the trailing Totals/Average columns
        header.append(f"{p['label']} {scenario}")
        if compare:
            header += [f"{p['label']} Variance", f"{p['label']} % variance", f"{p['label']} {compare}"]
    w.writerow(header)

    def row(section, code, name, path, period_values):
        line = [section, code, name, path]
        for v in period_values:
            line.append(v.get("base", ""))
            if compare:
                line += [v.get("variance", ""), v.get("pct", ""), v.get("comp", "")]
        w.writerow(line)

    for g in result["income_groups"]:
        for r in g["rows"]:
            row(g["name"], r["account_code"], r["account_name"], r["path"],
                [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                  "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in r["periods"]])
    row("Income", "", "Total income", "",
        [{"base": pt["total_base_income"], "comp": pt["total_compare_income"],
          "variance": pt["income_variance_amount"], "pct": pt["income_variance"]} for pt in result["periods_totals"]])
    for i, g in enumerate(result["expense_groups"]):
        w.writerow([])
        for r in g["rows"]:
            row(g["name"], r["account_code"], r["account_name"], r["path"],
                [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                  "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in r["periods"]])
        row(g["name"], "", f"Total {g['name']}", "",
            [{"base": gp["base_subtotal"], "comp": gp["compare_subtotal"],
              "variance": gp["variance"], "pct": gp["pct_variance"]} for gp in g["periods"]])
        is_last = i == len(result["expense_groups"]) - 1
        label = "Net income" if is_last else f"Net income after {g['name']}"
        row(g["name"], "", label, "",
            [{"base": gp["base_running_after"], "comp": gp["compare_running_after"],
              "variance": gp["running_variance"], "pct": gp["running_pct_variance"]} for gp in g["periods"]])
    if not result["expense_groups"]:
        row("Income", "", "Net income", "",
            [{"base": pt["net_income"], "comp": pt["compare_net_income"],
              "variance": pt["net_income_variance_amount"], "pct": pt["net_income_variance"]}
             for pt in result["periods_totals"]])
    return csv_response(buf, _income_statement_filename(scenario, compare, date_from, date_to, split, "csv"))


def income_statement_xlsx(result: dict, scenario: str, compare: str, date_from: str, date_to: str,
                           split: str, pct_of_base: bool) -> Response:
    """XLSX counterpart to `income_statement_csv` above — same
    `income_statement_rows`/`income_statement_matrix` data, same overall
    shape (income rows, then each expense group with its own running
    "Net income after X" row), styled with `export.xlsx`'s helpers
    instead of written as plain CSV rows. `depth` (from
    `domain.accounts.build_account_tree`) drives indentation the way the
    HTML report's own chevrons did, standing in for the CSV's separate
    Path column — a sighted spreadsheet reader gets the same hierarchy
    from indentation, and dropping it keeps the sheet's column count
    matched to the CSV's data columns instead of growing it.

    No separate "Total income"/"Total {group}" row the way the CSV
    export has one — a group's own top-level account (its `rows[0]`,
    always first since `flatten_tree` puts a node ahead of its children)
    already *is* that rolled-up total; see `domain.accounts.
    build_account_tree`'s "subtotal" comment. That first row gets the
    bold/ruled "group" treatment directly, in place — one real total per
    section, not two.

    Two reader aids from legacy's own live feedback against a real
    download: every Variance/% Variance column also carries real
    conditional-formatting rules (red negative, green positive —
    `xlsx.xlsx_variance_coloring`), not a color baked in at generation
    time. Split view also draws a heavier rule down the right edge of
    every period's own column group (but the last) —
    `xlsx.xlsx_thicken_right_border` — since a wide multi-period sheet
    with only the thin per-cell grid to go on is easy to lose your place
    scanning across. Split's trailing Total/Average column groups also
    get their own bold-plus-tint treatment (italic layered on for
    Average) — `xlsx.xlsx_period_agg_font` — matching the app's own
    `.period-agg`/`.period-agg-average` styling so the two aggregate
    columns stand out from a real period the same way in both places."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    subtitle = scenario
    if compare:
        subtitle += f" vs. {compare}"
    if date_from and date_to:
        subtitle += f" · {date_from} to {date_to}"
    elif date_from:
        subtitle += f" · from {date_from}"
    elif date_to:
        subtitle += f" · through {date_to}"
    if split:
        subtitle += f" · split {split}"
    ws.cell(row=1, column=1, value="Income Statement").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    if "periods" not in result:
        headers = ["Code", "Account", scenario or "Amount"]
        if compare:
            headers += ["Variance", "% Variance", compare]
        n_cols = len(headers)
        header_row, data_start = 4, 5
        xlsx.xlsx_header_row(ws, header_row, headers)

        def row(r: int, code, name, depth, base, comp=None, style="line"):
            # comp is written as a literal; variance/pct are live
            # formulas instead (xlsx_variance_formulas), referencing
            # this same row's own C{r}/F{r} cells — see this function's
            # own docstring on why every base/compare figure stays a
            # literal while Variance/% Variance don't.
            value_cols = [(3, base, xlsx.XLSX_MONEY_FMT)]
            if compare:
                var_f, pct_f = xlsx.xlsx_variance_formulas(f"C{r}", f"F{r}", pct_of_base)
                value_cols += [(4, var_f, xlsx.XLSX_MONEY_FMT), (5, pct_f, xlsx.XLSX_PCT_FMT),
                               (6, comp, xlsx.XLSX_MONEY_FMT)]
            xlsx.xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

        # Root-row cell references for the running-total formulas below —
        # every income group's own root row (base column C, compare
        # column F), then every expense group's as its own root row is
        # written. A "Net income after X" row is then just those income
        # roots minus however many expense roots have been seen so far.
        income_roots_c, income_roots_f, expense_roots_c, expense_roots_f = [], [], [], []
        r = data_start
        for g in result["income_groups"]:
            for i, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    line["base_net"], line["compare_net"], style="group" if i == 0 else "line")
                if i == 0:
                    income_roots_c.append(f"C{r}")
                    income_roots_f.append(f"F{r}")
                r += 1
        for i, g in enumerate(result["expense_groups"]):
            r += 1  # blank separator row — same breathing room the CSV gives with w.writerow([])
            for j, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    line["base_net"], line["compare_net"], style="group" if j == 0 else "line")
                if j == 0:
                    expense_roots_c.append(f"C{r}")
                    expense_roots_f.append(f"F{r}")
                r += 1
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            running_base = xlsx.xlsx_sum_formula(income_roots_c, expense_roots_c)
            running_comp = xlsx.xlsx_sum_formula(income_roots_f, expense_roots_f) if compare else None
            row(r, "", label, 1, running_base, running_comp, style="running")
            r += 1
        if not result["expense_groups"]:
            running_base = xlsx.xlsx_sum_formula(income_roots_c)
            running_comp = xlsx.xlsx_sum_formula(income_roots_f) if compare else None
            row(r, "", "Net income", 1, running_base, running_comp, style="running")
            r += 1
        last_row = r - 1
        if compare:
            xlsx.xlsx_variance_coloring(ws, 4, data_start, last_row)  # Variance
            xlsx.xlsx_variance_coloring(ws, 5, data_start, last_row)  # % Variance
    else:
        cols_per_period = 4 if compare else 1
        field_labels = [scenario] + (["Variance", "% Var.", compare] if compare else [])
        n_cols = 2 + cols_per_period * len(result["periods"])
        header_row, field_row, data_start = 4, 5, 6

        # Two-row header: the date ("2026-01", "Total", "Average") merged
        # and centered across that period's own field columns, with the
        # field names (ACTUAL/Variance/%/compare) on their own row right
        # below instead of repeated into every column header.
        xlsx.xlsx_merged_header(ws, header_row, 1, field_row, 1, "Code")
        xlsx.xlsx_merged_header(ws, header_row, 2, field_row, 2, "Account")
        for i, p in enumerate(result["periods"]):
            start_col = 3 + i * cols_per_period
            xlsx.xlsx_merged_header(ws, header_row, start_col, header_row, start_col + cols_per_period - 1,
                                     p["label"])
            xlsx.xlsx_header_row(ws, field_row, field_labels, start_col=start_col)

        def row(r: int, code, name, depth, period_vals, style="line"):
            value_cols, col = [], 3
            for v in period_vals:
                base_col = col
                value_cols.append((col, v.get("base"), xlsx.XLSX_MONEY_FMT))
                col += 1
                if compare:
                    comp_col = col + 2
                    var_f, pct_f = xlsx.xlsx_variance_formulas(
                        f"{get_column_letter(base_col)}{r}", f"{get_column_letter(comp_col)}{r}", pct_of_base)
                    value_cols += [(col, var_f, xlsx.XLSX_MONEY_FMT),
                                   (col + 1, pct_f, xlsx.XLSX_PCT_FMT),
                                   (comp_col, v.get("comp"), xlsx.XLSX_MONEY_FMT)]
                    col += 3
            xlsx.xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

        def running_period_vals(income_rows: list[int], expense_rows: list[int]) -> list[dict]:
            """One {"base", "comp"} formula pair per period column-group
            for a "Net income after X" row — `income_rows`/`expense_rows`
            name *row* numbers (one physical row per account, shared
            across every period) while the actual cell reference still
            needs that period's own column."""
            out = []
            for i in range(len(result["periods"])):
                base_col = get_column_letter(3 + i * cols_per_period)
                base_f = xlsx.xlsx_sum_formula([f"{base_col}{rr}" for rr in income_rows],
                                                [f"{base_col}{rr}" for rr in expense_rows])
                comp_f = None
                if compare:
                    comp_col = get_column_letter(3 + i * cols_per_period + 3)
                    comp_f = xlsx.xlsx_sum_formula([f"{comp_col}{rr}" for rr in income_rows],
                                                    [f"{comp_col}{rr}" for rr in expense_rows])
                out.append({"base": base_f, "comp": comp_f})
            return out

        # Root *row numbers* only here (not column letters) — one
        # physical row per account, the same row number reused across
        # every period's own column-group.
        income_root_rows, expense_root_rows = [], []
        r = data_start
        for g in result["income_groups"]:
            for i, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    [{"base": rp.get("base_net"), "comp": rp.get("compare_net")} for rp in line["periods"]],
                    style="group" if i == 0 else "line")
                if i == 0:
                    income_root_rows.append(r)
                r += 1
        for i, g in enumerate(result["expense_groups"]):
            r += 1
            for j, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    [{"base": rp.get("base_net"), "comp": rp.get("compare_net")} for rp in line["periods"]],
                    style="group" if j == 0 else "line")
                if j == 0:
                    expense_root_rows.append(r)
                r += 1
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            row(r, "", label, 1, running_period_vals(income_root_rows, expense_root_rows), style="running")
            r += 1
        if not result["expense_groups"]:
            row(r, "", "Net income", 1, running_period_vals(income_root_rows, []), style="running")
            r += 1
        last_row = r - 1

        if compare:
            # A heavier rule down the right edge of every period's own
            # column group but the last (Total/Average count as periods
            # here too — they're just two more column groups in
            # result["periods"]). Compare-only: with no compare scenario
            # each period is a single plain column (cols_per_period ==
            # 1), and a divider after literally every column would read
            # as clutter rather than a period boundary.
            for i, p in enumerate(result["periods"][:-1]):
                end_col = 3 + i * cols_per_period + cols_per_period - 1
                for rr in range(header_row, last_row + 1):
                    xlsx.xlsx_thicken_right_border(ws, rr, end_col)
            for i, p in enumerate(result["periods"]):
                start_col = 3 + i * cols_per_period
                xlsx.xlsx_variance_coloring(ws, start_col + 1, data_start, last_row)  # Variance
                xlsx.xlsx_variance_coloring(ws, start_col + 2, data_start, last_row)  # % Variance

        # Total/Average's own "stand out" treatment (bold + tint, italic
        # layered on for Average) — every data row, not the header rows:
        # the merged period-label header ("Total"/"Average" instead of a
        # date) already reads as different from a real period on its
        # own. Unconditional (not inside `if compare:` above) — Total/
        # Average exist whether or not there's a compare scenario.
        for i, p in enumerate(result["periods"]):
            if not (p.get("is_total") or p.get("is_average")):
                continue
            force_italic = bool(p.get("is_average"))
            start_col = 3 + i * cols_per_period
            for col in range(start_col, start_col + cols_per_period):
                for rr in range(data_start, last_row + 1):
                    cell = ws.cell(row=rr, column=col)
                    cell.font = xlsx.xlsx_period_agg_font(cell.font, force_italic=force_italic)
                    cell.fill = xlsx.XLSX_PERIOD_AGG_FILL

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, _income_statement_filename(scenario, compare, date_from, date_to, split, "xlsx"))


# ---------------------------------------------------------------------------
# Cash flow
# ---------------------------------------------------------------------------


def _cash_flow_filename(scenario: str, date_from: str, date_to: str, ext: str) -> str:
    name = f"postwarden-cash-flow-{scenario}"
    if date_from and date_to:
        name += f"_{date_from}_to_{date_to}"
    elif date_from:
        name += f"_from_{date_from}"
    elif date_to:
        name += f"_through_{date_to}"
    return f"{name}.{ext}"


def cash_flow_csv(result: dict, scenario: str, date_from: str, date_to: str) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)

    def netted_of(r: dict) -> str:
        return "; ".join(f"{n['account_name']} {n['amount']:.2f}" for n in r["netted_from"])

    w.writerow(["Section", "Code", "Account", "Amount", "Flagged for review", "Net of"])
    w.writerow(["", "", "Beginning cash balance", result["tie_out"]["beginning"], "", ""])
    w.writerow([])
    for r in result["inflows"]:
        w.writerow(["Inflows", r["account_code"], r["account_name"], r["amount"],
                    "yes" if r["flagged"] else "", netted_of(r)])
    w.writerow(["Inflows", "", "Total inflows", result["total_inflows"], "", ""])
    w.writerow([])
    for r in result["outflows"]:
        w.writerow(["Outflows", r["account_code"], r["account_name"], r["amount"],
                    "yes" if r["flagged"] else "", netted_of(r)])
    w.writerow(["Outflows", "", "Total outflows", result["total_outflows"], "", ""])
    w.writerow([])
    # Only present when non-empty — most periods have no equity-contra
    # activity at all, so an always-present empty section would just be
    # noise most exports.
    if result["ledger_adjustments"]:
        for r in result["ledger_adjustments"]:
            w.writerow(["Ledger adjustments", r["account_code"], r["account_name"], r["amount"],
                        "yes" if r["flagged"] else "", ""])
        w.writerow(["Ledger adjustments", "", "Total ledger adjustments", result["total_adjustments"], "", ""])
        w.writerow([])
    w.writerow(["", "", "Net change in cash", result["net_change"], "", ""])
    w.writerow(["", "", "Ending cash balance", result["tie_out"]["ending"], "", ""])
    w.writerow([])
    w.writerow(["", "", "Tie-out check", "PASS" if result["tie_out"]["ok"] else "FAIL — see app log", "", ""])
    return csv_response(buf, _cash_flow_filename(scenario, date_from, date_to, "csv"))


def cash_flow_xlsx(result: dict, scenario: str, date_from: str, date_to: str) -> Response:
    """XLSX counterpart to `cash_flow_csv` above. No account tree here
    (`fn_cash_flow_lines`' rows are already flat, one per contra
    account), so no depth/indent and no "first row is the total"
    duplication concern the tree-shaped reports have. Beginning/Net
    change/Ending get the same bold "group" headline treatment as a
    section title, and the closing Tie-out row reuses the grand/
    grand_bad split Trial Balance/Balance Sheet use for their own
    balance check — green ink for PASS, red for FAIL, same accountant's
    double-rule underneath."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    subtitle = scenario
    if date_from and date_to:
        subtitle += f" · {date_from} to {date_to}"
    elif date_from:
        subtitle += f" · from {date_from}"
    elif date_to:
        subtitle += f" · through {date_to}"
    ws.cell(row=1, column=1, value="Cash Flow Statement").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Amount", "Flagged", "Net of"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    xlsx.xlsx_header_row(ws, header_row, headers)

    def netted_of(line: dict) -> str:
        return "; ".join(f"{n['account_name']} {n['amount']:.2f}" for n in line["netted_from"])

    def row(r: int, code, name, amount, flagged="", netted="", style="line"):
        label_cols = [(1, code), (2, name)]
        value_cols = [(3, amount, xlsx.XLSX_MONEY_FMT)]
        xlsx.xlsx_data_row(ws, r, label_cols, value_cols, style)
        # Flagged/Net of are plain descriptive text, not money — no
        # banding/border/number-format, same as every report's non-money
        # label columns.
        font = xlsx.XLSX_ROW_FONTS[style]
        for col, text in ((4, flagged), (5, netted)):
            cell = ws.cell(row=r, column=col, value=text or None)
            cell.font = font

    r = data_start
    row(r, "", "Beginning cash balance", result["tie_out"]["beginning"], style="group")
    r += 2  # blank separator row, same breathing room the CSV gives with w.writerow([])
    for section, rows_, total_label, total in (
        ("Inflows", result["inflows"], "Total inflows", result["total_inflows"]),
        ("Outflows", result["outflows"], "Total outflows", result["total_outflows"]),
    ):
        xlsx.xlsx_data_row(ws, r, [(1, ""), (2, section)], [], style="group")
        r += 1
        for line in rows_:
            row(r, line["account_code"], line["account_name"], line["amount"],
                "yes" if line["flagged"] else "", netted_of(line))
            r += 1
        row(r, "", total_label, total, style="subtotal")
        r += 2
    if result["ledger_adjustments"]:
        xlsx.xlsx_data_row(ws, r, [(1, ""), (2, "Ledger adjustments")], [], style="group")
        r += 1
        for line in result["ledger_adjustments"]:
            row(r, line["account_code"], line["account_name"], line["amount"],
                "yes" if line["flagged"] else "")
            r += 1
        row(r, "", "Total ledger adjustments", result["total_adjustments"], style="subtotal")
        r += 2
    row(r, "", "Net change in cash", result["net_change"], style="group")
    r += 1
    row(r, "", "Ending cash balance", result["tie_out"]["ending"], style="group")
    r += 2
    tie_style = "grand" if result["tie_out"]["ok"] else "grand_bad"
    tie_text = "PASS" if result["tie_out"]["ok"] else "FAIL — see app log"
    xlsx.xlsx_data_row(ws, r, [(1, ""), (2, "Tie-out check")], [(3, tie_text, "General")], style=tie_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 36
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, _cash_flow_filename(scenario, date_from, date_to, "xlsx"))


# ---------------------------------------------------------------------------
# Variance
# ---------------------------------------------------------------------------


def _variance_filename(baseline: str, compare: str, as_of: str, ext: str) -> str:
    name = f"postwarden-variance-{baseline}-vs-{compare}"
    if as_of:
        name += f"_{as_of}"
    return f"{name}.{ext}"


def variance_csv(result: dict, baseline: str, as_of: str) -> Response:
    compare = result["compare"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Code", "Account", "Path", baseline, "Variance", compare])
    for r in result["merged"]:
        w.writerow([r["account_code"], r["account_name"], r["path"],
                    r["baseline_net"], r["variance"], r["compare_net"]])
    return csv_response(buf, _variance_filename(baseline, compare, as_of, "csv"))


def variance_xlsx(result: dict, baseline: str, as_of: str, pct_of_base: bool) -> Response:
    """XLSX counterpart to `variance_csv` above — built from
    `result["grouped"]` (one section per account type, same as Trial
    Balance) rather than the CSV's flat `result["merged"]` list, so it
    can add the section headers, the per-type subtotal, and a real %
    Variance column the CSV leaves out (the figure's already computed
    either way — CSV just never had a use for it without a companion
    section structure).

    Native mode (no rollup level) builds a real account tree, same as
    every other tree-shaped report here — a section's own depth-1 row
    gets the "group" bold+ruled treatment, and an explicit subtotal row
    only when a section actually has more than one top-level account.
    Rolled-up mode (`level_id` set) has no tree at all —
    `result["merged"]`'s rows are already one flat pooled figure per
    account at that level, so every row stays "line" style and every
    section always gets its own subtotal row, since no single row in a
    rolled-up section is ever "the" total the way a tree's own root row
    is.

    Variance and % Variance are live formulas
    (`xlsx.xlsx_variance_formulas`), same as Income Statement's own
    export — each one only references its own row's baseline/compare
    cells, never a range, so it's safe regardless of whether that row's
    own baseline/compare figures are a leaf balance, a rolled-up root,
    or a section subtotal."""
    compare = result["compare"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Variance"

    subtitle = f"{baseline} vs. {compare}"
    subtitle += f" · {'As of ' + as_of if as_of else 'Through today'}"
    if result["rolled_up"]:
        subtitle += " · rolled up"
    ws.cell(row=1, column=1, value="Variance").font = xlsx.XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = xlsx.XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", baseline, "Variance", "% Variance", compare]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    xlsx.xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, base, comp, style="line"):
        # Variance/% Variance are live formulas instead of the
        # already-computed figures — C (base/baseline), then F
        # (compare), same canonical order compute_variance's own
        # variance_amount(baseline_net, compare_net, ...) calls use.
        var_f, pct_f = xlsx.xlsx_variance_formulas(f"C{r}", f"F{r}", pct_of_base)
        value_cols = [(3, base, xlsx.XLSX_MONEY_FMT), (4, var_f, xlsx.XLSX_MONEY_FMT),
                      (5, pct_f, xlsx.XLSX_PCT_FMT), (6, comp, xlsx.XLSX_MONEY_FMT)]
        xlsx.xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

    r = data_start
    for g in result["grouped"]:
        xlsx.xlsx_data_row(ws, r, [(1, ""), (2, g["label"])], [], style="group")
        r += 1
        top_level_count = sum(1 for line in g["rows"] if not result["rolled_up"] and line.get("depth") == 1)
        for line in g["rows"]:
            is_root = not result["rolled_up"] and line.get("depth") == 1
            row(r, line["account_code"], line["account_name"], line.get("depth", 1),
                line["baseline_net"], line["compare_net"], style="group" if is_root else "line")
            r += 1
        if result["rolled_up"] or top_level_count > 1:
            row(r, "", f"{g['label']} subtotal", 1, g["sub_baseline"], g["sub_compare"], style="subtotal")
            r += 1
    row(r, "", "Total", 1, result["total_baseline"], result["total_compare"], style="grand")
    last_row = r
    xlsx.xlsx_variance_coloring(ws, 4, data_start, last_row)  # Variance
    xlsx.xlsx_variance_coloring(ws, 5, data_start, last_row)  # % Variance

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx.xlsx_response(wb, _variance_filename(baseline, compare, as_of, "xlsx"))

"""Unit tests for `export.xlsx` — pure openpyxl manipulation, no
database. Each helper is exercised against a throwaway `Workbook`'s
active worksheet, checking the cell properties it's actually documented
to set rather than re-deriving openpyxl's own rendering."""
import io

from openpyxl import Workbook, load_workbook

from postwarden.export import xlsx


def test_xlsx_header_row_styles_every_header_cell():
    ws = Workbook().active
    xlsx.xlsx_header_row(ws, 1, ["Code", "Account"])
    for col in (1, 2):
        cell = ws.cell(row=1, column=col)
        assert cell.font == xlsx.XLSX_HEADER_FONT
        assert cell.fill.fgColor.rgb == "FF1B2430"
        assert cell.alignment.horizontal == "center"


def test_xlsx_merged_header_merges_and_styles_the_anchor_only():
    ws = Workbook().active
    xlsx.xlsx_merged_header(ws, 1, 1, 2, 3, "2026-01")
    assert {str(r) for r in ws.merged_cells.ranges} == {"A1:C2"}
    assert ws.cell(row=1, column=1).value == "2026-01"
    assert ws.cell(row=1, column=1).font == xlsx.XLSX_HEADER_FONT


def test_xlsx_thicken_right_border_keeps_the_other_three_edges():
    ws = Workbook().active
    xlsx.xlsx_data_row(ws, 1, [(1, "1000"), (2, "Assets")], [(3, 100, xlsx.XLSX_MONEY_FMT)], "line")
    original = ws.cell(row=1, column=3).border
    xlsx.xlsx_thicken_right_border(ws, 1, 3)
    thickened = ws.cell(row=1, column=3).border
    assert thickened.right.style == "medium"
    assert thickened.left == original.left
    assert thickened.top == original.top
    assert thickened.bottom == original.bottom


def test_xlsx_variance_coloring_adds_two_conditional_rules():
    ws = Workbook().active
    xlsx.xlsx_variance_coloring(ws, 4, 5, 10)
    rules = list(ws.conditional_formatting)
    # openpyxl groups rules by the exact range string they were added
    # against — one ConditionalFormattingList entry for "D5:D10", holding
    # both the negative and positive CellIsRule.
    assert len(rules) == 1
    assert str(rules[0].sqref) == "D5:D10"
    assert len(rules[0].rules) == 2


def test_xlsx_variance_formulas_default_is_base_minus_compare_pct_of_compare():
    var_f, pct_f = xlsx.xlsx_variance_formulas("C6", "F6", pct_of_base=False)
    assert var_f == "=C6-F6"
    assert pct_f == '=IF(F6=0,"",ROUND((C6-F6)/ABS(F6)*100,1))'


def test_xlsx_variance_formulas_pct_of_base_flips_direction():
    var_f, pct_f = xlsx.xlsx_variance_formulas("C6", "F6", pct_of_base=True)
    assert var_f == "=F6-C6"
    assert pct_f == '=IF(C6=0,"",ROUND((F6-C6)/ABS(C6)*100,1))'


def test_xlsx_sum_formula_joins_plus_and_minus_cells():
    assert xlsx.xlsx_sum_formula(["C6", "C20"], ["C34"]) == "=C6+C20-C34"
    assert xlsx.xlsx_sum_formula(["C6"]) == "=C6"


def test_xlsx_sum_formula_falls_back_to_a_literal_zero_when_empty():
    assert xlsx.xlsx_sum_formula([], []) == 0


def test_xlsx_period_agg_font_forces_bold_and_carries_italic_forward():
    plain = xlsx.XLSX_LINE_FONT
    agg = xlsx.xlsx_period_agg_font(plain)
    assert agg.bold is True
    assert agg.italic is False
    running_in_total_column = xlsx.xlsx_period_agg_font(xlsx.XLSX_RUNNING_FONT)
    assert running_in_total_column.italic is True  # already italic, stays italic
    average_column = xlsx.xlsx_period_agg_font(plain, force_italic=True)
    assert average_column.italic is True


def test_xlsx_data_row_line_style_shades_and_grids_value_cells():
    ws = Workbook().active
    xlsx.xlsx_data_row(ws, 5, [(1, "1100"), (2, "Checking")], [(3, 100, xlsx.XLSX_MONEY_FMT)], "line", depth=1)
    label_cell = ws.cell(row=5, column=2)
    value_cell = ws.cell(row=5, column=3)
    assert label_cell.alignment.indent == 1
    assert value_cell.fill.fgColor.rgb == "FFEEF0F3"
    assert value_cell.number_format == xlsx.XLSX_MONEY_FMT
    assert value_cell.font == xlsx.XLSX_LINE_FONT


def test_xlsx_data_row_grand_bad_style_uses_the_red_double_rule():
    ws = Workbook().active
    xlsx.xlsx_data_row(ws, 1, [(1, ""), (2, "Out of balance")], [(3, 1, xlsx.XLSX_MONEY_FMT)], "grand_bad")
    cell = ws.cell(row=1, column=3)
    assert cell.font == xlsx.XLSX_GRAND_FONT_BAD
    assert cell.border.bottom.style == "double"
    assert cell.border.bottom.color.rgb == "FFB3392C"


def test_xlsx_response_round_trips_through_openpyxl():
    wb = Workbook()
    wb.active.cell(row=1, column=1, value="hello")
    resp = xlsx.xlsx_response(wb, "report.xlsx")
    assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.headers["content-disposition"] == 'attachment; filename="report.xlsx"'
    reloaded = load_workbook(io.BytesIO(resp.body))
    assert reloaded.active.cell(row=1, column=1).value == "hello"

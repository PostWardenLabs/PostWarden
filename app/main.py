"""PostWarden — a personal general ledger with scenarios.

HTML screens for humans, /api/* JSON for machines, PostgreSQL for the truth.
"""
import calendar
import csv
import io
import json
import logging
import os
import re
import secrets
from contextlib import asynccontextmanager
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import urlencode

import psycopg
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from markupsafe import Markup, escape
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import auth
from .db import q, q1, tx
from .migrate import run_migrations


@asynccontextmanager
async def lifespan(app: FastAPI):
    run_migrations()
    auth.bootstrap_admin_from_env()
    yield


BASE = Path(__file__).parent
# Used only by the Cash Flow Statement's tie-out check (see
# _cash_flow_tie_out below) to log a mismatch — nothing else in this app
# has needed a logger before now. `docker compose logs app` is already
# how this project verifies a deploy (CLAUDE.md), so a plain module
# logger writing to stderr, uvicorn's default handler, needs no setup.
logger = logging.getLogger(__name__)
app = FastAPI(title="PostWarden", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=BASE / "static"), name="static")
templates = Jinja2Templates(directory=BASE / "templates")

PUBLIC_PATHS = {"/login"}


@app.middleware("http")
async def auth_gate(request: Request, call_next):
    """Every route needs a valid session except /login and static assets.
    Sets request.state.user once per request so handlers and templates
    (`request.state.user`) don't each need their own session lookup."""
    path = request.url.path
    if path in PUBLIC_PATHS or path.startswith("/static/"):
        return await call_next(request)
    session = auth.get_session(request.cookies.get(auth.SESSION_COOKIE))
    if not session:
        if path.startswith("/api/"):
            return JSONResponse({"detail": "Not authenticated"}, status_code=401)
        return RedirectResponse("/login", status_code=303)
    request.state.user = session
    # No task runner in this deployment, so "auto-post on the date" is done
    # lazily here instead of on a real cron: cheap (an indexed SELECT that's
    # almost always empty) and each due schedule only ever materializes once
    # since materialize_due_schedules() advances next_date past today as
    # part of the same write. A failure here shouldn't take the app down.
    try:
        materialize_due_schedules()
    except Exception:
        pass
    return await call_next(request)


def money(v) -> Markup:
    """Renders as American-default text ("1,234.56", no symbol) so the page
    is correct even with JS disabled, but wraps it in a span carrying the
    plain numeric value — money-format.js rewrites every .money-fmt's
    displayed text from that raw value using whatever symbol/decimal/
    thousands preferences are saved in Settings (client-side only, same
    as the theme picker; the number stored in Postgres never changes)."""
    if v is None:
        return Markup("")
    # A flipped-sign zero (income rows negate their stored net — see
    # _income_statement_groups' own `sign` — so a zero-balance income
    # account's base_net is -1 * 0) is still a genuine Decimal/float
    # negative zero, which %.2f-style formatting renders as the visually
    # confusing "-0.00" rather than "0.00" — same magnitude, no reason to
    # show a sign on it. money-format.js's own client-side rewrite
    # already normalizes this away for display, but the plain data-value
    # attribute underneath (and anything reading the raw figure directly,
    # like a CSV export) doesn't get that rescue, so fix it at the source
    # instead of leaning on the client-side one being present.
    if v == 0:
        v = abs(v)
    return Markup(f'<span class="money-fmt" data-value="{v:.2f}">{v:,.2f}</span>')


def dateformat(v) -> Markup:
    """Same pattern as money() above: renders as plain ISO text
    ("2026-08-26") so the page is correct with JS disabled, wrapped in a
    span carrying that same ISO value — date-format.js rewrites every
    .date-fmt's displayed text using whatever format is saved in
    Settings (client-side only; the date stored in Postgres never
    changes, and every date in this app is a plain DATE, no time
    component, so there's no timezone conversion to get wrong here —
    just which order y/m/d print in)."""
    if v is None:
        return Markup("")
    iso = v.isoformat() if hasattr(v, "isoformat") else str(v)
    return Markup(f'<span class="date-fmt" data-value="{iso}">{iso}</span>')


def asset(filename: str) -> str:
    """Cache-busting URL for a static file: /static/x?v=<mtime>. Without a
    version query param, browsers can silently keep serving an old cached
    copy of app.js/style.css after a deploy — a plain reload isn't enough
    to guarantee a re-fetch. Appending the file's mtime changes the URL
    (and cache key) exactly when the file itself changes, no build step."""
    v = int((BASE / "static" / filename).stat().st_mtime)
    return f"/static/{filename}?v={v}"


def tojson(value) -> Markup:
    """Embed a value as JSON inside a <script type="application/json">
    block. Escapes <, >, & so a value containing e.g. "</script>" (account
    names are free text) can't break out of the tag; JSON.parse on the
    other end decodes the \\uXXXX escapes back to the real characters, so
    this only protects the HTML-embedding boundary — it says nothing about
    how the parsed value gets used afterwards. app.js only ever reads these
    parsed values via .textContent/.value (DOM APIs, never innerHTML), so
    there's no second escaping step needed there; keep it that way."""
    encoded = json.dumps(value).replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")
    return Markup(encoded)


templates.env.filters["money"] = money
templates.env.filters["dateformat"] = dateformat
templates.env.filters["tojson"] = tojson
templates.env.globals["asset"] = asset
# Read once at startup, not per-request — the footer's version number
# changes on deploy, not while the app is running. Bump VERSION (repo
# root, plain text, no build step needed to pick it up) as part of
# whatever commit it's marking, same as any other doc that ships with
# the feature it describes.
templates.env.globals["version"] = (BASE.parent / "VERSION").read_text().strip()
# Opt-in, off by default, and deliberately a *second* flag rather than just
# checking whether POSTWARDEN_ADMIN_USER/PASSWORD are set: those two are meant
# for a normal self-hoster's own private first-boot convenience (see
# auth.bootstrap_admin_from_env), and plenty of READMEs recommend setting
# them. If showing the banner just meant "those two are set," every such
# self-hoster's own real password would get echoed onto their own login
# page. POSTWARDEN_DEMO_MODE has to be set on top of them, only ever true on
# the actual public demo, for the banner to appear at all.
templates.env.globals["demo_banner"] = os.environ.get("POSTWARDEN_DEMO_MODE", "").lower() in ("1", "true", "yes")
templates.env.globals["demo_user"] = os.environ.get("POSTWARDEN_ADMIN_USER", "")
templates.env.globals["demo_password"] = os.environ.get("POSTWARDEN_ADMIN_PASSWORD", "")

ACCOUNT_TYPES = ["asset", "liability", "equity", "income", "expense"]
TYPE_LABELS = {
    "asset": "Assets", "liability": "Liabilities", "equity": "Equity",
    "income": "Income", "expense": "Expenses",
}
SCENARIO_TYPES = ["actual", "budget", "forecast", "what_if"]


def scenarios_all():
    return q("""SELECT s.*, al.name AS base_level_name,
                       (SELECT COUNT(*) FROM journal_entries e
                         WHERE e.scenario_id = s.id) AS entry_count
                  FROM scenarios s
                  LEFT JOIN account_levels al ON al.id = s.base_level_id
                 ORDER BY s.scenario_type, s.code""")


def account_levels_all():
    return q("""SELECT al.*,
                       (SELECT COUNT(*) FROM scenarios s
                         WHERE s.base_level_id = al.id) AS scenario_count
                  FROM account_levels al ORDER BY al.depth""")


def postable_accounts_for_pickers():
    """Every account ANY scenario could actually post a line to — the union
    across all of them. Used only where there's no single scenario to be
    precise about: entry_templates.html isn't scenario-bound at all."""
    return q("""SELECT id, code, name, path FROM v_dim_account
                WHERE is_active AND (
                    is_postable
                    OR depth IN (SELECT al.depth FROM account_levels al
                                  JOIN scenarios s ON s.base_level_id = al.id)
                )
                ORDER BY sort_path""")


def postable_accounts_by_scenario() -> dict:
    """{scenario_id: [{id, code, name, path}, ...]} — each scenario's own
    exact posting targets, matching fn_line_account_guard precisely (true
    leaves, plus anything at that scenario's own base_level depth if it
    has one). Powers entries.html's "New entry" panel and scheduled.html's
    account picker,
    which re-filters to this when the Scenario field changes (see
    app.js's refreshAccountsForScenario()) instead of showing the same
    broadened list regardless of which scenario is selected."""
    rows = q("""
        SELECT s.id AS scenario_id, d.id, d.code, d.name, d.path
          FROM scenarios s
          JOIN v_dim_account d ON d.is_active AND (
              d.is_postable
              OR (s.base_level_id IS NOT NULL AND d.depth = (
                  SELECT al.depth FROM account_levels al WHERE al.id = s.base_level_id))
          )
         ORDER BY s.id, d.sort_path""")
    by_scenario: dict = {}
    for r in rows:
        by_scenario.setdefault(r["scenario_id"], []).append(
            {"id": r["id"], "code": r["code"], "name": r["name"], "path": r["path"]})
    return by_scenario


def flash_url(url: str, ok: str = None, err: str = None) -> str:
    params = {}
    if ok:
        params["ok"] = ok
    if err:
        params["err"] = err
    sep = "&" if "?" in url else "?"
    return url + (sep + urlencode(params) if params else "")


def flash_redirect(url: str, ok: str = None, err: str = None):
    return RedirectResponse(flash_url(url, ok, err), status_code=303)


def require_csrf(request: Request, token: str | None):
    """Raise ValueError (caught the same way as any other bad input) if the
    submitted token doesn't match this session's. Every state-changing POST
    calls this — see the hidden csrf_token field templates render from
    request.state.user.csrf_token."""
    user = auth.current_user(request)
    if not user or not token or not secrets.compare_digest(token, user["csrf_token"]):
        raise ValueError("Your session expired or the form was stale — please retry.")


def csv_response(buf: io.StringIO, filename: str) -> Response:
    """Wrap a finished csv.writer buffer as a download. Excel — still the
    most likely destination for these files — assumes the system codepage
    for a UTF-8 file with no signature, so any accented name or currency
    symbol comes back as mojibake; a leading BOM is what tells it the file
    is actually UTF-8. /import already decodes with utf-8-sig, so a
    round-tripped export reads back in fine."""
    return Response("﻿" + buf.getvalue(), media_type="text/csv; charset=utf-8", headers={
        "Content-Disposition": f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# XLSX export styling — shared by every /export/*.xlsx route, not just
# Income Statement's. Colors come from style.css's default "Slate" theme
# (`--ink`/`--paper-deep`) rather than whatever theme the browser has
# active — the export is generated server-side with no idea which of the
# nine themes Settings has picked, and a report someone opens in Excel a
# year later shouldn't depend on that anyway. Kept as one small palette
# here rather than a full port of style.css's theme system: this only
# ever needs to look like "a PostWarden document," not match the live
# page pixel-for-pixel.
_XLSX_FONT = "Arial"
# Explicit "FF" alpha on every ARGB string below, not just the bare RGB —
# openpyxl silently zero-pads a 6-digit color to "00RRGGBB" (fully
# transparent alpha) rather than "FFRRGGBB" (opaque) if you don't. Excel
# itself ignores that byte for a solid fill and renders it opaque either
# way, but other readers (LibreOffice, Google Sheets) don't all make the
# same forgiving choice, so relying on Excel's leniency isn't worth it.
_XLSX_HEADER_FILL = PatternFill("solid", fgColor="FF1B2430")   # --ink
_XLSX_HEADER_FONT = Font(name=_XLSX_FONT, size=10, bold=True, color="FFFFFFFF")
_XLSX_GROUP_FONT = Font(name=_XLSX_FONT, size=10, bold=True)
_XLSX_LINE_FONT = Font(name=_XLSX_FONT, size=10)
_XLSX_RUNNING_FONT = Font(name=_XLSX_FONT, size=10, italic=True)
_XLSX_TITLE_FONT = Font(name=_XLSX_FONT, size=14, bold=True, color="FF1B2430")
_XLSX_SUBTITLE_FONT = Font(name=_XLSX_FONT, size=9, italic=True, color="FF5B6B7C")  # --ink-soft
_XLSX_LINE_FILL = PatternFill("solid", fgColor="FFEEF0F3")     # a shade off --paper-deep
_XLSX_BOTTOM_BORDER = Border(bottom=Side(style="thin", color="FFAEBBC7"))  # --rule-strong
_XLSX_RULE = Side(style="thin", color="FFAEBBC7")  # --rule-strong
_XLSX_LINE_BORDER = Border(left=_XLSX_RULE, right=_XLSX_RULE, top=_XLSX_RULE, bottom=_XLSX_RULE)
# Split's period-group divider — heavier than the plain grid rule above,
# so the eye catches "new period starts here" scanning across a wide
# sheet the same way a ruled column break would in a printed ledger.
_XLSX_PERIOD_DIVIDER = Side(style="medium", color="FF1B2430")  # --ink
# Same red/green the HTML report's own .neg (style.css) already uses for
# a negative figure — --red/--ok — so a variance reads the same way in
# the browser and in the spreadsheet. Real conditional-formatting rules
# (CellIsRule below), not a color baked in at generation time, so the
# color still tracks correctly if a variance cell is edited by hand
# later — it's font-only (no fill/border in the rule), so it layers over
# whatever base style (line/group/running) that cell already has rather
# than replacing it.
_XLSX_NEG_FONT = Font(color="FFB3392C")  # --red
_XLSX_POS_FONT = Font(color="FF1F7A52")  # --ok
# The other three reports' own row styles, matching style.css's own
# tr.subtotal/tr.grand treatment exactly (see style.css for both) so a
# Trial Balance/Balance Sheet/Variance export reads the same way in a
# spreadsheet as it does on the page: "subtotal" (a rolled-up figure
# across more than one top-level account, e.g. Trial Balance's own
# per-type subtotal when a type has multiple roots) is semi-bold and
# muted rather than fully bold, so it doesn't compete visually with a
# "group" row's own bold; "grand" is the accountant's double-rule under
# a total that balances — Excel's "double" border style — with a red
# variant for a report that doesn't (Trial Balance/Balance Sheet's own
# in_balance check, Cash Flow's tie-out).
_XLSX_SUBTOTAL_FONT = Font(name=_XLSX_FONT, size=10, bold=True, color="FF5B6B7C")  # --ink-soft
_XLSX_GRAND_FONT = Font(name=_XLSX_FONT, size=10, bold=True)
_XLSX_GRAND_FONT_BAD = Font(name=_XLSX_FONT, size=10, bold=True, color="FFB3392C")  # --red
_XLSX_GRAND_BORDER = Border(bottom=Side(style="double", color="FF1B2430"))  # --ink
_XLSX_GRAND_BORDER_BAD = Border(bottom=Side(style="double", color="FFB3392C"))  # --red
# No currency symbol — matches money()'s own plain-text convention above
# (the app never bakes a symbol into a stored/exported figure; display-only
# formatting is a client-side concern there, and there's no client here).
# Parens for negatives, a bare dash for zero, same shape as money()'s
# comment on negative-zero applies to here too.
_XLSX_MONEY_FMT = '#,##0.00;(#,##0.00);"-"'
# _pct_variance() already returns the percentage figure itself (12.3
# meaning "12.3%"), not a 0-1 fraction, so this appends a literal "%"
# rather than using Excel's built-in 0.0% format, which would multiply
# the already-multiplied number by 100 again.
_XLSX_PCT_FMT = '0.0"%";(0.0"%");"-"'


def _xlsx_header_row(ws, row: int, headers: list[str], start_col: int = 1):
    for col, text in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_merged_header(ws, r1: int, c1: int, r2: int, c2: int, text: str):
    """One header cell spanning r1:r2 × c1:c2, merged and centered —
    Split's per-period date ("2026-01") sitting above that period's own
    ACTUAL/Variance/%/compare columns, or (c1==c2 spanning both header
    rows) the Code/Account label sitting beside them. Only the anchor
    cell (top-left of the merge) needs the header styling — Excel and
    every reader that respects merges renders a merged range entirely
    from that cell, ignoring whatever the covered-but-hidden cells carry,
    so styling those individually would be dead work."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.font = _XLSX_HEADER_FONT
    cell.fill = _XLSX_HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_thicken_right_border(ws, row: int, col: int):
    """Replace just the right edge of an already-styled cell with the
    heavier period-divider rule, keeping whatever left/top/bottom border
    that cell already carries (the full grid on a shaded line row, a
    plain rule on a group row, none at all on a running-total or blank
    separator row). Works on a merged-but-not-anchor cell too (a period's
    header date spans several columns, but only its own left-most cell
    holds the merge) — openpyxl allows setting style on those, just not
    `.value`."""
    cell = ws.cell(row=row, column=col)
    b = cell.border
    cell.border = Border(left=b.left, right=_XLSX_PERIOD_DIVIDER, top=b.top, bottom=b.bottom)


def _xlsx_variance_coloring(ws, col: int, row_start: int, row_end: int):
    """Red text for a negative variance, green for a positive one, over
    one column's whole data range — applied to both the plain Variance
    column and the % Variance column, single-range or per-period alike.
    A zero (rendered as the money/pct format's own "-") gets neither."""
    cell_range = f"{get_column_letter(col)}{row_start}:{get_column_letter(col)}{row_end}"
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["0"], font=_XLSX_NEG_FONT))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=["0"], font=_XLSX_POS_FONT))


def _xlsx_variance_formulas(base_cell: str, compare_cell: str, pct_of_base: bool) -> tuple[str, str]:
    """Live Excel formulas for a Variance/% Variance pair, replicating
    _variance_amount()/_pct_variance()'s own two conventions exactly —
    default: base-minus-compare, % of compare (the standard percent-
    change reading, base as "new"); pct_of_base ("Flip variance
    direction" checked): compare-minus-base, % of base instead (see
    _pct_variance's own docstring for the full reasoning). Safe to derive
    live, unlike a group/subtotal row's own base/compare figures (see the
    Income Statement route's docstring on why those stay literals) —
    each formula only ever references the two cells already sitting in
    the same row, never a range that could double-count a rolled-up
    tree. IF(...,"",...) mirrors _pct_variance() returning None (blank,
    not a literal 0%) when there's nothing to divide by."""
    if pct_of_base:
        return (f"={compare_cell}-{base_cell}",
                f'=IF({base_cell}=0,"",ROUND(({compare_cell}-{base_cell})/ABS({base_cell})*100,1))')
    return (f"={base_cell}-{compare_cell}",
            f'=IF({compare_cell}=0,"",ROUND(({base_cell}-{compare_cell})/ABS({compare_cell})*100,1))')


def _xlsx_sum_formula(plus_cells: list[str], minus_cells: list[str] = ()) -> str:
    """A live formula adding/subtracting specific, individually-named
    cells — e.g. "=C6+C20-C34" — never a row range, so it stays safe
    regardless of how deep the tree under any one of those cells goes:
    each cell named here is a group's own root row, which (same
    reasoning as the "group" row style itself) already carries that
    subtree's full rolled-up total. Used for Income Statement's "Net
    income after X" running rows, each one just Income's root row(s)
    minus every expense group's root row seen so far. Falls back to a
    literal 0 rather than a bare "=" (not a valid formula) when both
    lists are empty — a report with no income and no expense rows at
    all, which shouldn't happen in practice but shouldn't crash either."""
    if not plus_cells and not minus_cells:
        return 0
    return "=" + "+".join(plus_cells) + "".join(f"-{c}" for c in minus_cells)


_XLSX_ROW_FONTS = {
    "group": _XLSX_GROUP_FONT, "line": _XLSX_LINE_FONT, "running": _XLSX_RUNNING_FONT,
    "subtotal": _XLSX_SUBTOTAL_FONT, "grand": _XLSX_GRAND_FONT, "grand_bad": _XLSX_GRAND_FONT_BAD,
}


def _xlsx_data_row(ws, row: int, label_cols: list, value_cols: list, style: str, depth: int = 0):
    """Write one report row. `label_cols` is [(col, text), ...] for the
    leading text columns (code, account name); `value_cols` is
    [(col, value, number_format), ...] for the money/percent columns.
    `style` picks one of six treatments (_XLSX_ROW_FONTS above names the
    font for each):
    - "group": a section's own top-level account, or a bare section-title
      row when `value_cols` is empty (Trial Balance/Balance Sheet's own
      "Assets"/"Liabilities" headers) — bold, ruled underneath. For a
      real account row, that figure already *is* the section's total,
      since it's the root of the rolled-up tree — see the Income
      Statement route's own docstring on why this replaced a separate
      "Total X" row.
    - "line": a plain account row — normal weight, its value cells shaded
      *and* fully gridded, matching the reference workbook's own "these
      are the numbers you'd read down a column" treatment.
    - "running": a running-total row like "Net income after Taxes" —
      italic, unshaded, unruled.
    - "subtotal": a rolled-up figure across more than one top-level
      account in the same section (Trial Balance's own per-type subtotal
      when a type has multiple roots) — semi-bold, muted, ruled, matching
      style.css's tr.subtotal exactly.
    - "grand"/"grand_bad": the report's own bottom-line total — bold,
      with the accountant's double-rule under the value cells (style.css's
      tr.grand), red instead of ink for "grand_bad" when that total
      doesn't actually balance/tie out.
    `depth` indents an account name under its parent, same meaning as the
    HTML report's own chevrons."""
    font = _XLSX_ROW_FONTS[style]
    for col, text in label_cols:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = font
        if depth and col != label_cols[0][0]:
            cell.alignment = Alignment(indent=depth)
        if style in ("group", "subtotal"):
            cell.border = _XLSX_BOTTOM_BORDER
    for col, value, number_format in value_cols:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.number_format = number_format
        if style in ("group", "subtotal"):
            cell.border = _XLSX_BOTTOM_BORDER
        elif style == "line":
            cell.fill = _XLSX_LINE_FILL
            cell.border = _XLSX_LINE_BORDER
        elif style == "grand":
            cell.border = _XLSX_GRAND_BORDER
        elif style == "grand_bad":
            cell.border = _XLSX_GRAND_BORDER_BAD


def xlsx_response(wb: Workbook, filename: str) -> Response:
    """Wrap a finished Workbook as a download — the XLSX counterpart to
    csv_response() above. No BOM/codepage concerns here (XLSX is a real
    zip container, not a bare text stream), so this is just a stream and
    a content type."""
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@app.get("/login")
def login_page(request: Request, err: str = None):
    if auth.get_session(request.cookies.get(auth.SESSION_COOKIE)):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(request, "login.html", {"err": err})


@app.post("/login")
def login_submit(username: str = Form(...), password: str = Form(...),
                 remember: str = Form(None)):
    username = username.strip().lower()
    if auth.is_rate_limited(username):
        return flash_redirect(
            "/login", err="Too many failed attempts — wait a few minutes and try again")
    row = q1("SELECT id, password_hash, is_active FROM users WHERE username = %s",
             (username,))
    if not row or not row["is_active"] or not auth.verify_password(password, row["password_hash"]):
        auth.record_failed_login(username)
        return flash_redirect("/login", err="Invalid username or password")
    auth.clear_failed_logins(username)
    token = auth.create_session(row["id"])
    resp = RedirectResponse("/", status_code=303)
    # The session itself is good for SESSION_TTL either way (see
    # auth.create_session) — "remember me" only decides whether the
    # *cookie* survives closing the browser. Unchecked, no max_age at all
    # makes it a session cookie the browser drops on its own; checked, it
    # gets an explicit lifetime matching the session behind it.
    resp.set_cookie(auth.SESSION_COOKIE, token, httponly=True, samesite="lax",
                    secure=auth.COOKIE_SECURE,
                    max_age=int(auth.SESSION_TTL.total_seconds()) if remember is not None else None)
    return resp


@app.post("/logout")
def logout(request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
    except ValueError:
        pass  # worst case of a bad token here is a no-op logout; just proceed
    user = auth.current_user(request)
    if user:
        auth.delete_session(user["token"])
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie(auth.SESSION_COOKIE)
    return resp


# ---------------------------------------------------------------------------
# Help — static reference content, one section per screen (see help.html).
# Keeps the explanatory prose that used to sit atop every page in one place
# instead of repeated inline; nothing here is dynamic, so no route params.
# ---------------------------------------------------------------------------
@app.get("/help")
def help_page(request: Request):
    return templates.TemplateResponse(request, "help.html", {"nav": "help"})


# ---------------------------------------------------------------------------
# User settings — username/password, and the theme picker (see settings.html)
# ---------------------------------------------------------------------------
USERNAME_PATTERN = re.compile(r"^[a-z0-9_.-]{3,32}$")


@app.get("/settings")
def settings_page(request: Request, ok: str = None, err: str = None):
    return templates.TemplateResponse(request, "settings.html", {
        "nav": "settings", "ok": ok, "err": err,
    })


@app.get("/settings/account")
def account_page(request: Request, ok: str = None, err: str = None):
    return templates.TemplateResponse(request, "account.html", {
        "nav": "settings", "ok": ok, "err": err,
    })


# postwarden_bi (SPEC.md decision 14) is a
# fixed, hardcoded-password role by design — same tradeoff docker-compose.yml
# already makes for the app's own postwarden/postwarden login: a real per-instance
# secret would need a place to be generated and stored, and this project has
# exactly one of those (Postgres itself). Host/port are the only things that
# actually vary per install, so those are the only two read from the
# environment; POSTWARDEN_BI_PORT is purely informational, see its
# docker-compose.yml comment.
BI_DB = "postwarden"
BI_USER = "postwarden_bi"
BI_OBJECTS = [
    ("v_dim_account", "Account dimension — hierarchy path, depth, normal side"),
    ("v_fact_lines", "Fact table — one row per journal line, fully denormalized"),
    ("v_dim_date", "Date dimension, 2020–2035"),
    ("v_monthly_activity", "v_fact_lines pre-aggregated to account × month × scenario"),
    ("fn_trial_balance('ACTUAL', '2026-08-31')", "Trial balance at any date, any scenario"),
]


@app.get("/settings/connect-bi")
def connect_bi_page(request: Request):
    return templates.TemplateResponse(request, "connect_bi.html", {
        "nav": "settings",
        "bi_host": request.url.hostname,
        "bi_port": os.environ.get("POSTWARDEN_BI_PORT", "5432"),
        "bi_db": BI_DB,
        "bi_user": BI_USER,
        "bi_objects": BI_OBJECTS,
    })


@app.get("/settings/connect-bi/download.pbids")
def connect_bi_pbids(request: Request):
    """A Power BI Data Source file — double-clicking it in Power BI Desktop
    opens straight to a PostgreSQL connection dialog pre-filled with this
    instance's host/port/database, nothing to type by hand. No credentials
    in it: Power BI still prompts for the postwarden_bi password itself, the same
    as it would connecting manually. See https://learn.microsoft.com/power-bi/connect-data/desktop-data-sources#pbids-files"""
    pbids = {
        "version": "0.1",
        "connections": [{
            "details": {"protocol": "postgresql", "address": {
                "server": f'{request.url.hostname}:{os.environ.get("POSTWARDEN_BI_PORT", "5432")}',
                "database": BI_DB,
            }},
            "mode": "Import",
        }],
    }
    return Response(json.dumps(pbids, indent=2), media_type="application/json", headers={
        "Content-Disposition": 'attachment; filename="PostWarden.pbids"'})


@app.post("/settings/username")
def change_username(request: Request, username: str = Form(...),
                    csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        username = username.strip().lower()
        if not USERNAME_PATTERN.match(username):
            raise ValueError(
                "Username must be 3-32 characters: lowercase letters, numbers, "
                "_ . or - only")
        user_id = auth.current_user(request)["user_id"]
        with tx() as cur:
            cur.execute("UPDATE users SET username = %s WHERE id = %s",
                       (username, user_id))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/settings/account", err=msg)
    return flash_redirect("/settings/account", ok=f"Username changed to {username!r}")


@app.post("/settings/password")
def change_password(request: Request, current_password: str = Form(...),
                    new_password: str = Form(...), confirm_password: str = Form(...),
                    csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        session = auth.current_user(request)
        row = q1("SELECT password_hash FROM users WHERE id = %s", (session["user_id"],))
        if not row or not auth.verify_password(current_password, row["password_hash"]):
            raise ValueError("Current password is incorrect")
        if new_password != confirm_password:
            raise ValueError("New password and confirmation don't match")
        if len(new_password) < 8:
            raise ValueError("New password must be at least 8 characters")
        with tx() as cur:
            cur.execute("UPDATE users SET password_hash = %s WHERE id = %s",
                       (auth.hash_password(new_password), session["user_id"]))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/settings/account", err=msg)
    # Same as the CLI's reset-password: revoke every session for this user,
    # including the current one — logging back in with the new password is
    # itself the confirmation that it was set correctly.
    auth.delete_all_sessions_for_user(session["user_id"])
    resp = flash_redirect("/login", ok="Password changed — please log in again")
    resp.delete_cookie(auth.SESSION_COOKIE)
    return resp


# ---------------------------------------------------------------------------
# Dashboard — the landing page. Always ACTUAL: "how are my real finances
# doing," not a scenario picker. Trial Balance/Income Statement/Balance
# Sheet/Variance are the tools for anything more specific than that.
# ---------------------------------------------------------------------------
@app.get("/")
def dashboard(request: Request):
    today = date.today()
    month_start = today.replace(day=1).isoformat()
    today_iso = today.isoformat()

    as_of_rows = q("SELECT * FROM fn_trial_balance(%s, %s)", ("ACTUAL", today_iso))
    net_worth = (sum(r["net"] for r in as_of_rows if r["acct_type"] == "asset")
                + sum(r["net"] for r in as_of_rows if r["acct_type"] == "liability"))

    mtd_rows = q("SELECT * FROM fn_trial_balance(%s, %s, %s)",
                ("ACTUAL", today_iso, month_start))
    mtd_income = -sum(r["net"] for r in mtd_rows if r["acct_type"] == "income")
    mtd_expenses = sum(r["net"] for r in mtd_rows if r["acct_type"] == "expense")

    recent = q("""
        SELECT e.id, e.entry_date, e.description, p.name AS payee_name,
               (SELECT COALESCE(SUM(l.debit), 0) FROM journal_lines l
                 WHERE l.entry_id = e.id) AS total_debits
          FROM journal_entries e
          JOIN scenarios s ON s.id = e.scenario_id
          LEFT JOIN payees p ON p.id = e.payee_id
         WHERE s.code = 'ACTUAL'
         ORDER BY e.entry_date DESC, e.seq DESC
         LIMIT 8""")

    # The widget's "Salary Income → Cash" label — which account(s) the
    # money came from and which it landed in. Batched by entry id (same
    # shape as the Journal page's own lines_by_entry) rather than one
    # query per row. Collapses to "multiple" on either side rather than
    # listing every account, since this is a compact recent-activity row,
    # not the full entry — see Journal's own expand-to-see-lines for that.
    recent_ids = [r["id"] for r in recent]
    debit_names, credit_names = {}, {}
    if recent_ids:
        for ln in q("""SELECT l.entry_id, a.name AS account_name, l.debit, l.credit
                         FROM journal_lines l
                         JOIN accounts a ON a.id = l.account_id
                        WHERE l.entry_id = ANY(%s)""", (recent_ids,)):
            bucket = debit_names if ln["debit"] > 0 else credit_names
            bucket.setdefault(ln["entry_id"], set()).add(ln["account_name"])
    def flow_side(names):
        # "multiple" is a stand-in label, not an account name, so it gets
        # italicized to read as a placeholder rather than a real account —
        # everything else is a real (user-controlled) account name and
        # goes through escape() rather than straight into the Markup.
        if len(names) != 1:
            return Markup("<em>multiple</em>")
        return escape(next(iter(names)))

    for r in recent:
        debits = debit_names.get(r["id"], set())
        credits = credit_names.get(r["id"], set())
        r["flow"] = Markup("{} → {}").format(flow_side(credits), flow_side(debits))

    # Upcoming transactions — every *active* schedule, soonest first. Never
    # shows one that's actually due: the auth middleware's
    # materialize_due_schedules() call (see there) already turns a due
    # occurrence into a real Staging entry and pushes next_date past today
    # before this route ever runs, so every row here is a real future date.
    upcoming = q("""
        SELECT se.id, se.next_date, se.description, p.name AS payee_name,
               (SELECT COALESCE(SUM(l.debit), 0) FROM scheduled_entry_lines l
                 WHERE l.scheduled_entry_id = se.id) AS total_debits
          FROM scheduled_entries se
          LEFT JOIN payees p ON p.id = se.payee_id
         WHERE se.is_active
         ORDER BY se.next_date, se.id
         LIMIT 8""")

    upcoming_ids = [r["id"] for r in upcoming]
    debit_names2, credit_names2 = {}, {}
    if upcoming_ids:
        for ln in q("""SELECT l.scheduled_entry_id AS entry_id, a.name AS account_name, l.debit, l.credit
                         FROM scheduled_entry_lines l
                         JOIN accounts a ON a.id = l.account_id
                        WHERE l.scheduled_entry_id = ANY(%s)""", (upcoming_ids,)):
            bucket = debit_names2 if ln["debit"] > 0 else credit_names2
            bucket.setdefault(ln["entry_id"], set()).add(ln["account_name"])
    for r in upcoming:
        debits = debit_names2.get(r["id"], set())
        credits = credit_names2.get(r["id"], set())
        r["flow"] = Markup("{} → {}").format(flow_side(credits), flow_side(debits))

    pending, _ = pending_staging_entries()

    return templates.TemplateResponse(request, "dashboard.html", {
        "nav": "dashboard", "net_worth": net_worth,
        "mtd_income": mtd_income, "mtd_expenses": mtd_expenses,
        "mtd_net": mtd_income - mtd_expenses,
        "month_label": today.strftime("%B %Y"),
        "recent": recent, "upcoming": upcoming, "pending_count": len(pending),
        "today": today_iso,
    })


# ---------------------------------------------------------------------------
# Trial balance — cumulative "as of" for Assets/Liabilities/Equity, same as
# always (that's the whole point of a trial balance: verifying the *entire*
# ledger's debits equal its credits). Income/Expense account rows default
# to a *simulated monthly close*: each account shows only month-to-date
# activity, as if every prior month had actually been closed to Equity.
# The two synthetic Equity lines this implies:
#   "Current Year Earnings (Unclosed)" — this fiscal year's earnings not
#   already reflected in this month (i.e. everything since Jan 1 except
#   MTD), and
#   "Prior Year Earnings (Unclosed)" — every fiscal year before this one,
# together account for exactly the gap between MTD and the true lifetime
# total, so the page's overall debit=credit check still holds — nothing
# is hidden, just regrouped the way it would look after an actual monthly
# close, without ever posting one. `raw=1` turns the simulation off
# entirely and shows every account's true, unmodified cumulative balance
# instead — useful for auditing the real numbers underneath.
# ---------------------------------------------------------------------------
def _pnl_net(accounts: list[dict], balances: dict) -> float:
    """Combined Income-minus-Expense across a {account_id: net} balance
    map, sign-corrected so a positive result means real earnings (credit
    side of Equity)."""
    income = sum(balances.get(a["id"], 0) for a in accounts if a["account_type"] == "income")
    expense = sum(balances.get(a["id"], 0) for a in accounts if a["account_type"] == "expense")
    return -income - expense


def _earnings_row(name: str, amount, depth: int = 2) -> dict:
    return {"account_code": "", "account_name": name, "path": "", "depth": depth,
            "has_children": False, "debit_balance": max(-amount, 0), "credit_balance": max(amount, 0)}


def _build_account_tree(accounts: list[dict], balances_by_id: dict,
                        compare_by_id: dict = None) -> list[dict]:
    """The account forest (roots = accounts.parent_id IS NULL), each node
    carrying a "subtotal" that rolls up every descendant's own direct
    balance — the actual Trial Balance/Balance Sheet display figure.
    "net" stays each account's own direct postings only, same as
    fn_trial_balance always showed; "subtotal" is the new thing a summary
    account with subdivisions (e.g. "Current Assets"/"Long-term Assets"
    under "Assets") needed and never had.

    `compare_by_id` is optional — a second {account_id: net} map rolled
    up alongside the first into "compare_subtotal"/"compare_net", for
    Income Statement/Variance's own second-scenario column. A single
    tree this way drives both a plain report and a two-scenario
    comparison; callers that pass nothing get compare_subtotal fixed at
    0 for every node, which _flatten_tree's zero-check treats as "no
    override" — the exact same hide-if-zero behavior as before this
    parameter existed."""
    compare_by_id = compare_by_id or {}
    nodes = {}
    for a in accounts:
        nodes[a["id"]] = {
            "id": a["id"], "parent_id": a["parent_id"], "account_code": a["code"],
            # parent_path (not path) — every caller renders this right next
            # to account_name, so it must exclude the account's own name or
            # the leaf reads twice (see v_dim_account's comment in schema.sql).
            "account_name": a["name"], "path": a["parent_path"], "account_type": a["account_type"],
            "depth": a["depth"], "net": balances_by_id.get(a["id"], 0),
            "compare_net": compare_by_id.get(a["id"], 0), "children": [],
        }
    roots = []
    for a in accounts:
        node = nodes[a["id"]]
        parent = nodes.get(a["parent_id"])
        (parent["children"] if parent else roots).append(node)

    def rollup(node):
        total, compare_total = node["net"], node["compare_net"]
        for c in node["children"]:
            b, cm = rollup(c)
            total += b
            compare_total += cm
        node["subtotal"] = total
        node["compare_subtotal"] = compare_total
        node["debit_balance"] = max(total, 0)
        node["credit_balance"] = max(-total, 0)
        return total, compare_total
    for r in roots:
        rollup(r)
    return roots


def _flatten_tree(nodes: list[dict], zeros: bool) -> list[dict]:
    """Depth-first flatten for template rendering, dropping any node (and
    its whole subtree) whose rolled-up subtotal is zero on *both* sides
    (own and compare — a row that only moved in one of the two scenarios
    is still activity worth showing), unless `zeros` — the same "hide
    accounts with no activity" rule Trial Balance always applied, just
    against the rollup instead of each account's own balance now. Adds
    has_children counting only what survives that filter, so a summary
    account left childless by it doesn't render a collapse arrow with
    nothing behind it."""
    out = []
    for node in nodes:
        if not zeros and node["subtotal"] == 0 and node.get("compare_subtotal", 0) == 0:
            continue
        kept_children = _flatten_tree(node["children"], zeros)
        out.append({**node, "has_children": bool(kept_children)})
        out.extend(kept_children)
    return out


def _trial_balance_rows(scenario: str, as_of: str, zeros: int, raw: int = 0) -> dict:
    as_of_date = as_of or None
    as_of_dt = date.fromisoformat(as_of_date) if as_of_date else date.today()
    accounts = q("SELECT * FROM v_dim_account WHERE is_active ORDER BY sort_path")
    full_balances = {r["account_id"]: r["net"] for r in
                     q("SELECT * FROM fn_account_balances(%s, %s)", (scenario, as_of_date))}
    total_debits = sum(max(v, 0) for v in full_balances.values())
    total_credits = sum(max(-v, 0) for v in full_balances.values())

    def build_sections(balances_by_id: dict, extra_equity: list[dict]) -> list[dict]:
        roots = _build_account_tree(accounts, balances_by_id)
        grouped = []
        for t in ACCOUNT_TYPES:
            type_roots = [r for r in roots if r["account_type"] == t]
            extra = extra_equity if t == "equity" else []
            flat = _flatten_tree(type_roots, zeros)
            if flat or extra:
                grouped.append({
                    "type": t, "label": TYPE_LABELS[t], "rows": flat + extra,
                    "sub_debits": sum(r["debit_balance"] for r in type_roots + extra),
                    "sub_credits": sum(r["credit_balance"] for r in type_roots + extra),
                    "show_type_total": len(type_roots) > 1 or bool(extra),
                })
        return grouped

    if raw:
        grouped = build_sections(full_balances, [])
        return {"grouped": grouped, "total_debits": total_debits, "total_credits": total_credits,
                "in_balance": total_debits == total_credits}

    fy_start = date(as_of_dt.year, 1, 1).isoformat()
    month_start = date(as_of_dt.year, as_of_dt.month, 1).isoformat()
    fy_balances = {r["account_id"]: r["net"] for r in
                  q("SELECT * FROM fn_account_balances(%s, %s, %s)", (scenario, as_of_date, fy_start))}
    mtd_balances = {r["account_id"]: r["net"] for r in
                   q("SELECT * FROM fn_account_balances(%s, %s, %s)", (scenario, as_of_date, month_start))}

    all_time_earnings = _pnl_net(accounts, full_balances)
    fy_earnings = _pnl_net(accounts, fy_balances)
    mtd_earnings = _pnl_net(accounts, mtd_balances)
    prior_year_earnings = all_time_earnings - fy_earnings
    current_year_earnings = fy_earnings - mtd_earnings

    merged_balances = {a["id"]: (mtd_balances.get(a["id"], 0)
                                 if a["account_type"] in ("income", "expense")
                                 else full_balances.get(a["id"], 0))
                       for a in accounts}
    extra_equity = []
    if zeros or current_year_earnings != 0 or prior_year_earnings != 0:
        extra_equity = [
            _earnings_row("Current Year Earnings (Unclosed)", current_year_earnings),
            _earnings_row("Prior Year Earnings (Unclosed)", prior_year_earnings),
        ]
    grouped = build_sections(merged_balances, extra_equity)

    return {"grouped": grouped, "total_debits": total_debits, "total_credits": total_credits,
            "in_balance": total_debits == total_credits, "fy_start": fy_start,
            "month_start": month_start}


@app.get("/trial-balance")
def trial_balance(request: Request, scenario: str = "ACTUAL",
                  as_of: str = None, zeros: int = 0, raw: int = 0):
    result = _trial_balance_rows(scenario, as_of, zeros, raw)
    return templates.TemplateResponse(request, "trial_balance.html", {
        "nav": "tb", "scenario": scenario, "as_of": as_of or "", "zeros": zeros, "raw": raw,
        "scenarios": scenarios_all(), "today": date.today().isoformat(), **result,
    })


def _trial_balance_export_filename(scenario: str, as_of: str, raw: int, ext: str) -> str:
    name = f"postwarden-trial-balance-{scenario}"
    if as_of:
        name += f"_{as_of}"
    if raw:
        name += "_raw"
    return f"{name}.{ext}"


@app.get("/export/trial-balance.csv")
def trial_balance_export_csv(scenario: str = "ACTUAL", as_of: str = None,
                             zeros: int = 0, raw: int = 0):
    result = _trial_balance_rows(scenario, as_of, zeros, raw)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Code", "Account", "Path", "Debit", "Credit"])
    for g in result["grouped"]:
        for r in g["rows"]:
            w.writerow([r["account_code"], r["account_name"], r["path"],
                       r["debit_balance"] or "", r["credit_balance"] or ""])
    return csv_response(buf, _trial_balance_export_filename(scenario, as_of, raw, "csv"))


@app.get("/export/trial-balance.xlsx")
def trial_balance_export_xlsx(scenario: str = "ACTUAL", as_of: str = None,
                              zeros: int = 0, raw: int = 0):
    """XLSX counterpart to trial_balance_export_csv() above, plus the
    section/subtotal/grand-total structure the CSV leaves out (it's a
    plain account list — no type-head, no per-type subtotal, no balance
    check) but result["grouped"] already carries: a bold section-title
    row per account type (g["label"]), that type's own subtotal row only
    when it actually sums more than one top-level account (g["show_type_
    total"] — a single-root type's own root row already *is* the total,
    same reasoning as Income Statement's own group-row treatment), and a
    bottom "In balance"/"Out of balance" row with the accountant's
    double-rule, red instead of ink when it doesn't."""
    result = _trial_balance_rows(scenario, as_of, zeros, raw)
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    subtitle = f"{scenario} · {'As of ' + as_of if as_of else 'Through today'}"
    if not raw:
        subtitle += " · simulated monthly close"
    ws.cell(row=1, column=1, value="Trial Balance").font = _XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Debit", "Credit"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    _xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, debit, credit, style="line"):
        value_cols = [(3, debit or None, _XLSX_MONEY_FMT), (4, credit or None, _XLSX_MONEY_FMT)]
        _xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

    r = data_start
    for g in result["grouped"]:
        _xlsx_data_row(ws, r, [(1, ""), (2, g["label"])], [], style="group")
        r += 1
        for line in g["rows"]:
            row(r, line["account_code"], line["account_name"], line.get("depth", 2),
                line["debit_balance"], line["credit_balance"],
                style="group" if line.get("depth") == 1 else "line")
            r += 1
        if g["show_type_total"]:
            row(r, "", f"{g['label']} subtotal", 1, g["sub_debits"], g["sub_credits"], style="subtotal")
            r += 1
    grand_style = "grand" if result["in_balance"] else "grand_bad"
    label = "In balance" if result["in_balance"] else "Out of balance (this scenario allows single-sided entries)"
    row(r, "", label, 1, result["total_debits"], result["total_credits"], style=grand_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx_response(wb, _trial_balance_export_filename(scenario, as_of, raw, "xlsx"))


# ---------------------------------------------------------------------------
# Income statement (P&L) — Income and Expense only, always period-ranged:
# these are flow accounts, so "as of a date" doesn't mean anything for them
# the way it does for Assets/Liabilities/Equity. Defaults to month-to-date.
# No physical period-close needed — see fn_trial_balance's p_from: the
# range is just a WHERE clause, not a journal entry zeroing anything out.
#
# Expense rows are broken into their own section per *top-level* expense
# account rather than one flat "Expenses" bucket — a user who adds a
# second top-level expense account (e.g. "6000 Other expenses" alongside
# the original "5000 Expenses") gets a waterfall: each section's own
# subtotal, followed by a running-total line reflecting everything
# subtracted so far, in account-code order. Every running line but the
# very last is labelled "Net income after {group}" so a waterfall of
# several doesn't read as the same "Net income" repeated; the last one
# always is, since it's the actual bottom line. With just the one usual
# top-level expense account this collapses to exactly the old single
# Expenses-section-then-Net-income layout.
#
# Optionally compares a second ("compare") scenario alongside the primary
# one — e.g. Actual vs. a Budget scenario — with a % variance and each
# subtotal/net line's share of total income, so a budget scenario reads
# next to the real numbers instead of needing the separate Variance page.
#
# Each top-level group's own rows come from a real parent/child tree
# (_build_account_tree/_flatten_tree, same machinery Trial Balance/
# Balance Sheet use), not a flat "every account with nonzero activity"
# list — so a mid-tree summary account (e.g. a "Housing" under
# "Expenses") gets its own collapsible row via report-tree.js instead of
# disappearing into a flat list of leaves, and `zeros` shows every
# account down to zero-balance leaves, same checkbox/meaning as Trial
# Balance's.
# ---------------------------------------------------------------------------
def _pct_variance(base, compare_val, pct_of_base: bool = False):
    """% variance between `base` (a report's own primary scenario figure
    — "Scenario" on Income Statement, "Baseline" on Variance, "Actual" on
    the Budget grid) and `compare_val` (whatever it's being measured
    against — "Compare to"/"Budgeted") — two conventions, user-toggleable
    per report via a "Flip variance direction" checkbox next to Hide
    zero balances (Income Statement, Variance, Budget Grid all share this
    one flag — see each route's own `pct_of_base` query param; the
    parameter name predates this docstring and stayed as-is since it's
    also a public, bookmarkable query string — only what it *does*
    changed here).

    Default (pct_of_base=False, unchecked): the standard percent-change
    reading, (new - old) / old, with `base` as the "new" figure and
    `compare_val` as the "old" one being measured against — (base -
    compare_val) / compare_val. "actual came in 12% ahead of budget."

    Checked (pct_of_base=True): the same reading with the two swapped —
    `compare_val` as "new", `base` as "old" — (compare_val - base) /
    base. "budget came in 12% ahead of actual."

    Both conventions divide by whichever figure is playing "old" in that
    state, not always the same one — so this takes the toggle as an
    explicit argument rather than being a caller-side negation. None
    (not 0%) when there's nothing to divide by."""
    if pct_of_base:
        if not base:
            return None
        return round((compare_val - base) / abs(base) * 100, 1)
    if not compare_val:
        return None
    return round((base - compare_val) / abs(compare_val) * 100, 1)


def _variance_amount(base, compare_val, pct_of_base: bool = False):
    """The plain-currency counterpart to _pct_variance() above — same
    toggle, same two conventions, kept as its own function (not baked
    into _pct_variance's return) since every call site needs both the
    dollar figure and the percentage rendered side by side, not one
    derived from the other. Default: base - compare_val (actual minus
    budget — positive when actual is ahead). Checked: compare_val - base
    (budget minus actual) — the numerator flips right along with which
    side of the % the toggle picks, so the sign of the dollar variance
    always agrees with whichever percentage is showing next to it."""
    return (compare_val - base) if pct_of_base else (base - compare_val)


def _pct_of(amount, total):
    if not total:
        return None
    return round(amount / total * 100, 1)


def _income_statement_groups(roots: list[dict], t: str, flip: bool, zeros: bool,
                             pct_of_base: bool = False) -> list[dict]:
    """One group per top-level account of type `t` — multiple, for a
    second top-level expense account like "6000 Other" (see module
    comment). Each group's rows are that root's own _flatten_tree()
    output, so the root itself opens the group as a normal (possibly
    collapsible) row rather than existing only as the header text above
    it, and any zero-balance root is dropped entirely unless `zeros` —
    same "no activity anywhere in this group" hiding the old flat merge
    gave for free by simply never creating the group. `flip` sign-
    corrects credit-normal Income rows (net < 0 for real income) so
    every amount from here on reads as a plain positive figure in its
    "normal" direction. `pct_of_base` — see _pct_variance()'s own
    comment — is this scenario's own net (base_net/base_subtotal), not
    the compare scenario's."""
    sign = -1 if flip else 1

    # sign * x, but a zero-balance row never comes out as a *negative*
    # zero (-1 * 0) — same value, but %.2f-style formatting renders that
    # as a confusing "-0.00", and unlike the HTML template's money()
    # filter (which guards against this itself), a raw CSV export writes
    # whatever numeric value it's handed with no such rescue.
    def signed(x):
        v = sign * x
        return abs(v) if v == 0 else v

    out = []
    for root in sorted((r for r in roots if r["account_type"] == t), key=lambda r: r["account_code"]):
        if not zeros and root["subtotal"] == 0 and root["compare_subtotal"] == 0:
            continue
        rows = _flatten_tree([root], zeros)
        for r in rows:
            r["base_net"] = signed(r["subtotal"])
            r["compare_net"] = signed(r["compare_subtotal"])
            r["variance"] = _variance_amount(r["base_net"], r["compare_net"], pct_of_base)
            r["pct_variance"] = _pct_variance(r["base_net"], r["compare_net"], pct_of_base)
        out.append({
            "name": root["account_name"], "rows": rows,
            "base_subtotal": signed(root["subtotal"]), "compare_subtotal": signed(root["compare_subtotal"]),
        })
    for g in out:
        g["variance"] = _variance_amount(g["base_subtotal"], g["compare_subtotal"], pct_of_base)
        g["pct_variance"] = _pct_variance(g["base_subtotal"], g["compare_subtotal"], pct_of_base)
    return out


def _income_statement_balances(scenario_code: str, accounts_by_id: dict,
                               date_to_v: str | None, date_from_v: str | None) -> dict:
    """Account balances for one side of the Income Statement (base or
    compare) — journal-based via fn_account_balances for a normal
    scenario, same as every other report, but an income-statement-only
    scenario (Budget Grid's own scenario type) never takes a journal
    entry at all (fn_income_statement_only_guard blocks it), so that
    path always came back empty — the Compare column silently showing
    nothing was actually correct given what it was querying, just not
    what "compare to a budget scenario" should mean. Its numbers live in
    budget_lines instead, one row per (account, month); summed across
    every month the report's date range touches, since Income Statement
    (unlike Budget Grid) covers an arbitrary range, not one month at a
    time. budget_lines.amount is a plain positive target with no debit/
    credit sign to it, flipped here into the same journal sign
    convention fn_account_balances returns (income negative) — the
    `sign` flip _income_statement_groups() applies next expects that
    convention from either source equally; same flip _budget_rows() does
    in the opposite direction for its own Actual column."""
    scen = q1("SELECT id, income_statement_only FROM scenarios WHERE code = %s", (scenario_code,))
    if not scen:
        return {}
    if not scen["income_statement_only"]:
        return {r["account_id"]: r["net"] for r in
                q("SELECT * FROM fn_account_balances(%s, %s, %s)", (scenario_code, date_to_v, date_from_v))}
    where, params = ["scenario_id = %s"], [scen["id"]]
    if date_from_v:
        where.append("period_month >= date_trunc('month', %s::date)")
        params.append(date_from_v)
    if date_to_v:
        where.append("period_month <= date_trunc('month', %s::date)")
        params.append(date_to_v)
    rows = q(f"""SELECT account_id, SUM(amount) AS amt FROM budget_lines
                  WHERE {' AND '.join(where)} GROUP BY account_id""", params)
    return {
        r["account_id"]: (-1 if accounts_by_id.get(r["account_id"], {}).get("account_type") == "income" else 1) * r["amt"]
        for r in rows
    }


def _income_statement_rows(scenario: str, date_from: str, date_to: str,
                           compare: str = "", zeros: int = 0,
                           pct_of_base: bool = False) -> dict:
    date_to_v, date_from_v = date_to or None, date_from or None
    accounts = q("""SELECT * FROM v_dim_account
                     WHERE is_active AND account_type IN ('income', 'expense')
                     ORDER BY sort_path""")
    accounts_by_id = {a["id"]: a for a in accounts}
    base_by_id = _income_statement_balances(scenario, accounts_by_id, date_to_v, date_from_v)
    compare_by_id = (_income_statement_balances(compare, accounts_by_id, date_to_v, date_from_v)
                     if compare else {})
    roots = _build_account_tree(accounts, base_by_id, compare_by_id)
    income_groups = _income_statement_groups(roots, "income", flip=True, zeros=zeros, pct_of_base=pct_of_base)
    expense_groups = _income_statement_groups(roots, "expense", flip=False, zeros=zeros, pct_of_base=pct_of_base)

    total_base_income = sum(g["base_subtotal"] for g in income_groups)
    total_compare_income = sum(g["compare_subtotal"] for g in income_groups)
    income_variance_amount = _variance_amount(total_base_income, total_compare_income, pct_of_base)
    income_variance = _pct_variance(total_base_income, total_compare_income, pct_of_base)

    base_running, compare_running = total_base_income, total_compare_income
    for g in expense_groups:
        base_running -= g["base_subtotal"]
        compare_running -= g["compare_subtotal"]
        g["base_running_after"] = base_running
        g["compare_running_after"] = compare_running
        g["running_variance"] = _variance_amount(base_running, compare_running, pct_of_base)
        g["running_pct_variance"] = _pct_variance(base_running, compare_running, pct_of_base)
        g["base_pct_of_income"] = _pct_of(g["base_subtotal"], total_base_income)
        g["compare_pct_of_income"] = _pct_of(g["compare_subtotal"], total_compare_income)
        g["base_running_pct_of_income"] = _pct_of(base_running, total_base_income)
        g["compare_running_pct_of_income"] = _pct_of(compare_running, total_compare_income)

    net_income = base_running if expense_groups else total_base_income
    compare_net_income = compare_running if expense_groups else total_compare_income
    return {
        "income_groups": income_groups, "expense_groups": expense_groups,
        "total_base_income": total_base_income, "total_compare_income": total_compare_income,
        "income_variance_amount": income_variance_amount, "income_variance": income_variance,
        "net_income": net_income, "compare_net_income": compare_net_income,
        "net_income_variance_amount": _variance_amount(net_income, compare_net_income, pct_of_base),
        "net_income_variance": _pct_variance(net_income, compare_net_income, pct_of_base),
        "net_income_pct_of_income": _pct_of(net_income, total_base_income),
        "compare_net_income_pct_of_income": _pct_of(compare_net_income, total_compare_income),
        "has_compare": bool(compare),
    }


def _split_periods(date_from: str, date_to: str, split: str) -> list[dict]:
    """Breaks [date_from, date_to] into calendar-aligned sub-periods for
    Income Statement's Split view (see income_statement_page's own
    `split` param) — real calendar months/quarters/years, not even
    day-slicing. Each period is clipped to the requested range at both
    ends rather than expanded outward to a whole calendar period, so a
    custom range like Aug 15-Oct 3 split quarterly never totals days
    outside what date_from/date_to actually asked for; `partial` flags a
    clipped edge so the template can show the real covered span next to
    the calendar-period label instead of silently implying a full
    quarter. An unrecognized/empty `split` (or an inverted range) returns
    [] — the caller's own signal to fall back to the single-range report,
    same as `compare=""` already means "no comparison" elsewhere. Capped
    at 60 periods (5 years monthly) as a plain sanity limit — nothing
    about the feature needs it, it's just guarding against an accidental
    date range turning into thousands of one-day SQL round trips. Also
    returns [] for an empty date_from/date_to — unlike
    income_statement_page (which always defaults both before calling
    this), income_statement_export_csv still allows either blank,
    meaning "unbounded" to _income_statement_rows; there's no calendar
    period to align an open-ended range to, so Split silently falls back
    to the single-range export rather than raising on an empty string."""
    if not date_from or not date_to:
        return []
    start, end = date.fromisoformat(date_from), date.fromisoformat(date_to)
    if start > end:
        return []
    if split == "monthly":
        step, label, first = 1, (lambda d: d.strftime("%Y-%m")), (lambda d: date(d.year, d.month, 1))
    elif split == "quarterly":
        step = 3
        label = lambda d: f"{d.year}-Q{(d.month - 1) // 3 + 1}"
        first = lambda d: date(d.year, (d.month - 1) // 3 * 3 + 1, 1)
    elif split == "yearly":
        step, label, first = 12, (lambda d: str(d.year)), (lambda d: date(d.year, 1, 1))
    else:
        return []

    out = []
    cur = first(start)
    while cur <= end and len(out) < 60:
        total = cur.month - 1 + step
        nxt = date(cur.year + total // 12, total % 12 + 1, 1)
        period_end = nxt - timedelta(days=1)
        period_from, period_to = max(cur, start), min(period_end, end)
        out.append({
            "label": label(cur), "date_from": period_from.isoformat(), "date_to": period_to.isoformat(),
            "partial": period_from != cur or period_to != period_end,
        })
        cur = nxt
    return out


def _divide(v, n):
    return v / n if v is not None else None


def _scale_income_statement_result(result: dict, n: int) -> dict:
    """The Average column's own figures — the Totals column's exact
    figures divided by the real period count `n`. Safe/exact as a plain
    division rather than a fresh computation because every dollar amount
    here is additive across periods (Totals.base_net already equals
    sum(period.base_net for period in periods) — Split's periods
    partition the date range with no overlap or gap), and every
    percentage/ratio field (pct_variance, income_variance,
    *_pct_of_income, ...) is scale-invariant: dividing both the amount
    and whatever it's a ratio *of* by the same n leaves the ratio
    identical to what Totals already computed — recomputing it from
    the divided figures would just be extra work to land on the exact
    same number. So only the plain dollar fields get divided here; every
    percentage field is carried through unchanged via the row/group
    dict's own shallow copy."""
    def scale_row(r):
        return {**r, "base_net": _divide(r["base_net"], n), "compare_net": _divide(r["compare_net"], n),
                "variance": _divide(r["variance"], n)}

    def scale_group(g, is_expense):
        g2 = {**g, "rows": [scale_row(r) for r in g["rows"]],
              "base_subtotal": _divide(g["base_subtotal"], n), "compare_subtotal": _divide(g["compare_subtotal"], n),
              "variance": _divide(g["variance"], n)}
        if is_expense:
            g2["base_running_after"] = _divide(g["base_running_after"], n)
            g2["compare_running_after"] = _divide(g["compare_running_after"], n)
            g2["running_variance"] = _divide(g["running_variance"], n)
        return g2

    return {
        **result,
        "income_groups": [scale_group(g, False) for g in result["income_groups"]],
        "expense_groups": [scale_group(g, True) for g in result["expense_groups"]],
        "total_base_income": _divide(result["total_base_income"], n),
        "total_compare_income": _divide(result["total_compare_income"], n),
        "income_variance_amount": _divide(result["income_variance_amount"], n),
        "net_income": _divide(result["net_income"], n),
        "compare_net_income": _divide(result["compare_net_income"], n),
        "net_income_variance_amount": _divide(result["net_income_variance_amount"], n),
    }


def _income_statement_matrix(scenario: str, periods: list[dict], date_from: str, date_to: str,
                             compare: str = "", zeros: int = 0, pct_of_base: bool = False) -> dict:
    """Split-view counterpart to _income_statement_rows() above — one
    column group per Split period instead of one range. A thin wrapper
    around that same single-period function rather than a parallel
    calculation: every period gets its own full _income_statement_rows()
    call with `zeros` forced on, which guarantees every account row/group
    exists in every period, aligned by account id — a plain lookup merge
    from there, with no risk of August ending up with a different set of
    rows than September because one had a zero-balance account the other
    didn't.

    A separate "combined activity" tree (the same _build_account_tree/
    _income_statement_groups machinery the single-period report already
    uses) decides which rows/groups actually render under the *real*
    `zeros` flag — fed the sum of |base_net| and |compare_net| across
    every real period (the Totals column below deliberately never
    contributes to this — see there), so a row shows if it had activity
    in *any* period and hides only if it was zero everywhere, the same
    meaning "show zero balances" already has today, just extended across
    the whole matrix instead of one range. The scaffold tree's own
    base_net/compare_net figures are otherwise meaningless (a sum of
    absolute values, not a real total) — every row/group below gets its
    *real* per-period numbers overlaid from `per_period` right after,
    keyed by account id either way (a group's own id is its root
    account's — rows[0]).

    A trailing "Totals" column — the same whole-range figures the
    unsplit report would show for this exact scenario/date_from/date_to
    — is appended after the real periods, same shape as any other period
    (its own `_income_statement_rows()` call, zeros forced on) so the
    template's own `{% for p in periods %}` renders it with zero special
    casing. Computed *after* the scaffold's zero/activity union above and
    never folded into it — Totals only ever restates rows the real
    periods already decided to show, so it never needs a vote of its
    own, and giving it one would be redundant at best (any real period
    with activity already makes the union nonzero) and risked at worst
    (a row that's genuinely zero in every real period but somehow
    nonzero in Totals shouldn't be able to happen, but "never asked
    the question" is a stronger guarantee than "computed the same
    answer twice"). Average follows right after Totals, same treatment
    — see _scale_income_statement_result()."""
    accounts = q("""SELECT * FROM v_dim_account
                     WHERE is_active AND account_type IN ('income', 'expense')
                     ORDER BY sort_path""")
    per_period = [
        _income_statement_rows(scenario, p["date_from"], p["date_to"], compare, zeros=1, pct_of_base=pct_of_base)
        for p in periods
    ]

    combined_base, combined_compare, period_rows_by_id = {}, {}, []
    for p in per_period:
        rows_by_id = {}
        for g in p["income_groups"] + p["expense_groups"]:
            for r in g["rows"]:
                rows_by_id[r["id"]] = r
                combined_base[r["id"]] = combined_base.get(r["id"], 0) + abs(r["base_net"])
                combined_compare[r["id"]] = combined_compare.get(r["id"], 0) + abs(r["compare_net"])
        period_rows_by_id.append(rows_by_id)

    roots = _build_account_tree(accounts, combined_base, combined_compare)
    income_groups = _income_statement_groups(roots, "income", flip=True, zeros=zeros, pct_of_base=pct_of_base)
    expense_groups = _income_statement_groups(roots, "expense", flip=False, zeros=zeros, pct_of_base=pct_of_base)

    # The Totals column: whole-range figures, appended as one more
    # "period" after the union check above (see the docstring's own
    # paragraph on why it never contributes to it). Its label is a
    # plain, JS-free-safe default — period-picker.js rewrites it
    # client-side to match whatever the Period dropdown currently reads
    # ("This Quarter", "Custom range", ...) once it knows, since the
    # backend itself never learns which preset (if any) was picked, only
    # the date_from/date_to it resolved to (see that script's own
    # comment). CSV export, which has no client-side rewrite to lean on,
    # keeps the plain "Total" — a reasonable, still-correct spreadsheet
    # column header either way.
    totals_result = _income_statement_rows(scenario, date_from, date_to, compare, zeros=1, pct_of_base=pct_of_base)
    totals_rows_by_id = {r["id"]: r for g in totals_result["income_groups"] + totals_result["expense_groups"]
                         for r in g["rows"]}
    period_rows_by_id.append(totals_rows_by_id)

    # Average: Totals' own figures divided by the real period count — see
    # _scale_income_statement_result's own docstring for why a plain
    # division is exact here rather than an approximation. Same "just one
    # more period" treatment as Totals, appended right after it so the
    # template needs no special casing for this column either.
    average_result = _scale_income_statement_result(totals_result, len(periods))
    average_rows_by_id = {r["id"]: r for g in average_result["income_groups"] + average_result["expense_groups"]
                          for r in g["rows"]}
    period_rows_by_id.append(average_rows_by_id)

    all_periods = per_period + [totals_result, average_result]
    periods_with_total = periods + [
        {"label": "Total", "date_from": date_from, "date_to": date_to, "partial": False, "is_total": True},
        {"label": "Average", "date_from": date_from, "date_to": date_to, "partial": False, "is_average": True},
    ]

    # Matched by the group's own root-account id (its rows[0], same as any
    # other row) within its own income/expense list specifically — not by
    # name, and not a combined search across both lists: two top-level
    # accounts sharing a name (nothing stops a user naming both an income
    # and an expense root "Adjustments") would otherwise risk a group
    # matching the wrong one.
    for scaffold_groups, key in ((income_groups, "income_groups"), (expense_groups, "expense_groups")):
        for g in scaffold_groups:
            root_id = g["rows"][0]["id"]
            for r in g["rows"]:
                r["periods"] = [rows_by_id.get(r["id"], {}) for rows_by_id in period_rows_by_id]
            g["periods"] = [next(pg for pg in p[key] if pg["rows"][0]["id"] == root_id) for p in all_periods]

    return {
        "income_groups": income_groups, "expense_groups": expense_groups,
        # Each entry is a *whole* single-period _income_statement_rows()
        # result (total_base_income, net_income, ... — every top-level
        # figure the unsplit template reads straight off the result dict),
        # kept as-is rather than reshaped, so the split template's
        # "Total income"/final "Net income" rows read periods_totals[i].x
        # the same way the unsplit one reads x directly. The last two
        # entries are the Totals and Average columns' own results.
        "periods_totals": all_periods,
        "has_compare": bool(compare),
        "periods": periods_with_total,
    }


@app.get("/income-statement")
def income_statement_page(request: Request, scenario: str = "ACTUAL", compare: str = "",
                          date_from: str = "", date_to: str = "", zeros: int = 0,
                          pct_of_base: int = 0, split: str = ""):
    today = date.today()
    date_from = date_from or today.replace(day=1).isoformat()
    date_to = date_to or today.isoformat()
    periods = _split_periods(date_from, date_to, split)
    if periods:
        result = _income_statement_matrix(scenario, periods, date_from, date_to, compare, zeros, bool(pct_of_base))
    else:
        result = _income_statement_rows(scenario, date_from, date_to, compare, zeros, bool(pct_of_base))
        result["periods"] = periods
    return templates.TemplateResponse(request, "income_statement.html", {
        "nav": "income_statement", "scenarios": scenarios_all(),
        "scenario": scenario, "compare": compare, "date_from": date_from, "date_to": date_to,
        "zeros": zeros, "pct_of_base": pct_of_base, "split": split, "today": today.isoformat(), **result,
    })


def _income_statement_export_filename(scenario: str, compare: str, date_from: str, date_to: str,
                                      split: str, ext: str) -> str:
    """Shared by the CSV and XLSX exports below — a filename that
    actually says what's in it (scenario, compare scenario, date range,
    split granularity) instead of a bare postwarden-income-statement-
    ACTUAL.csv that looks identical no matter what the export covers, so
    two different downloads from this report don't collide — or silently
    overwrite each other — in a Downloads folder."""
    name = f"postwarden-income-statement-{scenario}"
    if compare:
        name += f"-vs-{compare}"
    if date_from and date_to:
        name += f"_{date_from}_to_{date_to}"
    elif date_from:
        name += f"_from_{date_from}"
    elif date_to:
        name += f"_through_{date_to}"
    if split:
        name += f"_{split}"
    return f"{name}.{ext}"


@app.get("/export/income-statement.csv")
def income_statement_export_csv(scenario: str = "ACTUAL", compare: str = "",
                                date_from: str = "", date_to: str = "", zeros: int = 0,
                                pct_of_base: int = 0, split: str = ""):
    periods = _split_periods(date_from, date_to, split)
    buf = io.StringIO()
    w = csv.writer(buf)

    if not periods:
        result = _income_statement_rows(scenario, date_from, date_to, compare, zeros, bool(pct_of_base))
        header = ["Section", "Code", "Account", "Path", scenario or "Amount"]
        if compare:
            header += ["Variance", "% variance", compare]
        w.writerow(header)

        def row(section, code, name, path, base, comp=None, variance=None, pct=None):
            line = [section, code, name, path, base]
            if compare:
                line += [variance if variance is not None else "",
                         pct if pct is not None else "",
                         comp if comp is not None else ""]
            w.writerow(line)

        for g in result["income_groups"]:
            for r in g["rows"]:
                row(g["name"], r["account_code"], r["account_name"], r["path"],
                    r["base_net"], r["compare_net"], r["variance"], r["pct_variance"])
        row("Income", "", "Total income", "", result["total_base_income"],
            result["total_compare_income"], result["income_variance_amount"], result["income_variance"])
        for i, g in enumerate(result["expense_groups"]):
            w.writerow([])
            for r in g["rows"]:
                row(g["name"], r["account_code"], r["account_name"], r["path"],
                    r["base_net"], r["compare_net"], r["variance"], r["pct_variance"])
            row(g["name"], "", f"Total {g['name']}", "", g["base_subtotal"],
                g["compare_subtotal"], g["variance"], g["pct_variance"])
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            row(g["name"], "", label, "", g["base_running_after"],
                g["compare_running_after"], g["running_variance"], g["running_pct_variance"])
        if not result["expense_groups"]:
            row("Income", "", "Net income", "", result["net_income"],
                result["compare_net_income"], result["net_income_variance_amount"], result["net_income_variance"])
        return csv_response(buf, _income_statement_export_filename(scenario, compare, date_from, date_to, split, "csv"))

    # Split view: one wide row per account, one group of columns per
    # period instead of one. Each period's own column group is prefixed
    # with that period's label so the header stays legible in a plain
    # spreadsheet with no merged/two-row header the way the HTML table
    # gets one — "2026-08 ACTUAL" reads fine as a single Excel column
    # header, "ACTUAL" repeated 3x with a separate period row above it
    # wouldn't survive a CSV round trip at all.
    result = _income_statement_matrix(scenario, periods, date_from, date_to, compare, zeros, bool(pct_of_base))
    header = ["Section", "Code", "Account", "Path"]
    for p in result["periods"]:  # real periods + the trailing Totals column
        header.append(f"{p['label']} {scenario}")
        if compare:
            header += [f"{p['label']} Variance", f"{p['label']} % variance", f"{p['label']} {compare}"]
    w.writerow(header)

    def row(section, code, name, path, period_values):
        line = [section, code, name, path]
        for v in period_values:
            line.append(v.get("base", ""))
            if compare:
                line += [v.get("variance", ""), v.get("pct", ""), v.get("comp", "")]
        w.writerow(line)

    for g in result["income_groups"]:
        for r in g["rows"]:
            row(g["name"], r["account_code"], r["account_name"], r["path"],
                [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                  "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in r["periods"]])
    row("Income", "", "Total income", "",
        [{"base": pt["total_base_income"], "comp": pt["total_compare_income"],
          "variance": pt["income_variance_amount"], "pct": pt["income_variance"]} for pt in result["periods_totals"]])
    for i, g in enumerate(result["expense_groups"]):
        w.writerow([])
        for r in g["rows"]:
            row(g["name"], r["account_code"], r["account_name"], r["path"],
                [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                  "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in r["periods"]])
        row(g["name"], "", f"Total {g['name']}", "",
            [{"base": gp["base_subtotal"], "comp": gp["compare_subtotal"],
              "variance": gp["variance"], "pct": gp["pct_variance"]} for gp in g["periods"]])
        is_last = i == len(result["expense_groups"]) - 1
        label = "Net income" if is_last else f"Net income after {g['name']}"
        row(g["name"], "", label, "",
            [{"base": gp["base_running_after"], "comp": gp["compare_running_after"],
              "variance": gp["running_variance"], "pct": gp["running_pct_variance"]} for gp in g["periods"]])
    if not result["expense_groups"]:
        row("Income", "", "Net income", "",
            [{"base": pt["net_income"], "comp": pt["compare_net_income"],
              "variance": pt["net_income_variance_amount"], "pct": pt["net_income_variance"]}
             for pt in result["periods_totals"]])
    return csv_response(buf, _income_statement_export_filename(scenario, compare, date_from, date_to, split, "csv"))


@app.get("/export/income-statement.xlsx")
def income_statement_export_xlsx(scenario: str = "ACTUAL", compare: str = "",
                                 date_from: str = "", date_to: str = "", zeros: int = 0,
                                 pct_of_base: int = 0, split: str = ""):
    """XLSX counterpart to income_statement_export_csv() above — same
    _income_statement_rows()/_income_statement_matrix() data, same overall
    shape (income rows, then each expense group with its own running "Net
    income after X" row), styled with the helpers above instead of
    written as plain CSV rows. `depth` (from _build_account_tree) drives
    indentation the way the HTML report's own chevrons do, standing in
    for the CSV's separate Path/breadcrumb column — not carried over here
    since a sighted spreadsheet reader gets the same hierarchy from
    indentation, and dropping it keeps the sheet's column count matched
    to the CSV's data columns instead of growing it.

    No separate "Total income"/"Total {group}" row the way the CSV export
    has one — a group's own top-level account (its rows[0], always first
    since _flatten_tree puts a node ahead of its children) already *is*
    that rolled-up total; see _build_account_tree's "subtotal" comment.
    Writing it again a few rows down was the same number twice under two
    different labels. Instead, that first row gets the bold/ruled "group"
    treatment directly, in place — one real total per section, not two.

    Every account row's own base/compare figure is a literal, not a
    formula: this is a rolled-up multi-root account tree, so a plain
    SUM() over a *visible row range* would double-count wherever a group
    is more than one level deep. What's written there is exactly the
    same number the HTML report and the CSV export already show for that
    row. Three things layered on top of those literals *are* live
    formulas, though, each one safe for the same underlying reason —
    every cell referenced is named individually, by row, never swept in
    as part of a range that could double-count:
    - Variance and % Variance (_xlsx_variance_formulas) — each one only
      references the two cells already sitting in its own row.
    - Each "Net income after X" running row (_xlsx_sum_formula) — Income's
      root row(s) minus every expense group's own root row seen so far,
      each named as its own cell (e.g. "=C6+C20-C34"), not a range.
    Edit any base/compare figure by hand later and every running total
    and variance downstream of it recalculates instead of going stale.

    Two more reader aids from live feedback against a real download:
    every Variance/% Variance column also carries real conditional-
    formatting rules (red negative, green positive —
    _xlsx_variance_coloring), not a color baked in at generation time, so
    it still tracks correctly if a cell is edited by hand afterward.
    Split view also draws a heavier rule down the right edge of every
    period's own column group (but the last) — see
    _xlsx_thicken_right_border — since a wide multi-period sheet with
    only the thin per-cell grid to go on is easy to lose your place
    scanning across."""
    periods = _split_periods(date_from, date_to, split)
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    subtitle = scenario
    if compare:
        subtitle += f" vs. {compare}"
    if date_from and date_to:
        subtitle += f" · {date_from} to {date_to}"
    elif date_from:
        subtitle += f" · from {date_from}"
    elif date_to:
        subtitle += f" · through {date_to}"
    if split:
        subtitle += f" · split {split}"
    ws.cell(row=1, column=1, value="Income Statement").font = _XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _XLSX_SUBTITLE_FONT

    if not periods:
        result = _income_statement_rows(scenario, date_from, date_to, compare, zeros, bool(pct_of_base))
        headers = ["Code", "Account", scenario or "Amount"]
        if compare:
            headers += ["Variance", "% Variance", compare]
        n_cols = len(headers)
        header_row, data_start = 4, 5
        _xlsx_header_row(ws, header_row, headers)

        def row(r: int, code, name, depth, base, comp=None, variance=None, pct=None, style="line"):
            # variance/pct (the backend-computed figures, same ones the
            # CSV export writes) are accepted for call-site parity but
            # unused here — Variance/% Variance are live formulas instead
            # (_xlsx_variance_formulas), referencing this same row's own
            # C{r}/F{r} cells.
            value_cols = [(3, base, _XLSX_MONEY_FMT)]
            if compare:
                var_f, pct_f = _xlsx_variance_formulas(f"C{r}", f"F{r}", bool(pct_of_base))
                value_cols += [(4, var_f, _XLSX_MONEY_FMT), (5, pct_f, _XLSX_PCT_FMT), (6, comp, _XLSX_MONEY_FMT)]
            _xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

        # Root-row cell references for the running-total formulas below —
        # every income group's own root row (base column C, compare
        # column F), then every expense group's as its own root row is
        # written. A "Net income after X" row is then just those income
        # roots minus however many expense roots have been seen so far —
        # _xlsx_sum_formula, same "reference the exact cell, never a
        # range" safety as _xlsx_variance_formulas above.
        income_roots_c, income_roots_f, expense_roots_c, expense_roots_f = [], [], [], []
        r = data_start
        for g in result["income_groups"]:
            for i, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    line["base_net"], line["compare_net"], line["variance"], line["pct_variance"],
                    style="group" if i == 0 else "line")
                if i == 0:
                    income_roots_c.append(f"C{r}")
                    income_roots_f.append(f"F{r}")
                r += 1
        for i, g in enumerate(result["expense_groups"]):
            r += 1  # blank separator row — same breathing room the CSV gives with w.writerow([])
            for j, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    line["base_net"], line["compare_net"], line["variance"], line["pct_variance"],
                    style="group" if j == 0 else "line")
                if j == 0:
                    expense_roots_c.append(f"C{r}")
                    expense_roots_f.append(f"F{r}")
                r += 1
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            running_base = _xlsx_sum_formula(income_roots_c, expense_roots_c)
            running_comp = _xlsx_sum_formula(income_roots_f, expense_roots_f) if compare else None
            row(r, "", label, 1, running_base, running_comp, style="running")
            r += 1
        if not result["expense_groups"]:
            running_base = _xlsx_sum_formula(income_roots_c)
            running_comp = _xlsx_sum_formula(income_roots_f) if compare else None
            row(r, "", "Net income", 1, running_base, running_comp, style="running")
            r += 1
        last_row = r - 1
        if compare:
            _xlsx_variance_coloring(ws, 4, data_start, last_row)  # Variance
            _xlsx_variance_coloring(ws, 5, data_start, last_row)  # % Variance
    else:
        result = _income_statement_matrix(scenario, periods, date_from, date_to, compare, zeros, bool(pct_of_base))
        cols_per_period = 4 if compare else 1
        field_labels = [scenario] + (["Variance", "% Var.", compare] if compare else [])
        n_cols = 2 + cols_per_period * len(result["periods"])
        header_row, field_row, data_start = 4, 5, 6

        # Two-row header: the date ("2026-01", "Total", "Average") merged
        # and centered across that period's own field columns, with the
        # field names (ACTUAL/Variance/%/compare) on their own row right
        # below instead of repeated into every column header — the split
        # CSV export prefixes every column with its period's label
        # because a bare CSV has no merged cells to lean on; XLSX does.
        _xlsx_merged_header(ws, header_row, 1, field_row, 1, "Code")
        _xlsx_merged_header(ws, header_row, 2, field_row, 2, "Account")
        for i, p in enumerate(result["periods"]):
            start_col = 3 + i * cols_per_period
            _xlsx_merged_header(ws, header_row, start_col, header_row, start_col + cols_per_period - 1, p["label"])
            _xlsx_header_row(ws, field_row, field_labels, start_col=start_col)

        def row(r: int, code, name, depth, period_vals, style="line"):
            # v["variance"]/v["pct"] unused here — same live-formula
            # treatment as the single-range row() above, one Variance/%
            # Variance formula pair per period, each referencing only
            # that period's own base/compare cells in this row.
            value_cols, col = [], 3
            for v in period_vals:
                base_col = col
                value_cols.append((col, v.get("base"), _XLSX_MONEY_FMT))
                col += 1
                if compare:
                    comp_col = col + 2
                    var_f, pct_f = _xlsx_variance_formulas(
                        f"{get_column_letter(base_col)}{r}", f"{get_column_letter(comp_col)}{r}", bool(pct_of_base))
                    value_cols += [(col, var_f, _XLSX_MONEY_FMT),
                                   (col + 1, pct_f, _XLSX_PCT_FMT),
                                   (comp_col, v.get("comp"), _XLSX_MONEY_FMT)]
                    col += 3
            _xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

        def running_period_vals(income_rows: list[int], expense_rows: list[int]) -> list[dict]:
            """One {"base", "comp"} formula pair per period column-group
            for a "Net income after X" row — same _xlsx_sum_formula
            safety as the single-range branch above, just repeated once
            per period's own pair of columns instead of the fixed C/F
            pair, since `income_rows`/`expense_rows` name *row* numbers
            (one physical row per account, shared across every period)
            while the actual cell reference still needs that period's own
            column."""
            out = []
            for i in range(len(result["periods"])):
                base_col = get_column_letter(3 + i * cols_per_period)
                base_f = _xlsx_sum_formula([f"{base_col}{rr}" for rr in income_rows],
                                           [f"{base_col}{rr}" for rr in expense_rows])
                comp_f = None
                if compare:
                    comp_col = get_column_letter(3 + i * cols_per_period + 3)
                    comp_f = _xlsx_sum_formula([f"{comp_col}{rr}" for rr in income_rows],
                                               [f"{comp_col}{rr}" for rr in expense_rows])
                out.append({"base": base_f, "comp": comp_f})
            return out

        # Root *row numbers* only here (not column letters) — one physical
        # row per account, the same row number reused across every
        # period's own column-group, unlike the single-range branch above
        # where "C"/"F" are already the one-and-only base/compare columns.
        income_root_rows, expense_root_rows = [], []
        r = data_start
        for g in result["income_groups"]:
            for i, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                      "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in line["periods"]],
                    style="group" if i == 0 else "line")
                if i == 0:
                    income_root_rows.append(r)
                r += 1
        for i, g in enumerate(result["expense_groups"]):
            r += 1
            for j, line in enumerate(g["rows"]):
                row(r, line["account_code"], line["account_name"], line["depth"],
                    [{"base": rp.get("base_net"), "comp": rp.get("compare_net"),
                      "variance": rp.get("variance"), "pct": rp.get("pct_variance")} for rp in line["periods"]],
                    style="group" if j == 0 else "line")
                if j == 0:
                    expense_root_rows.append(r)
                r += 1
            is_last = i == len(result["expense_groups"]) - 1
            label = "Net income" if is_last else f"Net income after {g['name']}"
            row(r, "", label, 1, running_period_vals(income_root_rows, expense_root_rows), style="running")
            r += 1
        if not result["expense_groups"]:
            row(r, "", "Net income", 1, running_period_vals(income_root_rows, []), style="running")
            r += 1
        last_row = r - 1

        if compare:
            # A heavier rule down the right edge of every period's own
            # column group but the last (the very last one is just the
            # sheet's outer edge, not a boundary between two periods) —
            # spans the header rows too, not just the data, so it reads
            # as one continuous line separating "2026-01" from "2026-02"
            # the way a ruled column break would on a printed ledger.
            # Total/Average count as periods here same as any real month
            # — they're just two more column groups in `result["periods"]`
            # — so they get the same treatment. Compare-only: with no
            # compare scenario each period is a single plain column
            # (cols_per_period == 1), and a divider after literally every
            # column read as clutter rather than a period boundary — see
            # live feedback that asked for this to be scoped down.
            for i, p in enumerate(result["periods"][:-1]):
                end_col = 3 + i * cols_per_period + cols_per_period - 1
                for rr in range(header_row, last_row + 1):
                    _xlsx_thicken_right_border(ws, rr, end_col)
            for i, p in enumerate(result["periods"]):
                start_col = 3 + i * cols_per_period
                _xlsx_variance_coloring(ws, start_col + 1, data_start, last_row)  # Variance
                _xlsx_variance_coloring(ws, start_col + 2, data_start, last_row)  # % Variance

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx_response(wb, _income_statement_export_filename(scenario, compare, date_from, date_to, split, "xlsx"))


# ---------------------------------------------------------------------------
# Balance sheet — Assets, Liabilities, Equity, always "as of" a date (these
# are stock accounts, not flow accounts — a range doesn't apply). Same
# simulated-monthly-close split as Trial Balance, and for the same reason
# (see there): "Current Year Earnings (Unclosed)" is this fiscal year's
# earnings not already reflected this month, "Prior Year Earnings
# (Unclosed)" is every year before this one — together they total the
# exact same unclosed P&L as a single lifetime figure would, so nothing
# about the balance changes, only how it reads. `raw=1` collapses this
# back to one plain "Current earnings (unclosed)" line — the true,
# un-simulated all-time total.
# ---------------------------------------------------------------------------
def _balance_sheet_rows(scenario: str, as_of: str, raw: int = 0, zeros: int = 0) -> dict:
    as_of_date = as_of or None
    as_of_dt = date.fromisoformat(as_of_date) if as_of_date else date.today()
    accounts = q("SELECT * FROM v_dim_account WHERE is_active ORDER BY sort_path")
    full_balances = {r["account_id"]: r["net"] for r in
                     q("SELECT * FROM fn_account_balances(%s, %s)", (scenario, as_of_date))}
    total_pnl = _pnl_net(accounts, full_balances)

    if raw:
        earnings_lines = [("Current earnings (unclosed)", total_pnl)]
    else:
        # No MTD carve-out here, unlike Trial Balance: a balance sheet has
        # no Income/Expense section of its own to hold that money in, so
        # "Current Year" has to mean the *whole* fiscal year to date
        # (MTD included) or Assets would stop reconciling against
        # Liabilities + Equity by exactly the MTD amount.
        fy_start = date(as_of_dt.year, 1, 1).isoformat()
        fy_balances = {r["account_id"]: r["net"] for r in
                       q("SELECT * FROM fn_account_balances(%s, %s, %s)", (scenario, as_of_date, fy_start))}
        fy_earnings = _pnl_net(accounts, fy_balances)
        earnings_lines = [
            ("Current Year Earnings (Unclosed)", fy_earnings),
            ("Prior Year Earnings (Unclosed)", total_pnl - fy_earnings),
        ]
    # Same "hide a boring zero line" rule Trial Balance's own synthetic
    # earnings rows follow — a zero unclosed-earnings line is noise, not
    # information, unless zeros asked to see everything.
    if not zeros:
        earnings_lines = [(label, amt) for label, amt in earnings_lines if amt != 0]

    roots = _build_account_tree(accounts, full_balances)
    asset_roots = [r for r in roots if r["account_type"] == "asset"]
    liability_roots = [r for r in roots if r["account_type"] == "liability"]
    equity_roots = [r for r in roots if r["account_type"] == "equity"]
    assets = _flatten_tree(asset_roots, zeros=zeros)
    liabilities = _flatten_tree(liability_roots, zeros=zeros)
    equity = _flatten_tree(equity_roots, zeros=zeros)

    total_assets = sum(r["subtotal"] for r in asset_roots)
    total_liabilities = -sum(r["subtotal"] for r in liability_roots)
    total_equity = -sum(r["subtotal"] for r in equity_roots) + total_pnl
    return {
        "assets": assets, "liabilities": liabilities, "equity": equity,
        "earnings_lines": earnings_lines,
        "total_assets": total_assets, "total_liabilities": total_liabilities,
        "total_equity": total_equity,
        "total_liab_and_equity": total_liabilities + total_equity,
        "in_balance": total_assets == total_liabilities + total_equity,
    }


@app.get("/balance-sheet")
def balance_sheet_page(request: Request, scenario: str = "ACTUAL", as_of: str = None,
                       raw: int = 0, zeros: int = 0):
    result = _balance_sheet_rows(scenario, as_of, raw, zeros)
    return templates.TemplateResponse(request, "balance_sheet.html", {
        "nav": "balance_sheet", "scenarios": scenarios_all(), "scenario": scenario,
        "as_of": as_of or "", "raw": raw, "zeros": zeros, "today": date.today().isoformat(), **result,
    })


def _balance_sheet_export_filename(scenario: str, as_of: str, raw: int, ext: str) -> str:
    name = f"postwarden-balance-sheet-{scenario}"
    if as_of:
        name += f"_{as_of}"
    if raw:
        name += "_raw"
    return f"{name}.{ext}"


@app.get("/export/balance-sheet.csv")
def balance_sheet_export_csv(scenario: str = "ACTUAL", as_of: str = None, raw: int = 0, zeros: int = 0):
    result = _balance_sheet_rows(scenario, as_of, raw, zeros)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Section", "Code", "Account", "Path", "Amount"])
    for r in result["assets"]:
        w.writerow(["Assets", r["account_code"], r["account_name"], r["path"], r["subtotal"]])
    for r in result["liabilities"]:
        w.writerow(["Liabilities", r["account_code"], r["account_name"], r["path"], -r["subtotal"]])
    for r in result["equity"]:
        w.writerow(["Equity", r["account_code"], r["account_name"], r["path"], -r["subtotal"]])
    for label, amount in result["earnings_lines"]:
        w.writerow(["Equity", "", label, "", amount])
    w.writerow([])
    w.writerow(["Total assets", "", "", "", result["total_assets"]])
    w.writerow(["Total liabilities + equity", "", "", "", result["total_liab_and_equity"]])
    return csv_response(buf, _balance_sheet_export_filename(scenario, as_of, raw, "csv"))


@app.get("/export/balance-sheet.xlsx")
def balance_sheet_export_xlsx(scenario: str = "ACTUAL", as_of: str = None, raw: int = 0, zeros: int = 0):
    """XLSX counterpart to balance_sheet_export_csv() above. Each section
    (Assets/Liabilities/Equity) gets a bold section-title row, same as
    Trial Balance; a section's own top-level account row (depth 1) gets
    the same "group" bold+ruled treatment Income Statement's groups use,
    since — same reasoning as there — that row's figure already is that
    root's own rolled-up total, no separate "Total X" needed unless a
    section actually has more than one root (rare, but Trial Balance's
    own g["show_type_total"] doesn't exist here since _balance_sheet_rows
    never separated multi-root sections out that way; a second top-level
    Assets account just adds a second bold row in the same section,
    consistent with how Income Statement handles a second top-level
    expense account too). "Total assets"/"Total liabilities + equity" are
    a real cross-section identity, not a duplicate of anything above
    them, so they keep their own grand-total row with the accountant's
    double-rule — red instead of ink when the sheet doesn't balance."""
    result = _balance_sheet_rows(scenario, as_of, raw, zeros)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    subtitle = f"{scenario} · {'As of ' + as_of if as_of else 'Through today'}"
    if not raw:
        subtitle += " · simulated monthly close"
    ws.cell(row=1, column=1, value="Balance Sheet").font = _XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Amount"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    _xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, amount, style="line"):
        _xlsx_data_row(ws, r, [(1, code), (2, name)], [(3, amount, _XLSX_MONEY_FMT)], style, max(depth - 1, 0))

    r = data_start
    sections = [("Assets", result["assets"], 1), ("Liabilities", result["liabilities"], -1),
               ("Equity", result["equity"], -1)]
    for label, rows, sign in sections:
        _xlsx_data_row(ws, r, [(1, ""), (2, label)], [], style="group")
        r += 1
        for line in rows:
            row(r, line["account_code"], line["account_name"], line.get("depth", 2),
                sign * line["subtotal"], style="group" if line.get("depth") == 1 else "line")
            r += 1
        if label == "Equity":
            for earn_label, amount in result["earnings_lines"]:
                row(r, "", earn_label, 2, amount, style="line")
                r += 1
    r += 1  # blank separator, same breathing room the CSV gives with w.writerow([])
    grand_style = "grand" if result["in_balance"] else "grand_bad"
    row(r, "", "Total assets", 1, result["total_assets"], style=grand_style)
    r += 1
    row(r, "", "Total liabilities + equity", 1, result["total_liab_and_equity"], style=grand_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx_response(wb, _balance_sheet_export_filename(scenario, as_of, raw, "xlsx"))


# ---------------------------------------------------------------------------
# Cash flow statement — flat (no operating/investing/financing split, out
# of scope per SPEC.md decision 20), grouped by the contra-account each
# cash leg attributes to. fn_cash_flow_lines (db/schema.sql) does the
# real per-transaction attribution at full granularity (every non-cash
# leg, its own posted amount, sign-flipped — nothing netted or bucketed
# at the SQL layer); everything here is presentation on top of that raw
# truth — grouping rows into inflows/outflows, peeling equity-contra legs
# into their own "ledger adjustments" section, folding a reducible
# income entry's deduction legs into its own income row, and running the
# three-way tie-out the spec calls a hard invariant. See SPEC.md decision
# 20's addenda for the reasoning behind each of these three presentation
# rules — the underlying per-leg numbers this all runs on never change.
# ---------------------------------------------------------------------------
def _cash_flow_tie_out(scenario: str, date_from_v: str | None, date_to_v: str | None,
                       statement_total) -> dict:
    """The three numbers the spec says must agree to the cent: the
    statement's own total, the net leg activity on is_cashflow accounts
    for the same (post-exclusion) set of transactions, and the plain
    balance-sheet roll-forward (ending − beginning) of those same
    accounts. A mismatch means an untagged/mistagged account, a bad
    split attribution, or a pure-transfer wrongly included/excluded —
    surfaced as a warning banner on the report and logged, per the spec,
    rather than silently shown as if nothing were wrong.

    beginning/ending are returned (not just balance_delta) so the report
    can show them unconditionally as their own lines — previously they
    were computed here but only ever surfaced inside the failure banner,
    so a passing report gave no grounding for what "net change" was a
    change *from*. Presentation-layer bucketing (ledger adjustments,
    netting) never touches this function or its inputs: statement_total
    is still literally the same net_change number it always was, and
    beginning/ending are still the plain balance-sheet roll-forward,
    independent of how the statement chooses to group its rows."""
    cash_leg_net = q1("""
        SELECT COALESCE(SUM(f.amount), 0) AS net
          FROM v_fact_lines f JOIN accounts a ON a.id = f.account_id
         WHERE a.is_cashflow AND f.scenario_code = %s
           AND f.entry_date <= COALESCE(%s, 'infinity'::date)
           AND f.entry_date >= COALESCE(%s, '-infinity'::date)
           AND f.entry_id IN (SELECT DISTINCT entry_id FROM fn_cash_flow_lines(%s, %s, %s))
    """, (scenario, date_to_v, date_from_v, scenario, date_from_v, date_to_v))["net"]

    # Beginning balance is "as of the day before date_from", cumulative
    # since inception — the same balance a Balance Sheet run as of that
    # day would show. No date_from means the range is unbounded at the
    # start, so there's nothing before it to roll forward from.
    if date_from_v:
        begin_as_of = (date.fromisoformat(date_from_v) - timedelta(days=1)).isoformat()
        beginning = q1("""SELECT COALESCE(SUM(net), 0) AS net FROM fn_account_balances(%s, %s)
                            WHERE account_id IN (SELECT id FROM accounts WHERE is_cashflow)""",
                       (scenario, begin_as_of))["net"]
    else:
        beginning = 0
    ending = q1("""SELECT COALESCE(SUM(net), 0) AS net FROM fn_account_balances(%s, %s)
                     WHERE account_id IN (SELECT id FROM accounts WHERE is_cashflow)""",
               (scenario, date_to_v))["net"]
    balance_delta = ending - beginning

    ok = statement_total == cash_leg_net == balance_delta
    if not ok:
        logger.error(
            "Cash flow tie-out mismatch (scenario=%s, %s..%s): "
            "statement_total=%s cash_leg_net=%s balance_delta=%s",
            scenario, date_from_v, date_to_v, statement_total, cash_leg_net, balance_delta,
        )
    return {"ok": ok, "statement_total": statement_total, "cash_leg_net": cash_leg_net,
            "balance_delta": balance_delta, "beginning": beginning, "ending": ending}


def _cash_flow_rows(scenario: str, date_from: str, date_to: str) -> dict:
    """Groups fn_cash_flow_lines' raw per-leg rows into the report's three
    sections. Three routing rules apply, per entry, in this order — see
    SPEC.md decision 20's addenda for the full reasoning behind each:

      1. Equity-typed contra legs are always their own row, always in
         ledger_adjustments — never blended into inflows/outflows, and
         never excluded outright either (excluding them would break the
         tie-out's beginning+net_change==ending identity, since the cash
         genuinely did move; the fix is presentation, not deletion).
         Peeled off first so they can't interact with rule 2.

      2. Among what's left on that same entry: if there is exactly one
         income-typed leg and at least one expense-typed leg, they
         collapse into a single row under the income leg's own account —
         amount is their signed sum, which (the entry balances, so this
         is exact, not estimated) already equals that leg group's own
         net cash contribution. The folded-away expense legs ride along
         as that row's netted_from, so the detail is demoted to an
         annotation, not deleted — still reachable, just not cluttering
         the top-level view. Two or more income legs on one entry is
         deliberately left un-netted: there's no principled way to
         decide which income leg a shared deduction belongs to, so
         rather than guess, every leg itemizes on its own, same as if
         rule 2 had never fired.

      3. Everything else — asset/liability legs always, plus any
         income/expense leg rule 2 didn't consume — itemizes exactly as
         fn_cash_flow_lines returned it, unchanged from before this rule
         existed.
    """
    date_from_v = date_from or None
    date_to_v = date_to or None
    lines = q("SELECT * FROM fn_cash_flow_lines(%s, %s, %s)", (scenario, date_from_v, date_to_v))
    accounts_by_id = {a["id"]: a for a in q("SELECT * FROM v_dim_account")}

    by_entry: dict[str, list[dict]] = {}
    for l in lines:
        by_entry.setdefault(l["entry_id"], []).append(l)

    def bump(agg: dict, account_id: int, amount, flagged: bool, netted_from: dict | None = None):
        row = agg.setdefault(account_id, {"amount": 0, "flagged": False, "netted_from": {}})
        row["amount"] += amount
        row["flagged"] = row["flagged"] or flagged
        for nf_id, nf_amount in (netted_from or {}).items():
            row["netted_from"][nf_id] = row["netted_from"].get(nf_id, 0) + nf_amount

    activity: dict[int, dict] = {}     # real economic activity -> inflows/outflows
    adjustments: dict[int, dict] = {}  # equity-contra -> ledger adjustments
    for entry_id, entry_lines in by_entry.items():
        flagged = any(l["n_cash_legs"] > 1 for l in entry_lines)
        by_type: dict[str, list[dict]] = {}
        for l in entry_lines:
            a = accounts_by_id.get(l["contra_account_id"])
            if a:
                by_type.setdefault(a["account_type"], []).append(l)

        # Rule 1 — equity, always its own row, always ledger_adjustments.
        for l in by_type.pop("equity", []):
            bump(adjustments, l["contra_account_id"], l["amount"], flagged)

        # Rule 2 — fold expense legs into a single well-defined income leg.
        income_legs = by_type.pop("income", [])
        expense_legs = by_type.pop("expense", [])
        if len(income_legs) == 1 and expense_legs:
            inc = income_legs[0]
            total = inc["amount"] + sum(e["amount"] for e in expense_legs)
            netted_from = {e["contra_account_id"]: e["amount"] for e in expense_legs}
            bump(activity, inc["contra_account_id"], total, flagged, netted_from)
        else:
            for l in income_legs + expense_legs:
                bump(activity, l["contra_account_id"], l["amount"], flagged)

        # Rule 3 — everything left (asset/liability) itemizes as-is.
        for l in [x for legs in by_type.values() for x in legs]:
            bump(activity, l["contra_account_id"], l["amount"], flagged)

    def to_rows(agg: dict) -> list[dict]:
        out = []
        for account_id, r in agg.items():
            a = accounts_by_id.get(account_id)
            if not a:
                continue
            netted_from = sorted((
                {"account_code": accounts_by_id[nf_id]["code"],
                 "account_name": accounts_by_id[nf_id]["name"], "amount": nf_amount}
                for nf_id, nf_amount in r["netted_from"].items() if nf_id in accounts_by_id
            ), key=lambda n: n["account_code"])
            out.append({"account_id": account_id, "account_code": a["code"], "account_name": a["name"],
                        "parent_path": a["parent_path"], "amount": r["amount"],
                        "flagged": r["flagged"], "netted_from": netted_from})
        out.sort(key=lambda r: r["account_code"])
        return out

    activity_rows = to_rows(activity)
    inflows = [r for r in activity_rows if r["amount"] >= 0]
    outflows = [r for r in activity_rows if r["amount"] < 0]
    ledger_adjustments = to_rows(adjustments)

    total_inflows = sum(r["amount"] for r in inflows)
    total_outflows = sum(r["amount"] for r in outflows)
    total_adjustments = sum(r["amount"] for r in ledger_adjustments)
    # Unchanged from before rules 1/2 existed: still every non-cash leg's
    # own contribution summed once. Rules 1/2 only ever regroup rows that
    # already summed to the same total, so this, the tie-out, and the
    # beginning+net_change==ending identity are all exactly as before.
    net_change = total_inflows + total_outflows + total_adjustments

    # Distinct transactions with more than one cash leg (checking +
    # savings both funded from one payroll deposit, say) — the
    # attribution above already divides these correctly (see
    # fn_cash_flow_lines' own comment: the formula doesn't actually
    # guess), but the spec asks that they surface for a human glance
    # anyway rather than blend in silently.
    flagged_entries = q("""
        SELECT DISTINCT e.id, e.entry_date, e.description, p.name AS payee
          FROM journal_entries e
          LEFT JOIN payees p ON p.id = e.payee_id
         WHERE e.id IN (SELECT entry_id FROM fn_cash_flow_lines(%s, %s, %s) WHERE n_cash_legs > 1)
         ORDER BY e.entry_date, e.id
    """, (scenario, date_from_v, date_to_v))

    return {
        "inflows": inflows, "outflows": outflows, "ledger_adjustments": ledger_adjustments,
        "total_inflows": total_inflows, "total_outflows": total_outflows,
        "total_adjustments": total_adjustments,
        "net_change": net_change, "flagged_entries": flagged_entries,
        "tie_out": _cash_flow_tie_out(scenario, date_from_v, date_to_v, net_change),
    }


@app.get("/cash-flow")
def cash_flow_page(request: Request, scenario: str = "ACTUAL",
                   date_from: str = "", date_to: str = ""):
    today = date.today()
    date_from = date_from or today.replace(day=1).isoformat()
    date_to = date_to or today.isoformat()
    result = _cash_flow_rows(scenario, date_from, date_to)
    return templates.TemplateResponse(request, "cash_flow.html", {
        "nav": "cash_flow", "scenarios": scenarios_all(), "scenario": scenario,
        "date_from": date_from, "date_to": date_to, "today": today.isoformat(), **result,
    })


@app.get("/export/cash-flow.csv")
def cash_flow_export_csv(scenario: str = "ACTUAL", date_from: str = "", date_to: str = ""):
    # Blank date_from/date_to means "unbounded" here, deliberately not
    # defaulted to the current month the way the page is — same
    # already-established split between page and CSV export that
    # income_statement_export_csv documents on itself.
    result = _cash_flow_rows(scenario, date_from, date_to)
    buf = io.StringIO()
    w = csv.writer(buf)

    def netted_of(r: dict) -> str:
        return "; ".join(f"{n['account_name']} {n['amount']:.2f}" for n in r["netted_from"])

    w.writerow(["Section", "Code", "Account", "Amount", "Flagged for review", "Net of"])
    w.writerow(["", "", "Beginning cash balance", result["tie_out"]["beginning"], "", ""])
    w.writerow([])
    for r in result["inflows"]:
        w.writerow(["Inflows", r["account_code"], r["account_name"], r["amount"],
                    "yes" if r["flagged"] else "", netted_of(r)])
    w.writerow(["Inflows", "", "Total inflows", result["total_inflows"], "", ""])
    w.writerow([])
    for r in result["outflows"]:
        w.writerow(["Outflows", r["account_code"], r["account_name"], r["amount"],
                    "yes" if r["flagged"] else "", netted_of(r)])
    w.writerow(["Outflows", "", "Total outflows", result["total_outflows"], "", ""])
    w.writerow([])
    # Only present when non-empty — most periods have no equity-contra
    # activity at all (opening-balance seeding happens once), so an
    # always-present empty section would just be noise most exports.
    if result["ledger_adjustments"]:
        for r in result["ledger_adjustments"]:
            w.writerow(["Ledger adjustments", r["account_code"], r["account_name"], r["amount"],
                        "yes" if r["flagged"] else "", ""])
        w.writerow(["Ledger adjustments", "", "Total ledger adjustments", result["total_adjustments"], "", ""])
        w.writerow([])
    w.writerow(["", "", "Net change in cash", result["net_change"], "", ""])
    w.writerow(["", "", "Ending cash balance", result["tie_out"]["ending"], "", ""])
    w.writerow([])
    w.writerow(["", "", "Tie-out check", "PASS" if result["tie_out"]["ok"] else "FAIL — see app log", "", ""])
    return csv_response(buf, _cash_flow_export_filename(scenario, date_from, date_to, "csv"))


def _cash_flow_export_filename(scenario: str, date_from: str, date_to: str, ext: str) -> str:
    name = f"postwarden-cash-flow-{scenario}"
    if date_from and date_to:
        name += f"_{date_from}_to_{date_to}"
    elif date_from:
        name += f"_from_{date_from}"
    elif date_to:
        name += f"_through_{date_to}"
    return f"{name}.{ext}"


@app.get("/export/cash-flow.xlsx")
def cash_flow_export_xlsx(scenario: str = "ACTUAL", date_from: str = "", date_to: str = ""):
    """XLSX counterpart to cash_flow_export_csv() above. No account tree
    here (fn_cash_flow_lines' rows are already flat, one per contra
    account — see _cash_flow_rows' own docstring), so no depth/indent and
    no "first row is the total" duplication concern the tree-shaped
    reports have. Beginning/Net change/Ending get the same bold "group"
    headline treatment as a section title, and the closing Tie-out row
    reuses the grand/grand_bad split Trial Balance/Balance Sheet use for
    their own balance check — green ink for PASS, red for FAIL, same
    accountant's double-rule underneath."""
    result = _cash_flow_rows(scenario, date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    subtitle = scenario
    if date_from and date_to:
        subtitle += f" · {date_from} to {date_to}"
    elif date_from:
        subtitle += f" · from {date_from}"
    elif date_to:
        subtitle += f" · through {date_to}"
    ws.cell(row=1, column=1, value="Cash Flow Statement").font = _XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", "Amount", "Flagged", "Net of"]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    _xlsx_header_row(ws, header_row, headers)

    def netted_of(line: dict) -> str:
        return "; ".join(f"{n['account_name']} {n['amount']:.2f}" for n in line["netted_from"])

    def row(r: int, code, name, amount, flagged="", netted="", style="line"):
        label_cols = [(1, code), (2, name)]
        value_cols = [(3, amount, _XLSX_MONEY_FMT)]
        _xlsx_data_row(ws, r, label_cols, value_cols, style)
        # Flagged/Net of are plain descriptive text, not money — no
        # banding/border/number-format, same as every report's non-money
        # label columns.
        font = _XLSX_ROW_FONTS[style]
        for col, text in ((4, flagged), (5, netted)):
            cell = ws.cell(row=r, column=col, value=text or None)
            cell.font = font

    r = data_start
    row(r, "", "Beginning cash balance", result["tie_out"]["beginning"], style="group")
    r += 2  # blank separator row, same breathing room the CSV gives with w.writerow([])
    for section, rows_, total_label, total in (
        ("Inflows", result["inflows"], "Total inflows", result["total_inflows"]),
        ("Outflows", result["outflows"], "Total outflows", result["total_outflows"]),
    ):
        _xlsx_data_row(ws, r, [(1, ""), (2, section)], [], style="group")
        r += 1
        for line in rows_:
            row(r, line["account_code"], line["account_name"], line["amount"],
                "yes" if line["flagged"] else "", netted_of(line))
            r += 1
        row(r, "", total_label, total, style="subtotal")
        r += 2
    if result["ledger_adjustments"]:
        _xlsx_data_row(ws, r, [(1, ""), (2, "Ledger adjustments")], [], style="group")
        r += 1
        for line in result["ledger_adjustments"]:
            row(r, line["account_code"], line["account_name"], line["amount"],
                "yes" if line["flagged"] else "")
            r += 1
        row(r, "", "Total ledger adjustments", result["total_adjustments"], style="subtotal")
        r += 2
    row(r, "", "Net change in cash", result["net_change"], style="group")
    r += 1
    row(r, "", "Ending cash balance", result["tie_out"]["ending"], style="group")
    r += 2
    tie_style = "grand" if result["tie_out"]["ok"] else "grand_bad"
    tie_text = "PASS" if result["tie_out"]["ok"] else "FAIL — see app log"
    _xlsx_data_row(ws, r, [(1, ""), (2, "Tie-out check")], [(3, tie_text, "General")], style=tie_style)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 36
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx_response(wb, _cash_flow_export_filename(scenario, date_from, date_to, "xlsx"))


# ---------------------------------------------------------------------------
# Budget grid — the ActualBudget-style grid for an income-statement-only
# scenario (scenarios.income_statement_only): one month at a time, Actual
# (this month's real postings), the Variance against Budgeted (columns
# in between, both sharing the money-in-between reading order every
# other two-scenario report uses), and Budgeted itself (editable) last,
# income/expense accounts only, no journal entries anywhere in sight.
# Reuses _build_account_tree twice — once over
# budget_lines, once over ACTUAL's postings for the same month — and
# merges the two node-for-node rather than inventing a second rollup
# function, since both sides share the exact same account tree shape.
# ---------------------------------------------------------------------------
def _shift_month(month: str, delta_months: int) -> str:
    d = date.fromisoformat(month)
    total = d.month - 1 + delta_months
    return date(d.year + total // 12, total % 12 + 1, 1).isoformat()


def _budget_rows(scenario: str, month: str, pct_of_base: bool = False) -> dict:
    month_start = date.fromisoformat(month)
    month_end = date(month_start.year, month_start.month,
                     calendar.monthrange(month_start.year, month_start.month)[1])

    accounts = q("""SELECT * FROM v_dim_account
                     WHERE is_active AND account_type IN ('income', 'expense')
                     ORDER BY sort_path""")
    scen = q1("SELECT id FROM scenarios WHERE code = %s AND income_statement_only",
              (scenario,))
    budgeted_by_id = {}
    if scen:
        budgeted_by_id = {r["account_id"]: r["amount"] for r in q(
            "SELECT account_id, amount FROM budget_lines WHERE scenario_id = %s AND period_month = %s",
            (scen["id"], month_start))}
    actual_by_id = {r["account_id"]: r["net"] for r in q(
        "SELECT * FROM fn_account_balances(%s, %s, %s)",
        ("ACTUAL", month_end.isoformat(), month_start.isoformat()))}

    budget_roots = _build_account_tree(accounts, budgeted_by_id)
    actual_by_node_id = {}

    def index(nodes):
        for n in nodes:
            actual_by_node_id[n["id"]] = n
            index(n["children"])
    index(_build_account_tree(accounts, actual_by_id))

    def merge(nodes):
        out = []
        for n in nodes:
            # journal amounts are debit-positive, so an income account's
            # actual net comes out negative; budget_lines.amount is a
            # plain target with no sign to juggle — flip actual's sign
            # for income so both columns read as a positive "how much",
            # same as Income Statement already does for income rows.
            sign = -1 if n["account_type"] == "income" else 1
            budgeted = n["subtotal"]
            actual = sign * actual_by_node_id[n["id"]]["subtotal"]
            out.append({
                **n, "budgeted": budgeted, "actual": actual,
                "variance": _variance_amount(actual, budgeted, pct_of_base),
                "pct_variance": _pct_variance(actual, budgeted, pct_of_base),
                "children": merge(n["children"]),
            })
        return out
    merged_roots = merge(budget_roots)

    def flatten(nodes):
        out = []
        for n in nodes:
            out.append({**n, "has_children": bool(n["children"])})
            out.extend(flatten(n["children"]))
        return out

    grouped_by_type = {}
    for t in ("income", "expense"):
        type_roots = [n for n in merged_roots if n["account_type"] == t]
        sub_budgeted = sum(n["budgeted"] for n in type_roots)
        sub_actual = sum(n["actual"] for n in type_roots)
        grouped_by_type[t] = {
            "type": t, "label": TYPE_LABELS[t], "rows": flatten(type_roots),
            "sub_budgeted": sub_budgeted, "sub_actual": sub_actual,
            "sub_variance": _variance_amount(sub_actual, sub_budgeted, pct_of_base),
            "sub_pct_variance": _pct_variance(sub_actual, sub_budgeted, pct_of_base),
        }
    grouped = [grouped_by_type["income"], grouped_by_type["expense"]]
    net_budgeted = grouped_by_type["income"]["sub_budgeted"] - grouped_by_type["expense"]["sub_budgeted"]
    net_actual = grouped_by_type["income"]["sub_actual"] - grouped_by_type["expense"]["sub_actual"]

    return {
        "grouped": grouped, "month_start": month_start.isoformat(),
        "month_end": month_end.isoformat(),
        "net_budgeted": net_budgeted, "net_actual": net_actual,
        "net_variance": _variance_amount(net_actual, net_budgeted, pct_of_base),
        "net_pct_variance": _pct_variance(net_actual, net_budgeted, pct_of_base),
    }


@app.get("/budget")
def budget_page(request: Request, scenario: str = "", month: str = "",
                ok: str = None, err: str = None, pct_of_base: int = 0):
    scens = [s for s in scenarios_all() if s["income_statement_only"]]
    scenario = scenario or (scens[0]["code"] if scens else "")
    scen = next((s for s in scens if s["code"] == scenario), None)
    month_in = month or date.today().isoformat()
    if len(month_in) == 7:  # "YYYY-MM" — what <input type="month"> and the prev/next links send
        month_in += "-01"
    month = date.fromisoformat(month_in).replace(day=1).isoformat()

    data = (_budget_rows(scenario, month, bool(pct_of_base)) if scen else
           {"grouped": [], "net_budgeted": 0, "net_actual": 0, "net_variance": 0,
            "net_pct_variance": None, "month_start": month, "month_end": month})
    return templates.TemplateResponse(request, "budget.html", {
        "nav": "budget", "scenarios": scens, "scenario": scenario, "scen": scen,
        "month": month, "prev_month": _shift_month(month, -1),
        "next_month": _shift_month(month, 1), "pct_of_base": pct_of_base, **data,
        "ok": ok, "err": err,
    })


@app.post("/budget/cell")
async def save_budget_cell(request: Request):
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
        scenario_id = int(form.get("scenario_id"))
        code = (form.get("account") or "").strip()
        period_month = form.get("period_month") or ""
        amount_raw = (form.get("amount") or "").strip()
        try:
            amount = round(float(amount_raw), 2) if amount_raw else 0.0
        except ValueError:
            raise ValueError(f"{amount_raw!r} isn't a number")
        acct = q1("SELECT id FROM accounts WHERE code = %s", (code,))
        if not acct:
            raise ValueError(f"Unknown account code: {code}")
        with tx() as cur:
            cur.execute(
                """INSERT INTO budget_lines (scenario_id, account_id, period_month, amount)
                       VALUES (%s, %s, %s, %s)
                   ON CONFLICT (scenario_id, account_id, period_month)
                       DO UPDATE SET amount = EXCLUDED.amount""",
                (scenario_id, acct["id"], period_month, amount))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return JSONResponse({"ok": False, "error": msg}, status_code=400)
    return JSONResponse({"ok": True, "amount": amount})


# ---------------------------------------------------------------------------
# Variance — budget (or any scenario) vs. actual (or any other scenario),
# rolled up to a common level so a coarse scenario (posted straight to
# "Bank") lines up against a fine one (Checking + Savings) instead of
# just not matching up at all. Scoped to full scenarios only — an
# income-statement-only one never has the journal-entry facts this reads,
# by design; see the Budget grid above for that comparison instead.
# ---------------------------------------------------------------------------
def _compute_variance(baseline: str, compare: str, level_id: str, as_of: str, zeros: int = 0,
                      pct_of_base: bool = False) -> dict:
    """Shared by the variance page and its CSV export — same rollup, same
    baseline/compare resolution, so the export matches what's on screen.
    Excludes Staging same as income-statement-only scenarios: Staging is a
    layover for entries waiting on approval, not a real balance sheet a
    user would ever want to compare against — whatever happens to be
    sitting there is incidental and temporary, not information worth
    reading a variance off of."""
    scens = [s for s in scenarios_all() if not s["income_statement_only"] and not s["is_staging"]]
    codes = [s["code"] for s in scens]
    if not compare:
        others = [s["code"] for s in scens if s["code"] != baseline]
        compare = others[0] if others else ""

    level_depth = None
    if level_id:
        lvl = q1("SELECT depth FROM account_levels WHERE id = %s", (int(level_id),))
        level_depth = lvl["depth"] if lvl else None
    elif compare:
        # Default to the comparison scenario's own base level, if it has
        # one — the natural granularity it was actually entered at. Set
        # level_id too (not just level_depth) so the picker reflects what
        # was actually used instead of silently showing "no rollup".
        bl = q1("""SELECT al.id, al.depth FROM scenarios s
                    JOIN account_levels al ON al.id = s.base_level_id
                   WHERE s.code = %s""", (compare,))
        if bl:
            level_depth = bl["depth"]
            level_id = str(bl["id"])

    as_of_date = as_of or None

    if level_depth is None:
        # Native depth — build a real account tree (same
        # _build_account_tree/_flatten_tree Trial Balance/Balance Sheet/
        # Income Statement use) instead of fn_rollup_balance(scenario,
        # NULL, ...), which already amounted to the same thing minus
        # zero-balance rows and real ancestor branches — see that
        # function's own comment ("matches fn_trial_balance's rows, just
        # without the always-show-every-postable-leaf zero rows"). Gets
        # Variance real chevrons and a working zero-balances toggle, same
        # as every other report, in the one mode where that's actually
        # meaningful: once a rollup level below has genuinely collapsed
        # several accounts' postings into one pooled number, there's
        # nothing finer left underneath to expand or reveal.
        accounts = q("SELECT * FROM v_dim_account WHERE is_active ORDER BY sort_path")
        baseline_by_id = ({r["account_id"]: r["net"] for r in
                           q("SELECT * FROM fn_account_balances(%s, %s)", (baseline, as_of_date))}
                          if baseline in codes else {})
        compare_by_id = ({r["account_id"]: r["net"] for r in
                          q("SELECT * FROM fn_account_balances(%s, %s)", (compare, as_of_date))}
                         if compare in codes else {})
        roots = _build_account_tree(accounts, baseline_by_id, compare_by_id)
        grouped = []
        for t in ACCOUNT_TYPES:
            type_roots = [r for r in roots if r["account_type"] == t]
            rows = _flatten_tree(type_roots, zeros)
            for r in rows:
                r["baseline_net"] = r["subtotal"]
                r["compare_net"] = r["compare_subtotal"]
                r["variance"] = _variance_amount(r["baseline_net"], r["compare_net"], pct_of_base)
                r["pct_variance"] = _pct_variance(r["baseline_net"], r["compare_net"], pct_of_base)
            if rows:
                sub_baseline = sum(rr["subtotal"] for rr in type_roots)
                sub_compare = sum(rr["compare_subtotal"] for rr in type_roots)
                grouped.append({
                    "type": t, "label": TYPE_LABELS[t], "rows": rows,
                    "sub_baseline": sub_baseline, "sub_compare": sub_compare,
                    "sub_variance": _variance_amount(sub_baseline, sub_compare, pct_of_base),
                    "sub_pct_variance": _pct_variance(sub_baseline, sub_compare, pct_of_base),
                })
        merged = [r for g in grouped for r in g["rows"]]
        # Roots only, not the full (branch + leaf) `merged` list — a
        # branch row's own baseline_net/compare_net already double-counts
        # its descendants, same reason Balance Sheet totals from
        # `asset_roots`/etc. rather than its own flattened display rows.
        total_baseline = sum(r["subtotal"] for r in roots)
        total_compare = sum(r["compare_subtotal"] for r in roots)
    else:
        # Rolled up to a chosen level — genuine SQL-side aggregation
        # across accounts posted at different native depths (e.g. a
        # Budget scenario posted straight to "Bank" reconciled against
        # Actual's separate Checking/Savings postings), so this stays on
        # fn_rollup_balance: no tree to walk, no zero rows to add — a
        # rolled-up row already represents whatever was pooled into it.
        baseline_rows = {r["account_id"]: r for r in q(
            "SELECT * FROM fn_rollup_balance(%s, %s, %s)",
            (baseline, level_depth, as_of_date))} if baseline in codes else {}
        compare_rows = {r["account_id"]: r for r in q(
            "SELECT * FROM fn_rollup_balance(%s, %s, %s)",
            (compare, level_depth, as_of_date))} if compare in codes else {}

        merged = []
        for aid in set(baseline_rows) | set(compare_rows):
            b = baseline_rows.get(aid)
            c = compare_rows.get(aid)
            ref = b or c
            b_net = b["net"] if b else 0
            c_net = c["net"] if c else 0
            merged.append({
                "account_code": ref["account_code"], "account_name": ref["account_name"],
                "path": ref["path"], "sort_path": ref["sort_path"], "acct_type": ref["acct_type"],
                "baseline_net": b_net, "compare_net": c_net,
                "variance": _variance_amount(b_net, c_net, pct_of_base),
                "pct_variance": _pct_variance(b_net, c_net, pct_of_base),
                "has_children": False,
            })

        grouped = []
        for t in ACCOUNT_TYPES:
            sub = sorted((r for r in merged if r["acct_type"] == t), key=lambda r: r["sort_path"])
            if sub:
                sub_baseline = sum(r["baseline_net"] for r in sub)
                sub_compare = sum(r["compare_net"] for r in sub)
                grouped.append({
                    "type": t, "label": TYPE_LABELS[t], "rows": sub,
                    "sub_baseline": sub_baseline, "sub_compare": sub_compare,
                    "sub_variance": _variance_amount(sub_baseline, sub_compare, pct_of_base),
                    "sub_pct_variance": _pct_variance(sub_baseline, sub_compare, pct_of_base),
                })
        total_baseline = sum(r["baseline_net"] for r in merged)
        total_compare = sum(r["compare_net"] for r in merged)

    return {
        "scens": scens, "compare": compare, "level_id": level_id,
        "merged": merged, "grouped": grouped, "rolled_up": level_depth is not None,
        "total_baseline": total_baseline, "total_compare": total_compare,
        "total_variance": _variance_amount(total_baseline, total_compare, pct_of_base),
        "total_pct_variance": _pct_variance(total_baseline, total_compare, pct_of_base),
    }


@app.get("/variance")
def variance_page(request: Request, baseline: str = "ACTUAL", compare: str = "",
                  level_id: str = "", as_of: str = None, zeros: int = 0,
                  pct_of_base: int = 0):
    v = _compute_variance(baseline, compare, level_id, as_of, zeros, bool(pct_of_base))

    return templates.TemplateResponse(request, "variance.html", {
        "nav": "variance", "grouped": v["grouped"], "scenarios": v["scens"],
        "levels": account_levels_all(), "baseline": baseline, "compare": v["compare"],
        "level_id": v["level_id"], "as_of": as_of or "", "zeros": zeros,
        "pct_of_base": pct_of_base, "rolled_up": v["rolled_up"],
        "total_baseline": v["total_baseline"],
        "total_compare": v["total_compare"],
        "total_variance": v["total_variance"],
        "total_pct_variance": v["total_pct_variance"],
        "today": date.today().isoformat(),
    })


@app.get("/export/variance.csv")
def variance_export_csv(baseline: str = "ACTUAL", compare: str = "",
                        level_id: str = "", as_of: str = None, zeros: int = 0,
                        pct_of_base: int = 0):
    v = _compute_variance(baseline, compare, level_id, as_of, zeros, bool(pct_of_base))
    compare = v["compare"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Code", "Account", "Path", baseline, "Variance", compare])
    for r in v["merged"]:
        w.writerow([r["account_code"], r["account_name"], r["path"],
                   r["baseline_net"], r["variance"], r["compare_net"]])
    return csv_response(buf, _variance_export_filename(baseline, compare, as_of, "csv"))


def _variance_export_filename(baseline: str, compare: str, as_of: str, ext: str) -> str:
    name = f"postwarden-variance-{baseline}-vs-{compare}"
    if as_of:
        name += f"_{as_of}"
    return f"{name}.{ext}"


@app.get("/export/variance.xlsx")
def variance_export_xlsx(baseline: str = "ACTUAL", compare: str = "",
                         level_id: str = "", as_of: str = None, zeros: int = 0,
                         pct_of_base: int = 0):
    """XLSX counterpart to variance_export_csv() above — same
    _compute_variance() data, but built from v["grouped"] (one section
    per account type, same as Trial Balance) rather than the CSV's flat
    v["merged"] list, so it can add the section headers, the per-type
    subtotal, and a real % Variance column the CSV leaves out (the
    figure's already computed either way — _pct_variance() — CSV just
    never had a use for it without a companion section structure).

    Native mode (no rollup level) builds a real account tree, same as
    every other tree-shaped report here — a section's own depth-1 row
    gets the "group" bold+ruled treatment, and an explicit subtotal row
    only when a section actually has more than one top-level account
    (Trial Balance's own show_type_total reasoning). Rolled-up mode
    (`level_id` set) has no tree at all — v["merged"]'s rows are already
    one flat pooled figure per account at that level, so every row stays
    "line" style and every section always gets its own subtotal row,
    since no single row in a rolled-up section is ever "the" total the
    way a tree's own root row is.

    Variance and % Variance are live formulas (_xlsx_variance_formulas),
    same as Income Statement's own xlsx export — each one only
    references its own row's baseline/compare cells, never a range, so
    it's safe regardless of whether that row's own baseline/compare
    figures are a leaf balance, a rolled-up root, or a section subtotal."""
    v = _compute_variance(baseline, compare, level_id, as_of, zeros, bool(pct_of_base))
    compare = v["compare"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Variance"

    subtitle = f"{baseline} vs. {compare}"
    subtitle += f" · {'As of ' + as_of if as_of else 'Through today'}"
    if v["rolled_up"]:
        subtitle += " · rolled up"
    ws.cell(row=1, column=1, value="Variance").font = _XLSX_TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _XLSX_SUBTITLE_FONT

    headers = ["Code", "Account", baseline, "Variance", "% Variance", compare]
    n_cols = len(headers)
    header_row, data_start = 4, 5
    _xlsx_header_row(ws, header_row, headers)

    def row(r: int, code, name, depth, base, variance, pct, comp, style="line"):
        # variance/pct unused — Variance/% Variance are live formulas
        # instead (_xlsx_variance_formulas), same as Income Statement's
        # own xlsx export — C (base/baseline), then F (compare), same
        # canonical order _compute_variance's own _variance_amount(base
        # line_net, compare_net, ...) calls use.
        var_f, pct_f = _xlsx_variance_formulas(f"C{r}", f"F{r}", bool(pct_of_base))
        value_cols = [(3, base, _XLSX_MONEY_FMT), (4, var_f, _XLSX_MONEY_FMT),
                      (5, pct_f, _XLSX_PCT_FMT), (6, comp, _XLSX_MONEY_FMT)]
        _xlsx_data_row(ws, r, [(1, code), (2, name)], value_cols, style, max(depth - 1, 0))

    r = data_start
    for g in v["grouped"]:
        _xlsx_data_row(ws, r, [(1, ""), (2, g["label"])], [], style="group")
        r += 1
        top_level_count = sum(1 for line in g["rows"] if not v["rolled_up"] and line.get("depth") == 1)
        for line in g["rows"]:
            is_root = not v["rolled_up"] and line.get("depth") == 1
            row(r, line["account_code"], line["account_name"], line.get("depth", 1),
                line["baseline_net"], line["variance"], line["pct_variance"], line["compare_net"],
                style="group" if is_root else "line")
            r += 1
        if v["rolled_up"] or top_level_count > 1:
            row(r, "", f"{g['label']} subtotal", 1, g["sub_baseline"], g["sub_variance"],
                g["sub_pct_variance"], g["sub_compare"], style="subtotal")
            r += 1
    row(r, "", "Total", 1, v["total_baseline"], v["total_variance"], v["total_pct_variance"],
        v["total_compare"], style="grand")
    last_row = r
    _xlsx_variance_coloring(ws, 4, data_start, last_row)  # Variance
    _xlsx_variance_coloring(ws, 5, data_start, last_row)  # % Variance

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"C{data_start}"
    ws.sheet_view.showGridLines = False
    return xlsx_response(wb, _variance_export_filename(baseline, compare, as_of, "xlsx"))


# ---------------------------------------------------------------------------
# Chart of accounts
# ---------------------------------------------------------------------------
# Seed.sql's own convention (1xxx assets, 2xxx liabilities, ...) isn't
# DB-enforced, just a habit — quick-created accounts follow it anyway so
# the chart stays legible without asking the user to think about codes.
ACCOUNT_TYPE_CODE_PREFIX = {
    "asset": "1", "liability": "2", "equity": "3", "income": "4", "expense": "5",
}


def _next_account_code(account_type: str) -> str:
    prefix = ACCOUNT_TYPE_CODE_PREFIX[account_type]
    existing = {int(r["code"]) for r in q(
        "SELECT code FROM accounts WHERE code LIKE %s", (prefix + "%",))}
    candidate = (max(existing) + 10) if existing else int(prefix + "000")
    while candidate in existing:
        candidate += 1
    return str(candidate)


def _accounts_with_gaps(accounts: list[dict]) -> list[dict]:
    """Interleaves a "gap" placeholder before every account row (and one
    trailing) so accounts.html can render an Actual Budget-style "+" for
    adding a category between any two adjacent rows, instead of only via
    the form at the bottom.

    Deliberately does NOT decide each gap's parent_id/account_type here —
    which two rows are visually adjacent around a given gap depends on
    which summary accounts are currently collapsed, and that's a client-
    side (localStorage) preference this request has no way to see. Each
    gap just needs to know which account row to track for visibility
    (see accounts.js); accounts.js computes the actual parent/type at
    "+"-click time from whichever rows are visible right then."""
    out = []
    prev = None
    for acct in accounts:
        out.append({"kind": "gap", "track_id": acct["id"]})
        out.append({"kind": "account", **acct})
        prev = acct
    out.append({"kind": "gap", "track_id": prev["id"] if prev else None})
    return out


@app.get("/accounts")
def accounts_page(request: Request, level_id: str = "", ok: str = None, err: str = None):
    accounts = q("SELECT * FROM v_dim_account ORDER BY sort_path")
    selected_level = q1("SELECT * FROM account_levels WHERE id = %s",
                        (int(level_id),)) if level_id else None
    if selected_level:
        display_accounts = [a for a in accounts if a["depth"] == selected_level["depth"]]
        rows = None
    else:
        display_accounts = accounts
        rows = _accounts_with_gaps(accounts)
    return templates.TemplateResponse(request, "accounts.html", {
        "nav": "accounts", "accounts": accounts, "rows": rows,
        "display_accounts": display_accounts,
        "levels": account_levels_all(), "selected_level": selected_level,
        "account_types": ACCOUNT_TYPES, "type_labels": TYPE_LABELS,
        "ok": ok, "err": err,
    })


@app.post("/accounts")
def create_account(request: Request, code: str = Form(...), name: str = Form(...),
                   account_type: str = Form(...), parent_id: str = Form(""),
                   is_postable: str = Form(None), is_cashflow: str = Form(None),
                   csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                """INSERT INTO accounts (code, name, account_type, parent_id, is_postable, is_cashflow)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (code.strip(), name.strip(), account_type,
                 int(parent_id) if parent_id else None,
                 is_postable is not None, is_cashflow is not None),
            )
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/accounts", err=msg)
    return flash_redirect("/accounts", ok=f"Account {code} — {name} created")


@app.post("/accounts/quick-create")
def quick_create_account(request: Request, name: str = Form(...),
                         parent_id: str = Form(""), account_type: str = Form(""),
                         is_postable: str = Form(None), csrf_token: str = Form(...)):
    # Powers the "+" between two rows on /accounts (see accounts.js) — the
    # code is generated, not typed, same spirit as Actual Budget's category
    # picker not showing account numbers at all; /accounts's own bottom
    # form is still there for anyone who wants to set an exact code.
    try:
        require_csrf(request, csrf_token)
        name = name.strip()
        if not name:
            raise ValueError("Name is required")
        pid = int(parent_id) if parent_id else None
        if pid:
            parent = q1("SELECT account_type FROM accounts WHERE id = %s", (pid,))
            if not parent:
                raise ValueError("Unknown parent account")
            acct_type = parent["account_type"]
        else:
            acct_type = account_type
            if acct_type not in ACCOUNT_TYPES:
                raise ValueError("Choose an account type")
        code = _next_account_code(acct_type)
        with tx() as cur:
            cur.execute(
                """INSERT INTO accounts (code, name, account_type, parent_id, is_postable)
                   VALUES (%s, %s, %s, %s, %s)""",
                (code, name, acct_type, pid, is_postable is not None))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/accounts", err=msg)
    return flash_redirect("/accounts", ok=f"Account {code} — {name} created")


@app.post("/accounts/{account_id}/toggle-active")
def toggle_account(account_id: int, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                "UPDATE accounts SET is_active = NOT is_active WHERE id = %s",
                (account_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/accounts", err=msg)
    return flash_redirect("/accounts", ok="Account updated")


@app.post("/accounts/{account_id}/toggle-cashflow")
def toggle_account_cashflow(account_id: int, request: Request, csrf_token: str = Form(...)):
    """Flips accounts.is_cashflow — the Cash Flow Statement's own cash
    boundary (SPEC.md decision 20). Same shape as toggle-active above;
    a separate route rather than folding this into a generic "PATCH one
    boolean column" endpoint since that's not a pattern this app has
    anywhere else, and two clearly-named routes read better than one
    parameterized by column name typed into a form."""
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                "UPDATE accounts SET is_cashflow = NOT is_cashflow WHERE id = %s",
                (account_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/accounts", err=msg)
    return flash_redirect("/accounts", ok="Account updated")


# ---------------------------------------------------------------------------
# Journal entry — new entries are created inline on the Journal now (see
# entries_page below); this URL is kept as a redirect so old links and
# bookmarks land somewhere useful instead of 404ing.
# ---------------------------------------------------------------------------
@app.get("/entries/new")
def entry_new_redirect():
    return RedirectResponse("/entries?new=1", status_code=303)


TAG_PATTERN = re.compile(r"^[a-z0-9][a-z0-9 _-]{0,39}$")


def all_tags() -> list[str]:
    # Active only — same reasoning as payees' active_payees query below:
    # this feeds the tag-input's autocomplete (see tags.js), so an
    # archived tag stops being offered as something new to pick, without
    # touching what already carries it (tags_by_entry, the per-entry
    # badges, never filters on is_active at all).
    return [r["name"] for r in q("SELECT name FROM tags WHERE is_active ORDER BY name")]


def _parse_tags(raw: str) -> list[str]:
    """Comma-separated tag names from the tag-input widget (see tags.js) ->
    a clean, deduped, validated list — matches tags.name's CHECK constraint
    so a bad tag fails here with a plain message instead of a raw
    constraint-violation error."""
    seen = []
    for piece in (raw or "").split(","):
        name = piece.strip().lower()
        if not name or name in seen:
            continue
        if not TAG_PATTERN.match(name):
            raise ValueError(
                f"Invalid tag {name!r}: letters, numbers, spaces, - and _ only, max 40 chars")
        seen.append(name)
    return seen


def _sync_tags(cur, table: str, id_col: str, obj_id: int, tag_names: list[str]) -> None:
    """Shared by journal entries and scheduled entries — both have their own
    entry_id/tag_id junction table with the same shape."""
    cur.execute(f"DELETE FROM {table} WHERE {id_col} = %s", (obj_id,))
    for name in tag_names:
        # ON CONFLICT ... SET is_active = TRUE (not just the no-op `name
        # = EXCLUDED.name` this used before is_active existed) — same
        # reasoning as quick_create_payee: typing an existing tag's name
        # here is exactly the signal that it's back in use, so this
        # quietly reactivates one that was archived rather than leaving
        # it archived-but-now-attached-to-something, which would be a
        # tag with no way back into its own suggestion list.
        cur.execute(
            """INSERT INTO tags (name) VALUES (%s)
               ON CONFLICT (name) DO UPDATE SET is_active = TRUE
               RETURNING id""",
            (name,))
        tag_id = cur.fetchone()["id"]
        cur.execute(
            f"INSERT INTO {table} ({id_col}, tag_id) VALUES (%s, %s)",
            (obj_id, tag_id))


def _sync_entry_tags(cur, entry_id: str, tag_names: list[str]) -> None:
    _sync_tags(cur, "journal_entry_tags", "entry_id", entry_id, tag_names)


def _add_tag_to_entries(entry_ids: list[str], tag_name: str) -> None:
    """Adds one tag to every given entry that doesn't already have it —
    the Journal's bulk 'Edit tags' popup (see entries-select.js), one
    call per chip added. Deliberately additive, not _sync_entry_tags'
    full replace: different selected entries can have different
    existing tags, and this should only ever touch the one tag actually
    being added, leaving everything else on every entry alone.
    journal_entry_tags carries no immutability trigger — organizing a
    posted entry isn't the append-only rule tags editing would actually
    violate (see SPEC.md's tag-editing decision), so this works the
    same on a posted entry as a pending one."""
    with tx() as cur:
        # Reactivates on conflict — same reasoning as _sync_tags above.
        cur.execute(
            """INSERT INTO tags (name) VALUES (%s)
               ON CONFLICT (name) DO UPDATE SET is_active = TRUE
               RETURNING id""",
            (tag_name,))
        tag_id = cur.fetchone()["id"]
        for entry_id in entry_ids:
            cur.execute(
                """INSERT INTO journal_entry_tags (entry_id, tag_id) VALUES (%s, %s)
                   ON CONFLICT DO NOTHING""",
                (entry_id, tag_id))


def _remove_tag_from_entries(entry_ids: list[str], tag_name: str) -> None:
    """The other half of the bulk tag popup — one call per chip removed,
    dropping that one tag from whichever of the given entries actually
    have it. A tag nobody uses anymore is just left in `tags`, same as
    everywhere else in the app that removes a tag from something."""
    with tx() as cur:
        cur.execute(
            """DELETE FROM journal_entry_tags
                WHERE entry_id = ANY(%s)
                  AND tag_id = (SELECT id FROM tags WHERE name = %s)""",
            (entry_ids, tag_name))


def _parse_lines(form) -> list[dict]:
    """Turn parallel account[]/debit[]/credit[]/memo[] arrays into line dicts.

    Rules mirror the paper form: a line needs an account and exactly one of
    debit or credit, strictly positive. Blank rows are ignored. The account
    field is a <select> (searchable combobox in the UI — see app.js), so
    its value is already a bare code; no "code · name" text to split here
    the way a free-text field would need — this is just defense in depth
    against a client that isn't the browser UI.
    """
    accounts = form.getlist("account")
    debits = form.getlist("debit")
    credits = form.getlist("credit")
    memos = form.getlist("memo")
    lines = []
    for i, acct in enumerate(accounts):
        code = (acct or "").strip()
        d = (debits[i] if i < len(debits) else "").strip()
        c = (credits[i] if i < len(credits) else "").strip()
        memo = (memos[i] if i < len(memos) else "").strip() or None
        if not code and not d and not c:
            continue  # blank row
        if not code:
            raise ValueError(f"Line {i + 1}: missing account")
        try:
            dv = float(d) if d else 0.0
            cv = float(c) if c else 0.0
        except ValueError:
            raise ValueError(f"Line {i + 1}: debit and credit must be numbers")
        if dv < 0 or cv < 0:
            raise ValueError(f"Line {i + 1}: amounts must be positive")
        if (dv > 0) == (cv > 0):
            raise ValueError(
                f"Line {i + 1}: enter exactly one of debit or credit")
        lines.append({"code": code, "amount": round(dv - cv, 2), "memo": memo})
    if not lines:
        raise ValueError("The entry has no lines")
    return lines


@app.post("/entries")
async def create_entry(request: Request):
    form = await request.form()
    # The entry-grid page submits via fetch() so a rejected entry (e.g. a
    # bad account code) doesn't lose the lines the user already typed —
    # it asks for JSON back instead of a redirect. See app.js.
    wants_json = "application/json" in request.headers.get("accept", "")
    try:
        require_csrf(request, form.get("csrf_token"))
        lines = _parse_lines(form)
        entry_date = form.get("entry_date") or date.today().isoformat()
        scenario_id = int(form.get("scenario_id"))
        description = (form.get("description") or "").strip()
        reference = (form.get("reference") or "").strip() or None
        tag_names = _parse_tags(form.get("tags", ""))
        payee_id = int(form.get("payee_id")) if form.get("payee_id") else None
        if not description:
            raise ValueError("Description is required")

        codes = {ln["code"] for ln in lines}
        found = {r["code"] for r in q(
            "SELECT code FROM accounts WHERE code = ANY(%s)", (list(codes),))}
        missing = codes - found
        if missing:
            raise ValueError(f"Unknown account code: {', '.join(sorted(missing))}")

        with tx() as cur:
            cur.execute(
                """INSERT INTO journal_entries
                       (scenario_id, entry_date, description, reference,
                        payee_id, created_by_user_id)
                   VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                (scenario_id, entry_date, description, reference, payee_id,
                 auth.current_user(request)["user_id"]))
            entry_id = cur.fetchone()["id"]
            for n, ln in enumerate(lines, start=1):
                cur.execute(
                    """INSERT INTO journal_lines
                           (entry_id, line_no, account_id, amount, memo)
                       VALUES (%s, %s,
                               (SELECT id FROM accounts WHERE code = %s),
                               %s, %s)""",
                    (entry_id, n, ln["code"], ln["amount"], ln["memo"]))
            if tag_names:
                _sync_entry_tags(cur, entry_id, tag_names)
        # the deferred trigger has now blessed the entry at COMMIT
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        if wants_json:
            return JSONResponse({"ok": False, "error": msg}, status_code=400)
        return flash_redirect("/entries?new=1", err=msg)
    ok_msg = f"Entry #{entry_id} posted"
    if wants_json:
        return JSONResponse({"ok": True, "redirect": flash_url("/entries", ok=ok_msg)})
    return flash_redirect("/entries", ok=ok_msg)


# ---------------------------------------------------------------------------
# Journal browser
# ---------------------------------------------------------------------------
ENTRIES_PAGE_SIZE = 50


AMOUNT_OPS = {
    "gte": ">=", "lte": "<=", "gt": ">", "lt": "<", "eq": "=",
}


def _shared_journal_filters(where: list[str], params: list, date_from: str, date_to: str,
                            qtext: str, tag_list: list[str], account: str, payee: str,
                            amount_op: str, amount_value: str, amount_value2: str) -> None:
    """WHERE-clause fragments the Journal and Staging's filter bars share —
    date range, free-text search, tags, account, payee, and the amount
    operator (including 'between') — appended in place to where/params.
    Split out so Staging's own filters (added alongside its own bulk
    Reject/filter bar — see pending_staging_entries()) reuse the exact
    same logic instead of a second, easy-to-drift copy of it; only what's
    genuinely different per caller (the Journal's own scenario/
    hide_reversed, Staging's target-scenario) stays out of here."""
    if date_from:
        where.append("e.entry_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("e.entry_date <= %s")
        params.append(date_to)
    if qtext:
        where.append("(e.description ILIKE %s OR e.reference ILIKE %s)")
        params.extend([f"%{qtext}%", f"%{qtext}%"])
    if tag_list:
        # ANY of the given tags — a broadening filter, like most tag UIs.
        where.append("""e.id IN (SELECT jet.entry_id FROM journal_entry_tags jet
                                   JOIN tags tg ON tg.id = jet.tag_id
                                  WHERE tg.name = ANY(%s))""")
        params.append(tag_list)
    if account:
        # Powers the "click an amount on Income Statement/Balance Sheet to
        # see what posted to it" links (one specific account, exact code)
        # as well as the Journal's own Account filter dropdown.
        where.append("""e.id IN (SELECT jl.entry_id FROM journal_lines jl
                                   JOIN accounts a ON a.id = jl.account_id
                                  WHERE a.code = %s)""")
        params.append(account)
    if payee:
        # Powers the same click-through from Payees' entry count — both
        # queries this feeds already LEFT JOIN payees p ON e.payee_id, so
        # this can reference p.name directly rather than a subquery.
        where.append("p.name = %s")
        params.append(payee)
    if amount_op == "between" and amount_value and amount_value2:
        try:
            lo, hi = float(amount_value), float(amount_value2)
        except ValueError:
            lo = hi = None  # a hand-edited URL with garbage; just ignore it
        if lo is not None:
            # Bounds get sorted rather than trusted in whichever order
            # the two fields happen to hold — "between 500 and 100" reads
            # the same as "between 100 and 500" to anyone typing it.
            lo, hi = sorted((round(lo, 2), round(hi, 2)))
            where.append("""(SELECT COALESCE(SUM(l.debit), 0) FROM journal_lines l
                               WHERE l.entry_id = e.id) BETWEEN %s AND %s""")
            params.extend([lo, hi])
    elif amount_op in AMOUNT_OPS and amount_value:
        try:
            amount_num = float(amount_value)
        except ValueError:
            amount_num = None  # a hand-edited URL with garbage; just ignore it
        if amount_num is not None:
            # "Amount" for a whole entry is its total debit (= total credit,
            # by the balance invariant, when the scenario enforces one) —
            # the same figure the Journal already shows per entry. The
            # operator is looked up from a fixed whitelist, never
            # interpolated from user input directly, before it ever
            # touches the query string.
            op = AMOUNT_OPS[amount_op]
            where.append(f"""(SELECT COALESCE(SUM(l.debit), 0) FROM journal_lines l
                               WHERE l.entry_id = e.id) {op} %s""")
            params.append(round(amount_num, 2))


def _entries_filter(scenario: str, date_from: str, date_to: str, qtext: str,
                    tags: str, account: str = "", payee: str = "",
                    amount_op: str = "", amount_value: str = "", amount_value2: str = "",
                    hide_reversed: int = 0) -> tuple[list[str], list, list[str]]:
    """Shared by the paged HTML view and the CSV export — same filters,
    same WHERE clause, so what you see is exactly what you export."""
    try:
        tag_list = _parse_tags(tags) if tags else []
    except ValueError:
        tag_list = []  # a hand-edited URL with a malformed tag; just ignore it
    # Unconditional, not just "no scenario filter selected" — the Journal
    # is exclusively for posted, non-staging scenarios (Staging is its own
    # page with its own filter bar). A scenario filter value narrows
    # further within that; it never widens back into Staging, and this
    # holds even against a hand-edited query string.
    where, params = ["NOT s.is_staging"], []
    if scenario:
        where.append("s.code = %s")
        params.append(scenario)
    _shared_journal_filters(where, params, date_from, date_to, qtext, tag_list,
                            account, payee, amount_op, amount_value, amount_value2)
    if hide_reversed:
        # Excludes both halves of a reversal pair: the reversal itself
        # (reverses_entry_id set) and whatever it reversed (some other
        # entry's reverses_entry_id points back at this one) — someone
        # hiding "stuff I created by accident" wants neither the mistake
        # nor its own cleanup cluttering the view.
        where.append("""e.reverses_entry_id IS NULL
                         AND NOT EXISTS (SELECT 1 FROM journal_entries r
                                          WHERE r.reverses_entry_id = e.id)""")
    return where, params, tag_list


@app.get("/entries")
def entries_page(request: Request, scenario: str = "", date_from: str = "",
                 date_to: str = "", qtext: str = "", tags: str = "", account: str = "",
                 payee: str = "", amount_op: str = "", amount_value: str = "",
                 amount_value2: str = "", hide_reversed: int = 0, back: str = "", page: int = 1,
                 ok: str = None, err: str = None):
    page = max(page, 1)
    # Only ever a same-origin relative path — a bare "/x", never "//x"
    # (protocol-relative, i.e. an off-site redirect) or an absolute URL.
    if not back.startswith("/") or back.startswith("//"):
        back = ""
    where, params, tag_list = _entries_filter(scenario, date_from, date_to, qtext, tags,
                                              account, payee, amount_op, amount_value,
                                              amount_value2, hide_reversed)
    # Whether *any* filter is actually narrowing the list right now — not
    # "back"/"page" (navigation state, not a filter). Powers the "Clear
    # filters" link below: no point showing it over a plain, unfiltered
    # Journal. account/payee count too, even though each already has its
    # own scoped "clear" link right above the bar — this is the "reset
    # everything" version of the same idea.
    has_filters = bool(scenario or date_from or date_to or qtext or tag_list
                       or account or payee or amount_op or hide_reversed)
    clear_filters_qs = urlencode({"back": back}) if back else ""
    account_row = q1("SELECT code, name FROM accounts WHERE code = %s", (account,)) if account else None
    payee_row = q1("SELECT name FROM payees WHERE name = %s", (payee,)) if payee else None

    entries = q(f"""
        SELECT e.id, e.entry_date, e.description, e.reference,
               e.reverses_entry_id, s.code AS scenario_code,
               s.is_locked AS scenario_locked, u.username AS posted_by,
               p.name AS payee_name,
               (SELECT COALESCE(SUM(l.debit), 0) FROM journal_lines l
                 WHERE l.entry_id = e.id) AS total_debits,
               (SELECT COALESCE(SUM(l.credit), 0) FROM journal_lines l
                 WHERE l.entry_id = e.id) AS total_credits,
               (SELECT r.id FROM journal_entries r
                 WHERE r.reverses_entry_id = e.id LIMIT 1) AS reversed_by
          FROM journal_entries e
          JOIN scenarios s ON s.id = e.scenario_id
          LEFT JOIN users u ON u.id = e.created_by_user_id
          LEFT JOIN payees p ON p.id = e.payee_id
         WHERE {' AND '.join(where)}
         ORDER BY e.entry_date DESC, e.seq DESC
         LIMIT %s OFFSET %s""",
        params + [ENTRIES_PAGE_SIZE + 1, (page - 1) * ENTRIES_PAGE_SIZE])
    # Fetched one extra row purely to know whether a next page exists —
    # trim it back off before it's ever shown.
    has_next = len(entries) > ENTRIES_PAGE_SIZE
    entries = entries[:ENTRIES_PAGE_SIZE]

    ids = [e["id"] for e in entries]
    lines_by_entry = {}
    tags_by_entry = {}
    if ids:
        for ln in q("""SELECT l.entry_id, l.line_no, l.debit, l.credit,
                              l.memo, a.code AS account_code,
                              a.name AS account_name
                         FROM journal_lines l
                         JOIN accounts a ON a.id = l.account_id
                        WHERE l.entry_id = ANY(%s)
                        ORDER BY l.entry_id, l.line_no""", (ids,)):
            lines_by_entry.setdefault(ln["entry_id"], []).append(ln)
        for tg in q("""SELECT jet.entry_id, tg.name
                         FROM journal_entry_tags jet
                         JOIN tags tg ON tg.id = jet.tag_id
                        WHERE jet.entry_id = ANY(%s)
                        ORDER BY tg.name""", (ids,)):
            tags_by_entry.setdefault(tg["entry_id"], []).append(tg["name"])

    common_qs = {
        "scenario": scenario, "date_from": date_from, "date_to": date_to,
        "qtext": qtext, "tags": tags, "amount_op": amount_op, "amount_value": amount_value,
        "amount_value2": amount_value2,
        "hide_reversed": hide_reversed, "back": back,
    }
    export_qs = urlencode({**common_qs, "account": account, "payee": payee})
    clear_account_qs = urlencode({**common_qs, "payee": payee})
    clear_payee_qs = urlencode({**common_qs, "account": account})

    # For the "+ New entry" panel inline below the filters — same data
    # entry_new.html used to fetch as its own page, since creating an
    # entry now happens right here instead. Staging never accepts a manual
    # entry (fn_staging_manual_entry_guard) — only a schedule or an import
    # lands there — so it's excluded here the same as a locked or
    # income-statement-only scenario would be.
    new_entry_scenarios = [s for s in scenarios_all()
                          if not s["is_locked"] and not s["income_statement_only"]
                          and not s["is_staging"]]
    accounts_by_scenario = postable_accounts_by_scenario()
    new_entry_accounts = (accounts_by_scenario.get(new_entry_scenarios[0]["id"], [])
                          if new_entry_scenarios else [])
    active_payees = q("SELECT id, name FROM payees WHERE is_active ORDER BY name")
    # Filter dropdowns, unlike the New entry panel's own pickers, need
    # every account/payee ever usable historically — not just what's
    # postable/active right now, since a past entry can reference either.
    filter_accounts = postable_accounts_for_pickers()
    filter_payee_names = [r["name"] for r in q("SELECT name FROM payees ORDER BY name")]

    return templates.TemplateResponse(request, "entries.html", {
        "nav": "entries", "entries": entries, "lines_by_entry": lines_by_entry,
        "tags_by_entry": tags_by_entry, "tags": tags, "all_tags": all_tags(),
        # The Journal never shows Staging entries (see _entries_filter's
        # unconditional NOT s.is_staging), so its own Scenario filter
        # shouldn't offer Staging as an option either — nothing it could
        # ever actually filter to.
        "scenarios": [s for s in scenarios_all() if not s["is_staging"]], "scenario": scenario,
        "date_from": date_from, "date_to": date_to, "qtext": qtext,
        "account": account, "account_row": account_row, "clear_account_qs": clear_account_qs,
        "payee": payee, "payee_row": payee_row, "clear_payee_qs": clear_payee_qs,
        "amount_op": amount_op, "amount_value": amount_value, "amount_value2": amount_value2,
        "amount_ops": AMOUNT_OPS, "hide_reversed": hide_reversed,
        "has_filters": has_filters, "clear_filters_qs": clear_filters_qs,
        "filter_accounts": filter_accounts, "filter_payee_names": filter_payee_names,
        "back": back,
        "page": page, "page_size": ENTRIES_PAGE_SIZE,
        "has_next": has_next, "has_prev": page > 1, "export_qs": export_qs,
        "new_entry_scenarios": new_entry_scenarios, "new_entry_accounts": new_entry_accounts,
        "accounts_by_scenario": accounts_by_scenario, "payees": active_payees,
        "tpls": templates_full(), "today": date.today().isoformat(),
        "ok": ok, "err": err,
    })


@app.get("/entries/export.csv")
def entries_export_csv(scenario: str = "", date_from: str = "", date_to: str = "",
                       qtext: str = "", tags: str = "", account: str = "", payee: str = "",
                       amount_op: str = "", amount_value: str = "", amount_value2: str = "",
                       hide_reversed: int = 0):
    """Every entry matching the current filters (not just the current
    page) — one row per journal line, so it opens straight into a
    spreadsheet without the entry/line grouping the HTML view has."""
    where, params, _ = _entries_filter(scenario, date_from, date_to, qtext, tags,
                                       account, payee, amount_op, amount_value,
                                       amount_value2, hide_reversed)
    rows = q(f"""
        SELECT e.id AS entry_id, e.entry_date, s.code AS scenario_code,
               e.description, e.reference, p.name AS payee_name,
               a.code AS account_code, a.name AS account_name,
               l.debit, l.credit, l.memo
          FROM journal_lines l
          JOIN journal_entries e ON e.id = l.entry_id
          JOIN scenarios s ON s.id = e.scenario_id
          JOIN accounts a ON a.id = l.account_id
          LEFT JOIN payees p ON p.id = e.payee_id
         WHERE {' AND '.join(where)}
         ORDER BY e.entry_date DESC, e.seq DESC, l.line_no""", params)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Entry #", "Date", "Scenario", "Description", "Reference",
               "Payee", "Account code", "Account name", "Debit", "Credit", "Memo"])
    for r in rows:
        w.writerow([r["entry_id"], r["entry_date"], r["scenario_code"],
                   r["description"], r["reference"] or "", r["payee_name"] or "",
                   r["account_code"], r["account_name"],
                   r["debit"] or "", r["credit"] or "", r["memo"] or ""])
    return csv_response(buf, "postwarden-journal.csv")


def _reverse_one_entry(entry_id: str, user_id: int) -> str:
    """Posts the reversing entry for one already-posted entry — the actual
    work behind both the single-entry route and the bulk one below, so
    there's exactly one place that knows what a reversal looks like.
    Raises ValueError for anything that isn't a straightforward reversal
    (not found, already reversed); a locked scenario or similar still
    surfaces as psycopg.Error from the database itself. Returns the new
    entry's id."""
    orig = q1("""SELECT e.*, s.code AS scenario_code FROM journal_entries e
                 JOIN scenarios s ON s.id = e.scenario_id
                 WHERE e.id = %s""", (entry_id,))
    if not orig:
        raise ValueError(f"Entry #{entry_id} not found")
    already = q1("SELECT id FROM journal_entries WHERE reverses_entry_id = %s",
                 (entry_id,))
    if already:
        raise ValueError(f"Entry #{entry_id} was already reversed by #{already['id']}")
    with tx() as cur:
        cur.execute(
            """INSERT INTO journal_entries
                   (scenario_id, entry_date, description, reference,
                    reverses_entry_id, payee_id, created_by_user_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id""",
            (orig["scenario_id"], date.today(),
             f"Reversal of #{entry_id} — {orig['description']}",
             orig["reference"], entry_id, orig["payee_id"], user_id))
        new_id = cur.fetchone()["id"]
        cur.execute(
            """INSERT INTO journal_lines
                   (entry_id, line_no, account_id, amount, memo)
               SELECT %s, line_no, account_id, -amount, memo
                 FROM journal_lines WHERE entry_id = %s""",
            (new_id, entry_id))
        # Carry the original's tags over — a reversal is still "about"
        # whatever it was tagged for.
        cur.execute(
            """INSERT INTO journal_entry_tags (entry_id, tag_id)
               SELECT %s, tag_id FROM journal_entry_tags WHERE entry_id = %s""",
            (new_id, entry_id))
    return new_id


@app.post("/entries/{entry_id}/reverse")
def reverse_entry(entry_id: str, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        new_id = _reverse_one_entry(entry_id, auth.current_user(request)["user_id"])
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/entries", err=msg)
    return flash_redirect("/entries",
                          ok=f"Entry #{entry_id} reversed by #{new_id}")


@app.post("/entries/reverse")
async def reverse_entries_bulk(request: Request):
    """Bulk sibling of /entries/{id}/reverse, for the Journal's own
    'select entries' mode — same shape as Staging's bulk Approve/Reject:
    loop _reverse_one_entry over whatever's checked, collect successes
    and errors separately so one already-reversed or locked-scenario
    entry in the batch doesn't stop the rest from going through."""
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
    except ValueError as e:
        return flash_redirect("/entries", err=str(e))

    entry_ids = [v for v in form.getlist("entry_id") if v]
    if not entry_ids:
        return flash_redirect("/entries", err="Select at least one entry to reverse")

    user_id = auth.current_user(request)["user_id"]
    reversed_ids, errors = [], []
    for eid in entry_ids:
        try:
            reversed_ids.append(_reverse_one_entry(eid, user_id))
        except (ValueError, psycopg.Error) as e:
            errors.append(_pg_msg(e) if isinstance(e, psycopg.Error) else str(e))

    ok_msg = (f"Reversed {len(reversed_ids)} entr{'y' if len(reversed_ids) == 1 else 'ies'}"
             if reversed_ids else None)
    err_msg = "; ".join(errors) or None
    return flash_redirect("/entries", ok=ok_msg, err=err_msg)


@app.post("/entries/tags")
async def edit_entries_tags(request: Request):
    """Add or remove one tag across whatever's checked in the Journal's
    'select entries' mode — the 'Edit tags' popup (see entries-select.js)
    fires one of these per chip added/removed, live, since the popup has
    no Save button of its own to batch a set of changes behind. JSON in
    and out (fetch-driven, not a real page navigation) rather than the
    flash-redirect every other bulk action here uses."""
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
        entry_ids = [v for v in form.getlist("entry_id") if v]
        if not entry_ids:
            raise ValueError("No entries selected")
        action = form.get("action")
        tag_names = _parse_tags(form.get("tag", ""))
        if len(tag_names) != 1:
            raise ValueError("Expected exactly one tag")
        tag_name = tag_names[0]
        if action == "add":
            _add_tag_to_entries(entry_ids, tag_name)
        elif action == "remove":
            _remove_tag_from_entries(entry_ids, tag_name)
        else:
            raise ValueError(f"Unknown action {action!r}")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return JSONResponse({"ok": False, "error": msg}, status_code=400)
    return JSONResponse({"ok": True, "tag": tag_name, "action": action})


@app.post("/entries/{entry_id}/edit-description")
def edit_entry_description(entry_id: str, request: Request,
                           description: str = Form(...), csrf_token: str = Form(...)):
    """A typo in a posted entry's description is exactly the kind of
    mistake decision 4's append-only rule was never meant to trap
    someone with — same reasoning as tags (SPEC.md decision 16):
    organizational, not a fact about the transaction, so it's fair game
    to fix on something already posted. fn_entries_guard already allowed
    changing description/reference on a posted entry from the day it was
    written (only scenario/date/reverses_entry_id are actually blocked);
    this is the first route that exercises that door. Amounts, accounts,
    and every journal_lines row stay exactly as immutable as ever — this
    touches journal_entries.description and nothing else."""
    description = description.strip()
    try:
        require_csrf(request, csrf_token)
        if not description:
            raise ValueError("Description can't be empty")
        with tx() as cur:
            cur.execute("UPDATE journal_entries SET description = %s WHERE id = %s",
                       (description, entry_id))
            if cur.rowcount == 0:
                raise ValueError(f"Entry #{entry_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/entries", err=msg)
    return flash_redirect("/entries", ok=f"Entry #{entry_id} updated")


# ---------------------------------------------------------------------------
# Scenarios
# ---------------------------------------------------------------------------
@app.get("/scenarios")
def scenarios_page(request: Request, ok: str = None, err: str = None):
    return templates.TemplateResponse(request, "scenarios.html", {
        "nav": "scenarios", "scenarios": scenarios_all(),
        "levels": account_levels_all(),
        "scenario_types": SCENARIO_TYPES, "ok": ok, "err": err,
    })


@app.post("/scenarios")
def create_scenario(request: Request, code: str = Form(...), name: str = Form(...),
                    scenario_type: str = Form(...),
                    enforce_balance: str = Form(None),
                    income_statement_only: str = Form(None),
                    base_level_id: str = Form(""),
                    notes: str = Form(""), csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                """INSERT INTO scenarios
                       (code, name, scenario_type, enforce_balance,
                        income_statement_only, base_level_id, notes)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (code.strip().upper(), name.strip(), scenario_type,
                 enforce_balance is not None,
                 income_statement_only is not None,
                 int(base_level_id) if base_level_id else None,
                 notes.strip() or None))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/scenarios", err=msg)
    return flash_redirect("/scenarios", ok=f"Scenario {code.upper()} created")


@app.post("/scenarios/{scenario_id}/toggle-lock")
def toggle_lock(scenario_id: int, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                "UPDATE scenarios SET is_locked = NOT is_locked WHERE id = %s",
                (scenario_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/scenarios", err=msg)
    return flash_redirect("/scenarios", ok="Scenario updated")


# ---------------------------------------------------------------------------
# Account levels — user-named steps down the chart of accounts (see
# account_levels/scenarios.base_level_id in db/schema.sql). A scenario
# picks one of these as its base_level to post at a whole branch instead
# of every leaf under it — vertical extensibility, OneStream-style.
# ---------------------------------------------------------------------------
@app.get("/account-levels")
def account_levels_page(request: Request, ok: str = None, err: str = None):
    levels = account_levels_all()
    next_depth = max((lv["depth"] for lv in levels), default=0) + 1
    return templates.TemplateResponse(request, "account_levels.html", {
        "nav": "account_levels", "levels": levels, "next_depth": next_depth,
        "ok": ok, "err": err,
    })


@app.post("/account-levels")
def create_account_level(request: Request, name: str = Form(...), depth: str = Form(...),
                         csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        name = name.strip()
        if not name:
            raise ValueError("Name is required")
        depth_i = int(depth)
        if depth_i <= 0:
            raise ValueError("Depth must be a positive number")
        with tx() as cur:
            cur.execute(
                "INSERT INTO account_levels (name, depth) VALUES (%s, %s)",
                (name, depth_i))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/account-levels", err=msg)
    return flash_redirect("/account-levels", ok=f"Level {name!r} created")


@app.post("/account-levels/{level_id}/rename")
def rename_account_level(level_id: int, request: Request, name: str = Form(...),
                         csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        name = name.strip()
        if not name:
            raise ValueError("Name is required")
        with tx() as cur:
            cur.execute("UPDATE account_levels SET name = %s WHERE id = %s",
                       (name, level_id))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/account-levels", err=msg)
    return flash_redirect("/account-levels", ok="Level renamed")


@app.post("/account-levels/{level_id}/delete")
def delete_account_level(level_id: int, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute("DELETE FROM account_levels WHERE id = %s", (level_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/account-levels", err=msg)
    return flash_redirect("/account-levels", ok="Level deleted")


# ---------------------------------------------------------------------------
# Payees
# ---------------------------------------------------------------------------
def payees_all():
    return q("""SELECT p.*, (SELECT COUNT(*) FROM journal_entries e
                             WHERE e.payee_id = p.id) AS entry_count
                FROM payees p ORDER BY p.name""")


@app.get("/payees")
def payees_page(request: Request, ok: str = None, err: str = None):
    return templates.TemplateResponse(request, "payees.html", {
        "nav": "payees", "payees": payees_all(), "ok": ok, "err": err,
    })


@app.post("/payees")
def create_payee(request: Request, name: str = Form(...), csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        name = name.strip()
        if not name:
            raise ValueError("Payee name is required")
        with tx() as cur:
            cur.execute("INSERT INTO payees (name) VALUES (%s)", (name,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/payees", err=msg)
    return flash_redirect("/payees", ok=f"Payee {name!r} created")


@app.post("/payees/quick-create")
def quick_create_payee(request: Request, name: str = Form(...), csrf_token: str = Form(...)):
    # Called from the payee combobox on New entry via fetch() — "+ Create
    # <name>" there needs a real row back immediately so the <select> has
    # something to point payee_id at, unlike tags (synced by name at
    # submit time) or the /payees form (which redirects a whole page).
    # ON CONFLICT DO UPDATE (a no-op update) rather than DO NOTHING is the
    # standard trick to get RETURNING even when the name already exists —
    # it also quietly reactivates a deactivated payee the user is now
    # using again, which is what typing its name here signals.
    try:
        require_csrf(request, csrf_token)
        name = name.strip()
        if not name:
            raise ValueError("Payee name is required")
        with tx() as cur:
            cur.execute(
                """INSERT INTO payees (name) VALUES (%s)
                   ON CONFLICT (name) DO UPDATE
                       SET is_active = TRUE
                   RETURNING id, name""",
                (name,))
            row = cur.fetchone()
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return JSONResponse({"ok": False, "error": msg}, status_code=400)
    return JSONResponse({"ok": True, "id": row["id"], "name": row["name"]})


@app.post("/payees/{payee_id}/toggle-active")
def toggle_payee(payee_id: int, request: Request, csrf_token: str = Form(...)):
    # "Archive"/"Unarchive" on the page — is_active only ever controls
    # whether a payee still shows up in the New entry/Scheduled/Staging
    # pickers going forward; it never touches history (see /payees/{id}
    # /delete below for the route that actually does that). RETURNING the
    # new state, rather than assuming toggle flipped it the way the caller
    # expects, is what lets the flash message actually say which way it
    # went instead of a generic "updated".
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                """UPDATE payees SET is_active = NOT is_active WHERE id = %s
                   RETURNING name, is_active""",
                (payee_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError(f"Payee #{payee_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/payees", err=msg)
    verb = "unarchived" if row["is_active"] else "archived"
    return flash_redirect("/payees", ok=f"Payee {row['name']!r} {verb}")


@app.post("/payees/{payee_id}/rename")
def rename_payee(payee_id: int, request: Request, name: str = Form(...),
                 csrf_token: str = Form(...)):
    """The EDIT button's target — same "organizational, not a fact about
    the transaction" reasoning as an entry's own description (decision 16):
    a payee's name is metadata, safe to correct on something already used
    by posted entries. Nothing else about those entries changes; they just
    render whatever this payee is named *now*, the same as they always
    have (payees.name isn't copied onto journal_entries — see the FK)."""
    name = name.strip()
    try:
        require_csrf(request, csrf_token)
        if not name:
            raise ValueError("Payee name is required")
        with tx() as cur:
            cur.execute("UPDATE payees SET name = %s WHERE id = %s", (name, payee_id))
            if cur.rowcount == 0:
                raise ValueError(f"Payee #{payee_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/payees", err=msg)
    return flash_redirect("/payees", ok=f"Payee renamed to {name!r}")


@app.post("/payees/{payee_id}/delete")
def delete_payee(payee_id: int, request: Request, csrf_token: str = Form(...)):
    """A real delete, unlike Archive — every FK onto payees(id)
    (journal_entries, scheduled_entries, entry_templates) is ON DELETE SET
    NULL, so this is safe by construction: any entry that used this payee
    just goes back to having none, same as it started, rather than the
    delete being blocked or cascading into deleting entries themselves."""
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute("DELETE FROM payees WHERE id = %s RETURNING name", (payee_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError(f"Payee #{payee_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/payees", err=msg)
    return flash_redirect("/payees", ok=f"Payee {row['name']!r} deleted")


@app.post("/payees/merge")
async def merge_payees(request: Request):
    """The MERGE button's target, fired once per confirm from the popup
    entity-manage.js builds (see that file) — every selected payee's id,
    plus the final name typed into the popup (which may just be one of
    the originals, unchanged). The first selected id is kept as the
    surviving row (its id is what every FK below gets repointed to);
    which one survives is otherwise arbitrary, since the name is set
    explicitly afterward regardless of which row it started as. Deleting
    the others *before* the rename (not after) matters: if the typed name
    equals one of the about-to-be-deleted payees' own current name, naming
    the survivor that first would collide with payees.name's UNIQUE
    constraint — deleting that row first frees the name up."""
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
        ids = [int(v) for v in form.getlist("payee_id") if v]
        if len(ids) < 2:
            raise ValueError("Select at least two payees to merge")
        target_name = (form.get("target_name") or "").strip()
        if not target_name:
            raise ValueError("A name is required")
        survivor_id, other_ids = ids[0], ids[1:]
        with tx() as cur:
            cur.execute(
                "UPDATE journal_entries SET payee_id = %s WHERE payee_id = ANY(%s)",
                (survivor_id, other_ids))
            affected = cur.rowcount
            cur.execute(
                "UPDATE scheduled_entries SET payee_id = %s WHERE payee_id = ANY(%s)",
                (survivor_id, other_ids))
            cur.execute(
                "UPDATE entry_templates SET payee_id = %s WHERE payee_id = ANY(%s)",
                (survivor_id, other_ids))
            cur.execute("DELETE FROM payees WHERE id = ANY(%s)", (other_ids,))
            cur.execute("UPDATE payees SET name = %s WHERE id = %s",
                       (target_name, survivor_id))
            if cur.rowcount == 0:
                raise ValueError(f"Payee #{survivor_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/payees", err=msg)
    n = len(ids)
    return flash_redirect("/payees",
        ok=f'{n} payees merged to {target_name!r}. '
           f'{affected} {"entry" if affected == 1 else "entries"} affected')


# ---------------------------------------------------------------------------
# Tags — a management page mirroring Payees, for the entity itself rather
# than for tagging one entry (that's tags.js, on entries.html/scheduled.html/
# entry_templates.html). Same lifecycle as Payees now: Archive/Unarchive
# (is_active — hides a tag from the tag-input's suggestion list, all_tags()
# below, without touching any entry that already carries it) alongside a
# real Delete, plus Edit/Select+Merge.
# ---------------------------------------------------------------------------
def tags_all():
    return q("""SELECT t.*, (SELECT COUNT(*) FROM journal_entry_tags jet
                             WHERE jet.tag_id = t.id) AS entry_count
                FROM tags t ORDER BY t.name""")


@app.get("/tags")
def tags_page(request: Request, ok: str = None, err: str = None):
    return templates.TemplateResponse(request, "tags.html", {
        "nav": "tags", "tags": tags_all(), "ok": ok, "err": err,
    })


@app.post("/tags")
def create_tag(request: Request, name: str = Form(...), csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        names = _parse_tags(name)
        if len(names) != 1:
            raise ValueError("Enter exactly one tag name")
        with tx() as cur:
            cur.execute("INSERT INTO tags (name) VALUES (%s)", (names[0],))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/tags", err=msg)
    return flash_redirect("/tags", ok=f"Tag {names[0]!r} created")


@app.post("/tags/{tag_id}/toggle-active")
def toggle_tag(tag_id: int, request: Request, csrf_token: str = Form(...)):
    """Archive/Unarchive on the page — same shape as /payees/{id}/
    toggle-active (see its own comment): only ever changes whether this
    tag still shows up in the tag-input's suggestion list going forward,
    never touches an entry that already carries it."""
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                """UPDATE tags SET is_active = NOT is_active WHERE id = %s
                   RETURNING name, is_active""",
                (tag_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError(f"Tag #{tag_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/tags", err=msg)
    verb = "unarchived" if row["is_active"] else "archived"
    return flash_redirect("/tags", ok=f"Tag {row['name']!r} {verb}")


@app.post("/tags/{tag_id}/rename")
def rename_tag(tag_id: int, request: Request, name: str = Form(...),
              csrf_token: str = Form(...)):
    """Same "organizational, not a fact" reasoning as /payees/{id}/rename —
    a tag's name is metadata, safe to correct after the fact; every entry
    that already carries it just renders whatever it's named now (a tag
    name isn't copied anywhere, only referenced by journal_entry_tags'
    tag_id). Goes through _parse_tags for the same lowercase/pattern
    validation typing one into the chip input gets, not a raw UPDATE."""
    try:
        require_csrf(request, csrf_token)
        names = _parse_tags(name)
        if len(names) != 1:
            raise ValueError("Enter exactly one tag name")
        with tx() as cur:
            cur.execute("UPDATE tags SET name = %s WHERE id = %s", (names[0], tag_id))
            if cur.rowcount == 0:
                raise ValueError(f"Tag #{tag_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/tags", err=msg)
    return flash_redirect("/tags", ok=f"Tag renamed to {names[0]!r}")


@app.post("/tags/{tag_id}/delete")
def delete_tag(tag_id: int, request: Request, csrf_token: str = Form(...)):
    """journal_entry_tags/scheduled_entry_tags/entry_template_tags all
    reference tags(id) ON DELETE CASCADE (unlike payees' ON DELETE SET
    NULL) — deleting a tag just drops it from whatever it was on, rather
    than leaving a dangling reference to null out. Nothing else about
    those entries changes; they simply stop carrying this tag."""
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute("DELETE FROM tags WHERE id = %s RETURNING name", (tag_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError(f"Tag #{tag_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/tags", err=msg)
    return flash_redirect("/tags", ok=f"Tag {row['name']!r} deleted")


@app.post("/tags/merge")
async def merge_tags(request: Request):
    """MERGE's target for Tags — same shape as /payees/merge (see its own
    comment for the survivor/delete-then-rename reasoning), except a tag's
    associations are many-to-many, not a single FK column: journal_entry_
    tags/scheduled_entry_tags/entry_template_tags each get an "insert the
    survivor's own association wherever a merged-away tag had one, ON
    CONFLICT DO NOTHING" pass before the old tag rows are deleted, since a
    plain UPDATE ... SET tag_id could collide with a (entry_id, tag_id)
    pair that already exists (something tagged with *both* the survivor
    and a tag being folded into it) and violate the junction table's own
    primary key. "Entries affected" counts distinct journal_entries only,
    same as Payees — scheduled entries/templates carrying a merged tag
    aren't reflected in that count either, for the same reason."""
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
        ids = [int(v) for v in form.getlist("tag_id") if v]
        if len(ids) < 2:
            raise ValueError("Select at least two tags to merge")
        target_names = _parse_tags(form.get("target_name") or "")
        if len(target_names) != 1:
            raise ValueError("Enter exactly one tag name")
        target_name = target_names[0]
        survivor_id, other_ids = ids[0], ids[1:]
        with tx() as cur:
            cur.execute(
                "SELECT COUNT(DISTINCT entry_id) AS n FROM journal_entry_tags WHERE tag_id = ANY(%s)",
                (other_ids,))
            affected = cur.fetchone()["n"]
            for table, id_col in (("journal_entry_tags", "entry_id"),
                                  ("scheduled_entry_tags", "scheduled_entry_id"),
                                  ("entry_template_tags", "template_id")):
                cur.execute(
                    f"""INSERT INTO {table} ({id_col}, tag_id)
                        SELECT {id_col}, %s FROM {table} WHERE tag_id = ANY(%s)
                        ON CONFLICT DO NOTHING""",
                    (survivor_id, other_ids))
            cur.execute("DELETE FROM tags WHERE id = ANY(%s)", (other_ids,))
            cur.execute("UPDATE tags SET name = %s WHERE id = %s", (target_name, survivor_id))
            if cur.rowcount == 0:
                raise ValueError(f"Tag #{survivor_id} not found")
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/tags", err=msg)
    n = len(ids)
    return flash_redirect("/tags",
        ok=f'{n} tags merged to {target_name!r}. '
           f'{affected} {"entry" if affected == 1 else "entries"} affected')


# ---------------------------------------------------------------------------
# Scheduled entries — a template + recurrence rule. Due occurrences are
# auto-posted to the Staging scenario (materialize_due_schedules(), called
# from the auth middleware); a human approves each one from here before it
# becomes a real entry in its target scenario ("posting" = copy + mark the
# staging row promoted, never editing it in place — same append-only spirit
# as the rest of the ledger).
# ---------------------------------------------------------------------------
SCHEDULE_UNITS = ["day", "week", "month"]


def _advance_date(d: date, unit: str, count: int) -> date:
    if unit == "day":
        return d + timedelta(days=count)
    if unit == "week":
        return d + timedelta(weeks=count)
    if unit == "month":
        total = d.month - 1 + count
        year = d.year + total // 12
        month = total % 12 + 1
        day = min(d.day, calendar.monthrange(year, month)[1])  # clamp Jan 31 + 1mo -> Feb 28/29
        return date(year, month, day)
    raise ValueError(f"Unknown interval unit: {unit}")


def scheduled_all():
    return q("""
        SELECT se.*, s.code AS scenario_code, s.name AS scenario_name,
               p.name AS payee_name,
               (SELECT COUNT(*) FROM scheduled_entry_lines
                 WHERE scheduled_entry_id = se.id) AS line_count,
               (SELECT COALESCE(SUM(debit), 0) FROM scheduled_entry_lines
                 WHERE scheduled_entry_id = se.id) AS total_amount
          FROM scheduled_entries se
          JOIN scenarios s ON s.id = se.target_scenario_id
          LEFT JOIN payees p ON p.id = se.payee_id
         ORDER BY se.next_date, se.id""")


def _staging_filter(date_from: str = "", date_to: str = "", qtext: str = "", tags: str = "",
                    account: str = "", payee: str = "", amount_op: str = "",
                    amount_value: str = "", amount_value2: str = "",
                    target_scenario: str = "") -> tuple[list[str], list, list[str]]:
    """Same filter fields as the Journal's own bar (see
    _shared_journal_filters), reusing that exact logic — the one thing
    genuinely different here is 'Scenario': every row in Staging already
    shares one real scenario (STAGING itself), so filtering on *that*
    would be meaningless. What actually varies row to row is where each
    entry is headed once approved, so that's what this filters on
    instead — the ts/ib_ts aliases _entries_filter has no equivalent for,
    which is why this isn't just a call into that one."""
    try:
        tag_list = _parse_tags(tags) if tags else []
    except ValueError:
        tag_list = []
    where, params = ["e.promoted_entry_id IS NULL"], []
    if target_scenario:
        where.append("COALESCE(ts.code, ib_ts.code) = %s")
        params.append(target_scenario)
    _shared_journal_filters(where, params, date_from, date_to, qtext, tag_list,
                            account, payee, amount_op, amount_value, amount_value2)
    return where, params, tag_list


def pending_staging_entries(date_from: str = "", date_to: str = "", qtext: str = "",
                            tags: str = "", account: str = "", payee: str = "",
                            amount_op: str = "", amount_value: str = "", amount_value2: str = "",
                            target_scenario: str = ""):
    """Everything sitting in the Staging scenario, not yet approved — the
    Staging page's whole reason to exist (called with no arguments, i.e.
    unfiltered, by the Dashboard's banner count and Scheduled's own
    pending count). Not limited to schedule-sourced rows: this is
    *everything* Staging is holding regardless of which of the two
    producers put it there (a materialized schedule or a CSV import), so
    both joins are LEFT — for the "where's this headed, and where did it
    come from" display detail (and, now, the Scenario filter), not a
    filter on their own — and each entry has at most one of the two set,
    never both."""
    where, params, _ = _staging_filter(date_from, date_to, qtext, tags, account, payee,
                                       amount_op, amount_value, amount_value2, target_scenario)
    entries = q(f"""
        SELECT e.id, e.entry_date, e.description, e.reference,
               p.name AS payee_name,
               COALESCE(ts.code, ib_ts.code) AS target_scenario_code,
               COALESCE(ts.name, ib_ts.name) AS target_scenario_name,
               ib.filename AS import_filename,
               (SELECT COALESCE(SUM(l.debit), 0) FROM journal_lines l
                 WHERE l.entry_id = e.id) AS total_debits
          FROM journal_entries e
          JOIN scenarios stg ON stg.id = e.scenario_id AND stg.is_staging
          LEFT JOIN scheduled_entries se ON se.id = e.scheduled_entry_id
          LEFT JOIN scenarios ts ON ts.id = se.target_scenario_id
          LEFT JOIN import_batches ib ON ib.id = e.import_batch_id
          LEFT JOIN scenarios ib_ts ON ib_ts.id = ib.target_scenario_id
          LEFT JOIN payees p ON p.id = e.payee_id
         WHERE {' AND '.join(where)}
         ORDER BY e.entry_date, e.seq""", params)
    ids = [e["id"] for e in entries]
    lines_by_entry = {}
    if ids:
        for ln in q("""SELECT l.entry_id, l.debit, l.credit, l.memo,
                              a.code AS account_code, a.name AS account_name
                         FROM journal_lines l
                         JOIN accounts a ON a.id = l.account_id
                        WHERE l.entry_id = ANY(%s)
                        ORDER BY l.entry_id, l.line_no""", (ids,)):
            lines_by_entry.setdefault(ln["entry_id"], []).append(ln)
    return entries, lines_by_entry


def materialize_due_schedules() -> None:
    due = q("""SELECT * FROM scheduled_entries
               WHERE is_active AND next_date <= CURRENT_DATE
               ORDER BY id""")
    if not due:
        return
    staging = q1("SELECT id FROM scenarios WHERE is_staging")
    if not staging:
        return  # schema migrated but the seed row hasn't landed yet
    for sched in due:
        lines = q("""SELECT line_no, account_id, amount, memo
                       FROM scheduled_entry_lines
                      WHERE scheduled_entry_id = %s ORDER BY line_no""", (sched["id"],))
        if not lines:
            continue  # nothing to post; leave next_date alone rather than skip silently
        tag_names = [r["name"] for r in q(
            """SELECT tg.name FROM scheduled_entry_tags st
                JOIN tags tg ON tg.id = st.tag_id
               WHERE st.scheduled_entry_id = %s""", (sched["id"],))]
        with tx() as cur:
            cur.execute(
                """INSERT INTO journal_entries
                       (scenario_id, entry_date, description, reference,
                        payee_id, scheduled_entry_id)
                   VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                (staging["id"], sched["next_date"], sched["description"],
                 sched["reference"], sched["payee_id"], sched["id"]))
            entry_id = cur.fetchone()["id"]
            for ln in lines:
                cur.execute(
                    """INSERT INTO journal_lines
                           (entry_id, line_no, account_id, amount, memo)
                       VALUES (%s, %s, %s, %s, %s)""",
                    (entry_id, ln["line_no"], ln["account_id"], ln["amount"], ln["memo"]))
            if tag_names:
                _sync_entry_tags(cur, entry_id, tag_names)
            cur.execute(
                "UPDATE scheduled_entries SET next_date = %s WHERE id = %s",
                (_advance_date(sched["next_date"], sched["interval_unit"],
                               sched["interval_count"]), sched["id"]))


@app.get("/scheduled")
def scheduled_page(request: Request, ok: str = None, err: str = None):
    # income-statement-only scenarios can never receive a promoted journal
    # entry (fn_income_statement_only_guard) — same reason entries.html's
    # "New entry" panel excludes them from its own scenario picker. Staging
    # itself is excluded too: approving a schedule's occurrence *into*
    # Staging would be circular — Staging is the layover, not a destination.
    scen = [s for s in scenarios_all()
           if not s["is_locked"] and not s["income_statement_only"] and not s["is_staging"]]
    by_scenario = postable_accounts_by_scenario()
    postable = by_scenario.get(scen[0]["id"], []) if scen else []
    active_payees = q("SELECT id, name FROM payees WHERE is_active ORDER BY name")
    pending, _ = pending_staging_entries()
    return templates.TemplateResponse(request, "scheduled.html", {
        "nav": "scheduled", "schedules": scheduled_all(),
        "accounts": postable, "accounts_by_scenario": by_scenario,
        "scenarios": scen, "payees": active_payees,
        "all_tags": all_tags(), "today": date.today().isoformat(),
        "units": SCHEDULE_UNITS,
        "pending_count": len(pending),
        "ok": ok, "err": err,
    })


@app.post("/scheduled")
async def create_schedule(request: Request):
    form = await request.form()
    wants_json = "application/json" in request.headers.get("accept", "")
    try:
        require_csrf(request, form.get("csrf_token"))
        lines = _parse_lines(form)
        total = round(sum(ln["amount"] for ln in lines), 2)
        if total != 0:
            raise ValueError("Schedule lines must balance (debits = credits)")
        description = (form.get("description") or "").strip()
        if not description:
            raise ValueError("Description is required")
        reference = (form.get("reference") or "").strip() or None
        payee_id = int(form.get("payee_id")) if form.get("payee_id") else None
        tag_names = _parse_tags(form.get("tags", ""))
        target_scenario_id = int(form.get("scenario_id"))
        interval_unit = form.get("interval_unit") or ""
        if interval_unit not in SCHEDULE_UNITS:
            raise ValueError("Choose a valid repeat unit")
        interval_count = int(form.get("interval_count") or 1)
        if interval_count <= 0:
            raise ValueError("Repeat count must be positive")
        next_date = form.get("next_date") or date.today().isoformat()

        codes = {ln["code"] for ln in lines}
        found = {r["code"] for r in q(
            "SELECT code FROM accounts WHERE code = ANY(%s)", (list(codes),))}
        missing = codes - found
        if missing:
            raise ValueError(f"Unknown account code: {', '.join(sorted(missing))}")

        with tx() as cur:
            cur.execute(
                """INSERT INTO scheduled_entries
                       (description, reference, payee_id, target_scenario_id,
                        interval_unit, interval_count, next_date)
                   VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id""",
                (description, reference, payee_id, target_scenario_id,
                 interval_unit, interval_count, next_date))
            sched_id = cur.fetchone()["id"]
            for n, ln in enumerate(lines, start=1):
                cur.execute(
                    """INSERT INTO scheduled_entry_lines
                           (scheduled_entry_id, line_no, account_id, amount, memo)
                       VALUES (%s, %s, (SELECT id FROM accounts WHERE code = %s), %s, %s)""",
                    (sched_id, n, ln["code"], ln["amount"], ln["memo"]))
            if tag_names:
                _sync_tags(cur, "scheduled_entry_tags", "scheduled_entry_id",
                          sched_id, tag_names)
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        if wants_json:
            return JSONResponse({"ok": False, "error": msg}, status_code=400)
        return flash_redirect("/scheduled", err=msg)
    ok_msg = f"Schedule {description!r} created — next on {next_date}"
    if wants_json:
        return JSONResponse({"ok": True, "redirect": flash_url("/scheduled", ok=ok_msg)})
    return flash_redirect("/scheduled", ok=ok_msg)


@app.post("/scheduled/{scheduled_id}/toggle-active")
def toggle_schedule(scheduled_id: int, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute(
                "UPDATE scheduled_entries SET is_active = NOT is_active WHERE id = %s",
                (scheduled_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/scheduled", err=msg)
    return flash_redirect("/scheduled", ok="Schedule updated")


@app.get("/staging")
def staging_page(request: Request, date_from: str = "", date_to: str = "", qtext: str = "",
                 tags: str = "", account: str = "", payee: str = "", amount_op: str = "",
                 amount_value: str = "", amount_value2: str = "", target_scenario: str = "",
                 ok: str = None, err: str = None):
    pending, pending_lines = pending_staging_entries(
        date_from, date_to, qtext, tags, account, payee,
        amount_op, amount_value, amount_value2, target_scenario)
    # Same idea as the Journal's own has_filters — no point showing
    # "Clear filters" over an unfiltered list. No "back" to preserve here
    # (nothing links into Staging with one), so the clear link is just
    # a plain "/staging".
    has_filters = bool(date_from or date_to or qtext or tags or account or payee
                       or amount_op or target_scenario)
    return templates.TemplateResponse(request, "staging.html", {
        "nav": "staging", "pending": pending, "pending_lines": pending_lines,
        "date_from": date_from, "date_to": date_to, "qtext": qtext,
        "tags": tags, "all_tags": all_tags(),
        "account": account, "payee": payee,
        "amount_op": amount_op, "amount_value": amount_value, "amount_value2": amount_value2,
        "amount_ops": AMOUNT_OPS,
        "target_scenario": target_scenario,
        "target_scenarios": [s for s in scenarios_all() if not s["is_staging"]],
        "filter_accounts": postable_accounts_for_pickers(),
        "filter_payee_names": [r["name"] for r in q("SELECT name FROM payees ORDER BY name")],
        # For the inline edit panel's own Payee field (see
        # staging-inline-edit.js) — unlike accounts, payees aren't
        # scenario-scoped, so the one global list works for editing any
        # pending entry and doesn't need to be fetched per-entry.
        "payees": q("SELECT id, name FROM payees WHERE is_active ORDER BY name"),
        "has_filters": has_filters,
        "ok": ok, "err": err,
    })


@app.post("/staging/approve")
async def approve_staging_entries(request: Request):
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
    except ValueError as e:
        return flash_redirect("/staging", err=str(e))

    entry_ids = [v for v in form.getlist("entry_id") if v]
    if not entry_ids:
        return flash_redirect("/staging", err="Select at least one entry to approve")

    posted, errors = [], []
    for eid in entry_ids:
        try:
            # LEFT JOINs, not INNER: an entry carries its target via
            # exactly one of scheduled_entries or import_batches (or,
            # hypothetically, neither) — s.is_staging is what actually
            # proves this row belongs here, not either join.
            staged = q1("""SELECT e.*, s.is_staging,
                                  COALESCE(se.target_scenario_id, ib.target_scenario_id)
                                      AS target_scenario_id
                             FROM journal_entries e
                             JOIN scenarios s ON s.id = e.scenario_id
                             LEFT JOIN scheduled_entries se ON se.id = e.scheduled_entry_id
                             LEFT JOIN import_batches ib ON ib.id = e.import_batch_id
                            WHERE e.id = %s""", (eid,))
            if not staged or not staged["is_staging"]:
                raise ValueError(f"#{eid}: not a pending staging entry")
            if staged["promoted_entry_id"] is not None:
                raise ValueError(f"#{eid}: already approved")
            target_scenario_id = staged["target_scenario_id"]
            if target_scenario_id is None:
                # Neither producer said where this belongs — ACTUAL is the
                # only sensible default destination.
                actual = q1("SELECT id FROM scenarios WHERE code = 'ACTUAL'")
                target_scenario_id = actual["id"] if actual else None
            if target_scenario_id is None:
                raise ValueError(f"#{eid}: no target scenario to approve into")
            with tx() as cur:
                cur.execute(
                    """INSERT INTO journal_entries
                           (scenario_id, entry_date, description, reference,
                            payee_id, created_by_user_id)
                       VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                    (target_scenario_id, staged["entry_date"],
                     staged["description"], staged["reference"], staged["payee_id"],
                     auth.current_user(request)["user_id"]))
                new_id = cur.fetchone()["id"]
                cur.execute(
                    """INSERT INTO journal_lines
                           (entry_id, line_no, account_id, amount, memo)
                       SELECT %s, line_no, account_id, amount, memo
                         FROM journal_lines WHERE entry_id = %s""",
                    (new_id, eid))
                cur.execute(
                    """INSERT INTO journal_entry_tags (entry_id, tag_id)
                       SELECT %s, tag_id FROM journal_entry_tags WHERE entry_id = %s""",
                    (new_id, eid))
                cur.execute(
                    "UPDATE journal_entries SET promoted_entry_id = %s WHERE id = %s",
                    (new_id, eid))
            posted.append(eid)
        except (ValueError, psycopg.Error) as e:
            errors.append(_pg_msg(e) if isinstance(e, psycopg.Error) else str(e))

    ok_msg = f"Approved {len(posted)} entr{'y' if len(posted) == 1 else 'ies'}" if posted else None
    err_msg = "; ".join(errors) or None
    return flash_redirect("/staging", ok=ok_msg, err=err_msg)


def _pending_staging_entry(entry_id: str):
    """The one entry /staging/{id}/edit and /staging/{id}/reject both act
    on — raises ValueError (same as any other bad input in these routes)
    unless it's actually eligible: a real Staging entry, not yet approved.
    Mirrors the same is_staging/promoted_entry_id check approve_staging_
    entries() does per row, and (see db/schema.sql's fn_lines_immutable/
    fn_entries_guard) the same condition the database itself relaxes
    immutability for — this is a friendlier error in front of that, not
    a substitute for it."""
    staged = q1("""SELECT e.*, s.is_staging,
                          COALESCE(se.target_scenario_id, ib.target_scenario_id)
                              AS target_scenario_id
                     FROM journal_entries e
                     JOIN scenarios s ON s.id = e.scenario_id
                     LEFT JOIN scheduled_entries se ON se.id = e.scheduled_entry_id
                     LEFT JOIN import_batches ib ON ib.id = e.import_batch_id
                    WHERE e.id = %s""", (entry_id,))
    if not staged or not staged["is_staging"]:
        raise ValueError(f"#{entry_id}: not a pending staging entry")
    if staged["promoted_entry_id"] is not None:
        raise ValueError(f"#{entry_id}: already approved")
    target_scenario_id = staged["target_scenario_id"]
    if target_scenario_id is None:
        actual = q1("SELECT id FROM scenarios WHERE code = 'ACTUAL'")
        target_scenario_id = actual["id"] if actual else None
    if target_scenario_id is None:
        raise ValueError(f"#{entry_id}: no target scenario")
    staged["target_scenario_id"] = target_scenario_id
    return staged


@app.get("/staging/{entry_id}/edit")
def staging_edit_data(entry_id: str):
    """A JSON data endpoint now, not a page — Staging's own page
    (staging.html) edits an entry inline, reusing the exact "+ New
    entry" grid component the Journal uses (see staging-inline-edit.js),
    instead of navigating here first. This is what that panel fetches
    to fill itself in: everything specific to *this* entry. Everything
    NOT entry-specific (payees, all_tags) is already on the Staging page
    itself, rendered once, not refetched per entry opened."""
    try:
        staged = _pending_staging_entry(entry_id)
    except ValueError as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    target_scenario = q1("SELECT id, code, name, enforce_balance FROM scenarios WHERE id = %s",
                         (staged["target_scenario_id"],))
    # debit/credit come back as Decimal — not JSON-serializable as-is (see
    # templates_full()'s identical str()-or-None conversion for the same
    # reason) — turned into plain strings/None here, before
    # staging-inline-edit.js reads them to build the grid.
    lines = [{
        "code": ln["code"],
        "debit": str(ln["debit"]) if ln["debit"] else None,
        "credit": str(ln["credit"]) if ln["credit"] else None,
        "memo": ln["memo"],
    } for ln in q("""SELECT l.line_no, l.debit, l.credit, l.memo, a.code
                       FROM journal_lines l JOIN accounts a ON a.id = l.account_id
                      WHERE l.entry_id = %s ORDER BY l.line_no""", (entry_id,))]
    tag_names = [r["name"] for r in q(
        """SELECT tg.name FROM journal_entry_tags jet
            JOIN tags tg ON tg.id = jet.tag_id
           WHERE jet.entry_id = %s ORDER BY tg.name""", (entry_id,))]
    by_scenario = postable_accounts_by_scenario()
    accounts = by_scenario.get(staged["target_scenario_id"], [])

    return JSONResponse({
        "ok": True,
        "entry": {
            "id": staged["id"],
            "entry_date": staged["entry_date"].isoformat(),
            "description": staged["description"],
            "reference": staged["reference"] or "",
            "payee_id": staged["payee_id"],
        },
        "lines": lines, "tags": tag_names,
        "target_scenario": target_scenario, "accounts": accounts,
    })


@app.post("/staging/{entry_id}/edit")
async def staging_edit_save(entry_id: str, request: Request):
    form = await request.form()
    # Same reason as create_entry (app.js's grid submits every page it's
    # on the same way, via fetch requesting JSON) — a rejected edit needs
    # to report back without reloading the page and losing whatever the
    # person had just changed.
    wants_json = "application/json" in request.headers.get("accept", "")
    try:
        require_csrf(request, form.get("csrf_token"))
        _pending_staging_entry(entry_id)
        lines = _parse_lines(form)
        entry_date = form.get("entry_date") or date.today().isoformat()
        description = (form.get("description") or "").strip()
        reference = (form.get("reference") or "").strip() or None
        payee_id = int(form.get("payee_id")) if form.get("payee_id") else None
        tag_names = _parse_tags(form.get("tags", ""))
        if not description:
            raise ValueError("Description is required")

        codes = {ln["code"] for ln in lines}
        found = {r["code"] for r in q(
            "SELECT code FROM accounts WHERE code = ANY(%s)", (list(codes),))}
        missing = codes - found
        if missing:
            raise ValueError(f"Unknown account code: {', '.join(sorted(missing))}")

        with tx() as cur:
            cur.execute(
                """UPDATE journal_entries
                       SET entry_date = %s, description = %s, reference = %s, payee_id = %s
                     WHERE id = %s""",
                (entry_date, description, reference, payee_id, entry_id))
            # Delete-then-reinsert rather than trying to patch individual
            # lines in place — journal_lines stays UPDATE-blocked even for
            # a pending Staging entry (see db/schema.sql), only DELETE is
            # relaxed, so this is the only shape an edit can take. Both
            # happen in the same transaction as the header UPDATE above,
            # so the deferred balance/has-lines checks only ever see the
            # final, complete set at commit — never the momentarily-empty
            # state in between.
            cur.execute("DELETE FROM journal_lines WHERE entry_id = %s", (entry_id,))
            for n, ln in enumerate(lines, start=1):
                cur.execute(
                    """INSERT INTO journal_lines
                           (entry_id, line_no, account_id, amount, memo)
                       VALUES (%s, %s,
                               (SELECT id FROM accounts WHERE code = %s),
                               %s, %s)""",
                    (entry_id, n, ln["code"], ln["amount"], ln["memo"]))
            _sync_entry_tags(cur, entry_id, tag_names)
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        if wants_json:
            return JSONResponse({"ok": False, "error": msg}, status_code=400)
        return flash_redirect(f"/staging/{entry_id}/edit", err=msg)
    ok_msg = f"#{entry_id} updated"
    if wants_json:
        return JSONResponse({"ok": True, "redirect": flash_url("/staging", ok=ok_msg)})
    return flash_redirect("/staging", ok=ok_msg)


@app.post("/staging/{entry_id}/reject")
def staging_reject(entry_id: str, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        _pending_staging_entry(entry_id)
        with tx() as cur:
            # journal_lines does have ON DELETE CASCADE on entry_id, so
            # deleting just the entry would clean these up too — spelled
            # out as its own statement anyway so a failure here (the
            # trigger refusing it, on some future entry this helper's
            # check didn't catch) reports "couldn't delete the lines"
            # rather than a less obvious cascade-related error.
            cur.execute("DELETE FROM journal_lines WHERE entry_id = %s", (entry_id,))
            cur.execute("DELETE FROM journal_entries WHERE id = %s", (entry_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/staging", err=msg)
    return flash_redirect("/staging", ok=f"#{entry_id} rejected and deleted")


@app.post("/staging/reject")
async def reject_staging_entries(request: Request):
    """Bulk sibling of /staging/{id}/reject, for the top-of-page Reject
    button next to Approve entries — same per-id validation
    (_pending_staging_entry) and the same permanent delete, just looped
    over a checked set the way Approve already is. Shares Approve's own
    checkboxes (name="entry_id") since both buttons submit the same
    outer <form> — see staging.html."""
    form = await request.form()
    try:
        require_csrf(request, form.get("csrf_token"))
    except ValueError as e:
        return flash_redirect("/staging", err=str(e))

    entry_ids = [v for v in form.getlist("entry_id") if v]
    if not entry_ids:
        return flash_redirect("/staging", err="Select at least one entry to reject")

    rejected, errors = [], []
    for eid in entry_ids:
        try:
            _pending_staging_entry(eid)
            with tx() as cur:
                cur.execute("DELETE FROM journal_lines WHERE entry_id = %s", (eid,))
                cur.execute("DELETE FROM journal_entries WHERE id = %s", (eid,))
            rejected.append(eid)
        except (ValueError, psycopg.Error) as e:
            errors.append(_pg_msg(e) if isinstance(e, psycopg.Error) else str(e))

    ok_msg = f"Rejected {len(rejected)} entr{'y' if len(rejected) == 1 else 'ies'}" if rejected else None
    err_msg = "; ".join(errors) or None
    return flash_redirect("/staging", ok=ok_msg, err=err_msg)


# ---------------------------------------------------------------------------
# CSV import — the other producer Staging accepts entries from, alongside
# Scheduled entries. Deliberately round-trips /entries/export.csv's own
# column layout ("Entry #" groups rows back into one entry per journal
# entry, everything else lines up 1:1) so export → edit in a spreadsheet →
# re-import is a real workflow, not just a one-way dump.
# ---------------------------------------------------------------------------
IMPORT_REQUIRED_COLUMNS = ["Entry #", "Date", "Description", "Account code"]
IMPORT_MAX_ERRORS_SHOWN = 20


def _parse_csv_import(content: str) -> tuple[list[dict], list[str]]:
    """(groups, errors) — every group in `groups` already passed every
    check (a real account code, exactly one of debit/credit per line, and
    the whole entry nets to zero) and is ready to insert; `errors`
    describes every row/group that didn't, by original CSV row number,
    and never touches the database. The "Entry #"/"Scenario"/"Account
    name" columns an export produces are read only to group rows and are
    otherwise ignored — the id isn't reused, and the scenario a batch
    lands in comes from the import form, never trusted from inside the
    file itself."""
    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        return [], ["The file is empty"]
    missing = [c for c in IMPORT_REQUIRED_COLUMNS if c not in reader.fieldnames]
    if missing:
        return [], [f"Missing required column(s): {', '.join(missing)}"]

    raw_groups: dict[str, list[tuple[int, dict]]] = {}
    order: list[str] = []
    errors = []
    for i, row in enumerate(reader, start=2):  # header is row 1
        key = (row.get("Entry #") or "").strip()
        if not key:
            errors.append(f"Row {i}: missing Entry #")
            continue
        if key not in raw_groups:
            raw_groups[key] = []
            order.append(key)
        raw_groups[key].append((i, row))

    codes = {(row.get("Account code") or "").strip()
             for rows in raw_groups.values() for _, row in rows}
    codes.discard("")
    found = ({r["code"] for r in q(
        "SELECT code FROM accounts WHERE code = ANY(%s)", (list(codes),))}
             if codes else set())

    groups = []
    for key in order:
        rows = raw_groups[key]
        first_row_no, first = rows[0]
        lines, ok = [], True
        for row_no, row in rows:
            code = (row.get("Account code") or "").strip()
            if not code:
                errors.append(f"Row {row_no} (entry {key}): missing Account code")
                ok = False
                continue
            if code not in found:
                errors.append(f"Row {row_no} (entry {key}): unknown account code {code!r}")
                ok = False
                continue
            d, c = (row.get("Debit") or "").strip(), (row.get("Credit") or "").strip()
            try:
                dv, cv = (float(d) if d else 0.0), (float(c) if c else 0.0)
            except ValueError:
                errors.append(f"Row {row_no} (entry {key}): Debit/Credit must be numeric")
                ok = False
                continue
            if dv < 0 or cv < 0 or (dv > 0) == (cv > 0):
                errors.append(f"Row {row_no} (entry {key}): enter exactly one positive Debit or Credit")
                ok = False
                continue
            lines.append({"code": code, "amount": round(dv - cv, 2),
                          "memo": (row.get("Memo") or "").strip() or None})
        if not ok:
            continue
        total = round(sum(ln["amount"] for ln in lines), 2)
        if total != 0:
            errors.append(f"Entry {key} (row {first_row_no}): doesn't balance (off by {total:+.2f})")
            continue
        entry_date = (first.get("Date") or "").strip()
        try:
            date.fromisoformat(entry_date)
        except ValueError:
            errors.append(f"Entry {key} (row {first_row_no}): invalid Date {entry_date!r} — expected YYYY-MM-DD")
            continue
        description = (first.get("Description") or "").strip()
        if not description:
            errors.append(f"Entry {key} (row {first_row_no}): missing Description")
            continue
        groups.append({
            "entry_date": entry_date, "description": description,
            "reference": (first.get("Reference") or "").strip() or None,
            "payee_name": (first.get("Payee") or "").strip() or None,
            "lines": lines,
        })
    return groups, errors


@app.get("/import")
def import_page(request: Request, ok: str = None, err: str = None):
    # Same exclusions as Scheduled's target-scenario picker: an import has
    # to land somewhere it can eventually become real postings.
    scen = [s for s in scenarios_all()
           if not s["is_locked"] and not s["income_statement_only"] and not s["is_staging"]]
    recent = q("""SELECT ib.id, ib.filename, ib.row_count, ib.created_at,
                         s.code AS target_scenario_code, u.username AS imported_by
                    FROM import_batches ib
                    JOIN scenarios s ON s.id = ib.target_scenario_id
                    LEFT JOIN users u ON u.id = ib.imported_by_user_id
                   ORDER BY ib.created_at DESC LIMIT 10""")
    return templates.TemplateResponse(request, "import.html", {
        "nav": "import", "scenarios": scen, "recent": recent, "ok": ok, "err": err,
    })


@app.post("/import")
async def import_csv(request: Request, target_scenario_id: str = Form(...),
                     csrf_token: str = Form(...), file: UploadFile = File(...)):
    try:
        require_csrf(request, csrf_token)
        raw = await file.read()
        try:
            content = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise ValueError("Could not read the file as UTF-8 text")

        groups, errors = _parse_csv_import(content)
        if not groups:
            raise ValueError("; ".join(errors[:IMPORT_MAX_ERRORS_SHOWN]) or "No valid entries found in the file")

        staging = q1("SELECT id FROM scenarios WHERE is_staging")
        if not staging:
            raise ValueError("No Staging scenario configured")

        with tx() as cur:
            cur.execute(
                """INSERT INTO import_batches
                       (filename, target_scenario_id, imported_by_user_id, row_count)
                   VALUES (%s, %s, %s, %s) RETURNING id""",
                (file.filename or "import.csv", int(target_scenario_id),
                 auth.current_user(request)["user_id"], len(groups)))
            batch_id = cur.fetchone()["id"]
            for g in groups:
                payee_id = None
                if g["payee_name"]:
                    cur.execute(
                        """INSERT INTO payees (name) VALUES (%s)
                           ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name
                           RETURNING id""",
                        (g["payee_name"],))
                    payee_id = cur.fetchone()["id"]
                cur.execute(
                    """INSERT INTO journal_entries
                           (scenario_id, entry_date, description, reference,
                            payee_id, import_batch_id)
                       VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                    (staging["id"], g["entry_date"], g["description"],
                     g["reference"], payee_id, batch_id))
                entry_id = cur.fetchone()["id"]
                for n, ln in enumerate(g["lines"], start=1):
                    cur.execute(
                        """INSERT INTO journal_lines
                               (entry_id, line_no, account_id, amount, memo)
                           VALUES (%s, %s, (SELECT id FROM accounts WHERE code = %s), %s, %s)""",
                        (entry_id, n, ln["code"], ln["amount"], ln["memo"]))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/import", err=msg)

    ok_msg = f"Staged {len(groups)} entr{'y' if len(groups) == 1 else 'ies'} for review in Staging"
    err_msg = None
    if errors:
        shown = errors[:IMPORT_MAX_ERRORS_SHOWN]
        if len(errors) > len(shown):
            shown.append(f"...and {len(errors) - len(shown)} more")
        err_msg = f"{len(errors)} row(s) skipped: " + "; ".join(shown)
    return flash_redirect("/import", ok=ok_msg, err=err_msg)


# ---------------------------------------------------------------------------
# Entry templates — reusable scaffolding for New entry's "Load template"
# picker. Not postings themselves and not tracked once loaded: loading one
# just fills the form (entirely client-side — see entry_templates.js and
# the #templates-data blob below), the same as typing it by hand.
# ---------------------------------------------------------------------------
def templates_full():
    tpls = q("""SELECT t.id, t.name, t.description, t.reference, t.payee_id,
                       p.name AS payee_name
                  FROM entry_templates t
                  LEFT JOIN payees p ON p.id = t.payee_id
                 ORDER BY t.name""")
    ids = [t["id"] for t in tpls]
    lines_by_t, tags_by_t = {}, {}
    if ids:
        for ln in q("""SELECT l.template_id, a.code, l.debit, l.credit, l.memo
                         FROM entry_template_lines l
                         JOIN accounts a ON a.id = l.account_id
                        WHERE l.template_id = ANY(%s)
                        ORDER BY l.template_id, l.line_no""", (ids,)):
            lines_by_t.setdefault(ln["template_id"], []).append({
                "code": ln["code"],
                "debit": str(ln["debit"]) if ln["debit"] else None,
                "credit": str(ln["credit"]) if ln["credit"] else None,
                "memo": ln["memo"],
            })
        for tg in q("""SELECT ett.template_id, tg.name FROM entry_template_tags ett
                        JOIN tags tg ON tg.id = ett.tag_id
                       WHERE ett.template_id = ANY(%s) ORDER BY tg.name""", (ids,)):
            tags_by_t.setdefault(tg["template_id"], []).append(tg["name"])
    for t in tpls:
        t["lines"] = lines_by_t.get(t["id"], [])
        t["tags"] = tags_by_t.get(t["id"], [])
    return tpls


@app.get("/templates")
def templates_page(request: Request, ok: str = None, err: str = None):
    postable = postable_accounts_for_pickers()
    active_payees = q("SELECT id, name FROM payees WHERE is_active ORDER BY name")
    return templates.TemplateResponse(request, "entry_templates.html", {
        "nav": "templates", "tpls": templates_full(),
        "accounts": postable, "payees": active_payees, "all_tags": all_tags(),
        "ok": ok, "err": err,
    })


@app.post("/templates")
async def create_template(request: Request):
    form = await request.form()
    wants_json = "application/json" in request.headers.get("accept", "")
    try:
        require_csrf(request, form.get("csrf_token"))
        name = (form.get("name") or "").strip()
        if not name:
            raise ValueError("Template name is required")
        lines = _parse_lines(form)
        total = round(sum(ln["amount"] for ln in lines), 2)
        if total != 0:
            raise ValueError("Template lines must balance (debits = credits)")
        description = (form.get("description") or "").strip()
        if not description:
            raise ValueError("Description is required")
        reference = (form.get("reference") or "").strip() or None
        payee_id = int(form.get("payee_id")) if form.get("payee_id") else None
        tag_names = _parse_tags(form.get("tags", ""))

        codes = {ln["code"] for ln in lines}
        found = {r["code"] for r in q(
            "SELECT code FROM accounts WHERE code = ANY(%s)", (list(codes),))}
        missing = codes - found
        if missing:
            raise ValueError(f"Unknown account code: {', '.join(sorted(missing))}")

        with tx() as cur:
            cur.execute(
                """INSERT INTO entry_templates (name, description, reference, payee_id)
                   VALUES (%s, %s, %s, %s) RETURNING id""",
                (name, description, reference, payee_id))
            tpl_id = cur.fetchone()["id"]
            for n, ln in enumerate(lines, start=1):
                cur.execute(
                    """INSERT INTO entry_template_lines
                           (template_id, line_no, account_id, amount, memo)
                       VALUES (%s, %s, (SELECT id FROM accounts WHERE code = %s), %s, %s)""",
                    (tpl_id, n, ln["code"], ln["amount"], ln["memo"]))
            if tag_names:
                _sync_tags(cur, "entry_template_tags", "template_id", tpl_id, tag_names)
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        if wants_json:
            return JSONResponse({"ok": False, "error": msg}, status_code=400)
        return flash_redirect("/templates", err=msg)
    ok_msg = f"Template {name!r} saved"
    if wants_json:
        return JSONResponse({"ok": True, "redirect": flash_url("/templates", ok=ok_msg)})
    return flash_redirect("/templates", ok=ok_msg)


@app.post("/templates/{template_id}/delete")
def delete_template(template_id: int, request: Request, csrf_token: str = Form(...)):
    try:
        require_csrf(request, csrf_token)
        with tx() as cur:
            cur.execute("DELETE FROM entry_templates WHERE id = %s", (template_id,))
    except (ValueError, psycopg.Error) as e:
        msg = _pg_msg(e) if isinstance(e, psycopg.Error) else str(e)
        return flash_redirect("/templates", err=msg)
    return flash_redirect("/templates", ok="Template deleted")


# ---------------------------------------------------------------------------
# JSON API — same data, no HTML. Point scripts (or curiosity) here.
# ---------------------------------------------------------------------------
@app.get("/api/trial-balance")
def api_trial_balance(scenario: str = "ACTUAL", as_of: str = None):
    return q("SELECT * FROM fn_trial_balance(%s, %s)", (scenario, as_of))


@app.get("/api/accounts")
def api_accounts():
    return q("SELECT * FROM v_dim_account ORDER BY sort_path")


@app.get("/api/scenarios")
def api_scenarios():
    return scenarios_all()


@app.get("/api/entries")
def api_entries(scenario: str = None, date_from: str = None,
                date_to: str = None):
    where, params = ["TRUE"], []
    if scenario:
        where.append("scenario_code = %s")
        params.append(scenario)
    if date_from:
        where.append("entry_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("entry_date <= %s")
        params.append(date_to)
    return q(f"""SELECT * FROM v_fact_lines WHERE {' AND '.join(where)}
                 ORDER BY entry_date DESC, seq DESC, line_id LIMIT 1000""",
             params)


@app.get("/api/monthly-activity")
def api_monthly(scenario: str = None):
    if scenario:
        return q("""SELECT * FROM v_monthly_activity WHERE scenario_code = %s
                    ORDER BY month, account_code""", (scenario,))
    return q("SELECT * FROM v_monthly_activity ORDER BY month, account_code")


# ---------------------------------------------------------------------------
def _pg_msg(e: Exception) -> str:
    """Surface the RAISE EXCEPTION message from a trigger, without the noise."""
    diag = getattr(e, "diag", None)
    if diag is not None and diag.message_primary:
        return diag.message_primary
    return str(e).splitlines()[0]
